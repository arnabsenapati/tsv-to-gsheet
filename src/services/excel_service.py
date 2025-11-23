"""
Excel workbook operations service.

This module handles all interactions with Excel files:
- Reading TSV files
- Appending rows to Excel workbooks
- Detecting duplicates
- Managing worksheet data
"""

import csv
import re
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from utils.helpers import (
    normalize_magazine_edition,
    normalize_qno,
    normalize_page,
    validate_tsv,
    convert_value_for_column,
)


def read_tsv_rows(tsv_path: Path) -> list[list[str]]:
    """
    Read all rows from a TSV file, skipping the header.
    
    Args:
        tsv_path: Path to the TSV file
        
    Returns:
        List of rows, where each row is a list of column values
        
    Example:
        rows = read_tsv_rows(Path("questions.tsv"))
        # [["1", "25", "JEE Main 2023", "Physics For You Jan 2023"], ...]
    """
    with tsv_path.open("r", encoding="utf-8", newline="") as tsv_file:
        reader = csv.reader(tsv_file, delimiter="\t")
        next(reader, None)  # Skip header row
        return [row for row in reader]


def collect_existing_triplets(
    worksheet, 
    magazine_col: int, 
    qno_col: int, 
    page_col: int
) -> dict[tuple[str, str, str], tuple[str, str, str]]:
    """
    Collect all existing (magazine, qno, page) triplets from the worksheet.
    
    This is used to detect duplicates before appending new rows.
    All values are normalized for consistent comparison.
    
    Args:
        worksheet: openpyxl worksheet object
        magazine_col: Column index for magazine name (1-based)
        qno_col: Column index for question number (1-based)
        page_col: Column index for page number (1-based)
        
    Returns:
        Dictionary mapping normalized triplets to original values:
        {
            ("physics_for_you_jan_2023", "1", "25"): ("Physics For You Jan 2023", "1", "25"),
            ...
        }
    """
    triplets: dict[tuple[str, str, str], tuple[str, str, str]] = {}
    
    # Start from row 2 (skip header row 1)
    for row_idx in range(2, worksheet.max_row + 1):
        magazine_value = worksheet.cell(row=row_idx, column=magazine_col).value
        if not magazine_value:
            continue
        
        qno_value = worksheet.cell(row=row_idx, column=qno_col).value
        page_value = worksheet.cell(row=row_idx, column=page_col).value
        
        # Normalize all values for comparison
        normalized_magazine = normalize_magazine_edition(str(magazine_value))
        normalized_qno = normalize_qno(qno_value)
        normalized_page = normalize_page(page_value)
        
        if normalized_qno and normalized_page:
            key = (normalized_magazine, normalized_qno, normalized_page)
            # Store original values for error messages
            triplets.setdefault(
                key,
                (
                    str(magazine_value),
                    str(qno_value) if qno_value is not None else "",
                    str(page_value) if page_value is not None else "",
                ),
            )
    
    return triplets


def extract_file_metadata(
    rows: list[list[str]],
    magazine_col: int,
    question_set_col: int,
    qno_col: int,
    page_col: int,
) -> tuple[str, list[tuple[str, str, str, str, str, str]]]:
    """
    Extract and validate metadata from TSV rows.
    
    Validates:
    - All rows belong to same magazine edition
    - Required columns are present and non-empty
    - No duplicate questions within the file
    - Values can be normalized
    
    Args:
        rows: List of TSV rows (each row is list of strings)
        magazine_col: Magazine column index (1-based)
        question_set_col: Question set column index (1-based)
        qno_col: Question number column index (1-based)
        page_col: Page number column index (1-based)
        
    Returns:
        Tuple of (magazine_identifier, row_signatures) where:
        - magazine_identifier: Normalized magazine name
        - row_signatures: List of tuples with (normalized_magazine, normalized_qno, 
          normalized_page, original_magazine, original_qno, original_page)
          
    Raises:
        ValueError: If validation fails
    """
    magazine_identifier = None
    row_signatures: list[tuple[str, str, str, str, str, str]] = []
    seen_row_signatures: set[tuple[str, str, str]] = set()
    
    for row in rows:
        # ========================================================================
        # Validate Row Length
        # ========================================================================
        required_columns = max(magazine_col, question_set_col, qno_col, page_col)
        if len(row) < required_columns:
            raise ValueError("TSV row does not contain all required columns.")
        
        # ========================================================================
        # Magazine Edition Validation
        # ========================================================================
        magazine_value = row[magazine_col - 1].strip()
        if not magazine_value:
            raise ValueError("Magazine edition must be provided for every row.")
        
        normalized_magazine = normalize_magazine_edition(magazine_value)
        
        # Ensure all rows belong to same edition
        if magazine_identifier is None:
            magazine_identifier = normalized_magazine
        elif magazine_identifier != normalized_magazine:
            raise ValueError(
                "All rows in the TSV must belong to the same magazine edition. "
                "Please split files by edition before importing."
            )
        
        # ========================================================================
        # Question Set Validation
        # ========================================================================
        question_value = row[question_set_col - 1].strip()
        if not question_value:
            raise ValueError("Question set must be provided for every row.")
        
        # ========================================================================
        # Question Number Validation
        # ========================================================================
        qno_value = row[qno_col - 1].strip()
        if not qno_value:
            raise ValueError("Question number must be provided for every row.")
        
        normalized_qno = normalize_qno(qno_value)
        if not normalized_qno:
            raise ValueError(f"Unable to normalize question number '{qno_value}'.")
        
        # ========================================================================
        # Page Number Validation
        # ========================================================================
        page_value = row[page_col - 1].strip()
        if not page_value:
            raise ValueError("Page number must be provided for every row.")
        
        normalized_page = normalize_page(page_value)
        if not normalized_page:
            raise ValueError(f"Unable to normalize page number '{page_value}'.")
        
        # ========================================================================
        # Duplicate Detection Within File
        # ========================================================================
        combo_signature = (magazine_identifier, normalized_qno, normalized_page)
        if combo_signature in seen_row_signatures:
            raise ValueError(
                "Duplicate question/page detected within TSV for magazine edition "
                f"'{magazine_value}', question number '{qno_value}', page '{page_value}'."
            )
        seen_row_signatures.add(combo_signature)
        
        # Store signatures for further processing
        row_signatures.append(
            (
                normalized_magazine,
                normalized_qno,
                normalized_page,
                magazine_value,
                qno_value,
                page_value,
            )
        )
    
    if magazine_identifier is None:
        raise ValueError("Unable to identify magazine edition in the TSV file.")
    
    return magazine_identifier, row_signatures


def append_rows_to_excel(
    workbook_path: Path,
    worksheet,
    header_row: list[str],
    column_types: dict[int, type],
    rows: list[list[str]],
    insert_row: int,
    page_col: int,
) -> tuple[str, str]:
    """
    Append TSV rows to Excel worksheet and save the workbook.
    
    Features:
    - Converts values to correct types (int, float, str)
    - Tracks page numbers for range calculation
    - Saves workbook after appending
    - Returns status message and page range
    
    Args:
        workbook_path: Path to Excel file
        worksheet: openpyxl worksheet object
        header_row: List of header column names
        column_types: Dict mapping column index to Python type
        rows: List of rows to append (each row is list of values)
        insert_row: Row number where insertion starts (1-based)
        page_col: Column index for page numbers (1-based)
        
    Returns:
        Tuple of (status_message, page_range) where:
        - status_message: "Appended N rows to 'SheetName'"
        - page_range: "25" or "25-30" or "N/A"
        
    Example:
        msg, pages = append_rows_to_excel(
            Path("questions.xlsx"),
            worksheet,
            ["Qno", "Page", "Set", "Magazine"],
            {1: int, 2: int},
            [["1", "25", "JEE Main", "PFY Jan 2023"]],
            100,
            2
        )
        # msg: "Appended 1 rows to 'Sheet1'"
        # pages: "25"
    """
    appended_rows = 0
    page_numbers = []
    
    # ============================================================================
    # Write Rows to Worksheet
    # ============================================================================
    for row in rows:
        # Extract page number for tracking
        if page_col - 1 < len(row):
            page_value = row[page_col - 1].strip()
            if page_value:
                page_numbers.append(page_value)
        
        # Write each column with type conversion
        for col_idx, value in enumerate(row, start=1):
            target_type = column_types.get(col_idx)
            converted = convert_value_for_column(value, target_type, header_row, col_idx)
            worksheet.cell(row=insert_row, column=col_idx, value=converted)
        
        insert_row += 1
        appended_rows += 1
    
    # Save workbook
    worksheet.parent.save(workbook_path)
    
    # ============================================================================
    # Calculate Page Range
    # ============================================================================
    page_range = "N/A"
    if page_numbers:
        try:
            # Extract numeric values from page strings
            numeric_pages = []
            for page in page_numbers:
                # Find first number in page string (handles "25", "p25", "25-26")
                match = re.search(r'\d+', page)
                if match:
                    numeric_pages.append(int(match.group()))
            
            if numeric_pages:
                min_page = min(numeric_pages)
                max_page = max(numeric_pages)
                if min_page == max_page:
                    page_range = str(min_page)
                else:
                    page_range = f"{min_page}-{max_page}"
            else:
                # No numeric pages found - use first and last as-is
                if len(page_numbers) == 1:
                    page_range = page_numbers[0]
                else:
                    page_range = f"{page_numbers[0]}-{page_numbers[-1]}"
        except Exception:
            # Fallback to simple first-last format
            if len(page_numbers) == 1:
                page_range = page_numbers[0]
            else:
                page_range = f"{page_numbers[0]}-{page_numbers[-1]}"
    
    status_msg = f"Appended {appended_rows} rows to '{worksheet.title}'"
    return status_msg, page_range


def infer_column_types(worksheet, num_columns: int) -> dict[int, type]:
    """
    Infer data types for each column by examining first 100 rows.
    
    Args:
        worksheet: openpyxl worksheet object
        num_columns: Number of columns to analyze
        
    Returns:
        Dict mapping column index (1-based) to Python type (int, float, or str)
        
    Example:
        types = infer_column_types(worksheet, 4)
        # {1: int, 2: int, 3: str, 4: str}
    """
    column_types: dict[int, type] = {}
    
    for col_idx in range(1, num_columns + 1):
        sample_values = []
        # Sample first 100 rows (skip header)
        for row_idx in range(2, min(102, worksheet.max_row + 1)):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                sample_values.append(value)
        
        # Determine type from samples
        if not sample_values:
            column_types[col_idx] = str
        elif all(isinstance(v, int) for v in sample_values):
            column_types[col_idx] = int
        elif all(isinstance(v, (int, float)) for v in sample_values):
            column_types[col_idx] = float
        else:
            column_types[col_idx] = str
    
    return column_types


def _find_qno_column(header_row: list[str]) -> int:
    """Find the question number column index (1-based)."""
    for idx, header in enumerate(header_row, 1):
        if header and "qno" in header.lower():
            return idx
    raise ValueError("Could not find 'Qno' column in worksheet header.")


def _find_magazine_column(header_row: list[str]) -> int:
    """Find the magazine column index (1-based)."""
    for idx, header in enumerate(header_row, 1):
        if header and "magazine" in header.lower():
            return idx
    raise ValueError("Could not find 'Magazine' column in worksheet header.")


def _find_question_set_column(header_row: list[str]) -> int:
    """Find the question set column index (1-based)."""
    for idx, header in enumerate(header_row, 1):
        if header and "question set name" in header.lower():
            return idx
    raise ValueError("Could not find 'Question Set Name' column in worksheet header.")


def _find_page_column(header_row: list[str]) -> int:
    """Find the page number column index (1-based)."""
    for idx, header in enumerate(header_row, 1):
        if header and "page" in header.lower():
            return idx
    raise ValueError("Could not find 'Page' column in worksheet header.")


def _find_insert_row(worksheet) -> int:
    """
    Find the next available row for insertion.
    
    Returns the first empty row after existing data.
    """
    return worksheet.max_row + 1


def process_tsv(tsv_path: Path, workbook_path: Path) -> str:
    """
    Process a TSV file and append its contents to an Excel workbook.
    
    This is the main entry point for TSV processing. It:
    1. Validates the TSV file format
    2. Loads the Excel workbook
    3. Checks for duplicates
    4. Appends new rows
    5. Returns status message
    
    Args:
        tsv_path: Path to TSV file to import
        workbook_path: Path to target Excel workbook
        
    Returns:
        Status message: "Appended N rows to 'SheetName', Pages: 25-30"
        
    Raises:
        FileNotFoundError: If TSV or workbook not found
        ValueError: If validation fails or duplicates detected
    """
    try:
        # Validate TSV file format
        validate_tsv(tsv_path)
        rows = read_tsv_rows(tsv_path)
        
        # Load Excel workbook
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        
        workbook = load_workbook(workbook_path)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        
        # Read header row
        header_row = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if not header_row:
            raise ValueError("Worksheet header row is empty.")
        
        # Find required columns
        qno_column = _find_qno_column(header_row)
        magazine_col = _find_magazine_column(header_row)
        question_set_col = _find_question_set_column(header_row)
        page_col = _find_page_column(header_row)
        insert_row = _find_insert_row(worksheet)
        
        # Prepare for insertion
        column_types = infer_column_types(worksheet, len(header_row))
        existing_triplets = collect_existing_triplets(worksheet, magazine_col, qno_column, page_col)
        
        # Validate TSV metadata
        magazine_identifier, row_signatures = extract_file_metadata(
            rows, magazine_col, question_set_col, qno_column, page_col
        )
        
        # Check for duplicates with existing data
        duplicates = []
        for normalized_magazine, normalized_qno, normalized_page, original_mag, original_qno, original_page in row_signatures:
            combo = (normalized_magazine, normalized_qno, normalized_page)
            if combo in existing_triplets:
                existing_mag, existing_qno, existing_page = existing_triplets[combo]
                duplicates.append(
                    (
                        original_mag or existing_mag,
                        original_qno,
                        original_page,
                        existing_qno,
                        existing_page,
                    )
                )
        
        if duplicates:
            readable = "; ".join(
                f"Magazine '{mag}' Question '{qno}' Page '{page}' already exists (Workbook has Qno '{ex_qno}', Page '{ex_page}')"
                for mag, qno, page, ex_qno, ex_page in duplicates
            )
            raise ValueError(
                "Duplicate questions detected: "
                f"{readable}. Remove or update these entries before importing."
            )
        
        # Append rows to workbook
        status_message, page_range = append_rows_to_excel(
            workbook_path=workbook_path,
            worksheet=worksheet,
            header_row=header_row,
            column_types=column_types,
            rows=rows,
            insert_row=insert_row,
            page_col=page_col,
        )
        
        # Add page range to status message
        status_message = f"{status_message}, Pages: {page_range}"
        return status_message
        
    except ValueError as exc:
        raise ValueError(f"TSV processing failed: {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Unexpected error during TSV processing: {exc}") from exc
