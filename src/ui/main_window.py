"""
Main application window for TSV to Excel Watcher.

This module contains the TSVWatcherWindow class which is the main UI window
for the application. It handles:
- Workbook analysis and magazine edition tracking
- Question list management with grouping and tagging
- Chapter grouping and organization
- TSV file monitoring and import
- Custom question list creation
"""

from __future__ import annotations

import datetime as dt
import json
import queue
import re
import threading
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QColor, QFont, QPalette, QTextCursor, QIcon
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpacerItem,
    QSpinBox,
    QSplitter,
    QStackedWidget,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from config.constants import (
    LAST_SELECTION_FILE,
    MAGAZINE_GROUPING_MAP,
    PHYSICS_GROUPING_FILE,
    QUESTION_LIST_DIR,
    TAGS_CONFIG_FILE,
    TAG_COLORS,
)
from services.excel_service import process_tsv
from ui.dialogs import MultiSelectTagDialog
from ui.widgets import (
    ChapterCardView,
    ChapterTableWidget,
    DashboardView,
    GroupingChapterListWidget,
    GroupListWidget,
    NavigationSidebar,
    QuestionCardWidget,
    QuestionCardWithRemoveButton,
    QuestionTreeWidget,
    QuestionListCardView,
    TagBadge,
)
from utils.helpers import (
    _find_high_level_chapter_column,
    _find_magazine_column,
    _find_page_column,
    _find_qno_column,
    _find_question_set_column,
    _find_question_set_name_column,
    _find_question_text_column,
    normalize_magazine_edition,
)


class TSVWatcherWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("TSV to Excel Watcher")
        self.resize(1200, 820)

        self.event_queue: queue.Queue[tuple] = queue.Queue()
        self.watch_thread: threading.Thread | None = None
        self.stop_event = threading.Event()
        self.file_rows: dict[str, int] = {}
        self.file_errors: dict[str, str] = {}
        self.metrics_request_id = 0
        self.chapter_questions: dict[str, list[dict[str, str]]] = {}
        self.current_questions: list[dict[str, str]] = []
        self.all_questions: list[dict[str, str]] = []  # Unfiltered questions
        self.question_set_search_term: str = ""
        self.magazine_search_term: str = ""
        self.tag_filter_term: str = ""
        self.selected_tag_filters: list[str] = []  # Multiple selected tags for filtering
        self.current_magazine_name: str = ""  # Track current magazine for grouping
        self.current_selected_chapter: str | None = None  # Track selected chapter
        self.canonical_chapters: list[str] = []
        self.chapter_lookup: dict[str, str] = {}
        self.chapter_groups: dict[str, list[str]] = {}
        self.current_workbook_path: Path | None = None
        self.workbook_df: pd.DataFrame | None = None  # Cached DataFrame
        self.high_level_column_index: int | None = None
        self.question_lists: dict[str, list[dict]] = {}  # name -> list of questions
        self.question_lists_metadata: dict[str, dict] = {}  # name -> metadata (filters, magazine, etc)
        self.current_list_name: str | None = None
        self.current_selected_chapter: str | None = None  # Track selected chapter for filtering
        self.current_list_questions: list[dict] = []  # Questions in currently selected list
        self.group_tags: dict[str, list[str]] = {}  # group_key -> list of tags
        self.tag_colors: dict[str, str] = {}  # tag -> color
        self.copy_mode: str = "Copy: Text"  # Default copy mode for question cards
        
        # Custom list search variables
        self.list_question_set_search_term: str = ""
        
        # JEE Main Papers analysis
        self.jee_papers_df: pd.DataFrame | None = None
        self.jee_papers_file: Path | None = None
        
        # Predefined color palette for tags (20 colors)
        self.available_tag_colors = [
            "#2563eb",  # Blue
            "#10b981",  # Green
            "#f59e0b",  # Amber
            "#ef4444",  # Red
            "#8b5cf6",  # Purple
            "#06b6d4",  # Cyan
            "#ec4899",  # Pink
            "#14b8a6",  # Teal
            "#f97316",  # Orange
            "#6366f1",  # Indigo
            "#84cc16",  # Lime
            "#f43f5e",  # Rose
            "#0ea5e9",  # Sky Blue
            "#a855f7",  # Violet
            "#22c55e",  # Green Light
            "#eab308",  # Yellow
            "#d946ef",  # Fuchsia
            "#3b82f6",  # Blue Light
            "#fb923c",  # Orange Light
            "#38bdf8",  # Sky
        ]

        # Ensure QuestionList directory exists
        QUESTION_LIST_DIR.mkdir(parents=True, exist_ok=True)
        
        # Status bar animation state
        self.status_animation_frame = 0
        self.status_animation_timer = QTimer()
        self.status_animation_timer.timeout.connect(self._animate_status)
        
        self._load_group_tags()

        self._build_ui()
        self._load_last_selection()
        self._setup_timer()
        # Delay initial load to ensure UI is ready and timer is running
        QTimer.singleShot(100, self.update_row_count)

    def _build_ui(self) -> None:
        self._apply_palette()
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(10, 10, 10, 10)
        root_layout.setSpacing(10)

        # Create horizontal splitter for sidebar + content
        main_splitter = QSplitter(Qt.Horizontal)
        root_layout.addWidget(main_splitter, 1)

        # Left: Navigation sidebar
        self.sidebar = NavigationSidebar()
        self.sidebar.navigation_changed.connect(self._on_navigation_changed)
        main_splitter.addWidget(self.sidebar)

        # Right: Stacked widget for content pages
        self.content_stack = QStackedWidget()
        main_splitter.addWidget(self.content_stack)

        # Set splitter sizes (sidebar: 200px, content: remaining)
        main_splitter.setSizes([200, 1000])
        main_splitter.setCollapsible(0, False)  # Don't allow sidebar to collapse to 0

        # Create all view pages
        self._create_dashboard_page()      # Index 0
        self._create_magazine_page()        # Index 1
        self._create_questions_page()       # Index 2
        self._create_grouping_page()        # Index 3
        self._create_lists_page()           # Index 4
        self._create_import_page()          # Index 5
        self._create_jee_page()             # Index 6

        # Status bar and log toggle row
        status_log_layout = QHBoxLayout()
        
        # Status label with animation
        self.status_label = QLabel("")
        self.status_label.setStyleSheet("""
            QLabel {
                padding: 8px 16px;
                background-color: #f1f5f9;
                border-radius: 6px;
                color: #475569;
                font-size: 13px;
            }
        """)
        self.status_label.setVisible(True)
        self.status_label.setWordWrap(False)
        self.status_label.setMinimumWidth(200)
        status_log_layout.addWidget(self.status_label, 1)
        
        status_log_layout.addSpacing(10)
        
        self.log_toggle = QPushButton("Show Log")
        self.log_toggle.setCheckable(True)
        self.log_toggle.toggled.connect(self.toggle_log_visibility)
        status_log_layout.addWidget(self.log_toggle)
        
        root_layout.addLayout(status_log_layout)

        # Log card (collapsible)
        self.log_card = self._create_card()
        log_layout = QVBoxLayout(self.log_card)
        log_layout.addWidget(self._create_label("Log"))
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setObjectName("logView")
        log_layout.addWidget(self.log_view)
        root_layout.addWidget(self.log_card, 0)
        self.log_card.setVisible(False)
        
        self._refresh_grouping_ui()
        self._load_saved_question_lists()

    def _on_navigation_changed(self, index: int):
        """Handle navigation sidebar item selection."""
        self.content_stack.setCurrentIndex(index)

    def _create_dashboard_page(self):
        """Create Dashboard page (index 0)."""
        dashboard_container = QWidget()
        dashboard_layout = QVBoxLayout(dashboard_container)
        dashboard_layout.setContentsMargins(20, 20, 20, 20)
        dashboard_layout.setSpacing(20)
        
        # Workbook selector card
        workbook_card = self._create_card()
        workbook_layout = QVBoxLayout(workbook_card)
        workbook_layout.setSpacing(12)
        
        self.output_edit = QLineEdit()
        self.output_edit.editingFinished.connect(self.update_row_count)
        output_row = QHBoxLayout()
        output_row.addWidget(self._create_label("Workbook"))
        output_row.addWidget(self.output_edit)
        browse_output = QPushButton("Browse…")
        browse_output.clicked.connect(self.select_output_file)
        output_row.addWidget(browse_output)
        workbook_layout.addLayout(output_row)
        
        # Create compatibility labels (hidden, used for logging/status updates)
        self.row_count_label = QLabel("Total rows: N/A")
        self.row_count_label.setVisible(False)
        self.mag_summary_label = QLabel("Magazines: N/A")
        self.mag_summary_label.setVisible(False)
        self.mag_missing_label = QLabel("Missing ranges: N/A")
        self.mag_missing_label.setVisible(False)
        
        dashboard_layout.addWidget(workbook_card)
        
        # Add the new DashboardView for statistics
        self.dashboard_view = DashboardView(self)
        dashboard_layout.addWidget(self.dashboard_view, 1)
        
        self.content_stack.addWidget(dashboard_container)

    def _create_magazine_page(self):
        """Create Magazine Editions page (index 1)."""
        magazine_page = QWidget()
        magazine_tab_layout = QVBoxLayout(magazine_page)
        
        # Top summary card
        mag_summary_card = self._create_card()
        mag_summary_layout = QHBoxLayout(mag_summary_card)
        self.mag_total_editions_label = QLabel("Total Editions: 0")
        self.mag_total_editions_label.setObjectName("headerLabel")
        self.mag_total_sets_label = QLabel("Question Sets: 0")
        self.mag_total_sets_label.setObjectName("infoLabel")
        mag_summary_layout.addWidget(self.mag_total_editions_label)
        mag_summary_layout.addStretch()
        mag_summary_layout.addWidget(self.mag_total_sets_label)
        magazine_tab_layout.addWidget(mag_summary_card)
        
        mag_split = QSplitter(Qt.Horizontal)
        magazine_tab_layout.addWidget(mag_split, 1)
        
        # Left side - Magazine tree with search
        mag_tree_card = self._create_card()
        mag_tree_layout = QVBoxLayout(mag_tree_card)
        mag_tree_layout.addWidget(self._create_label("Magazine Editions"))
        
        # Search box for magazine tree
        mag_search_layout = QHBoxLayout()
        mag_search_layout.addWidget(QLabel("Search:"))
        self.mag_tree_search = QLineEdit()
        self.mag_tree_search.setPlaceholderText("Filter magazines or editions...")
        self.mag_tree_search.textChanged.connect(self.on_mag_tree_search_changed)
        mag_search_layout.addWidget(self.mag_tree_search)
        clear_mag_search_btn = QPushButton("Clear")
        clear_mag_search_btn.setMaximumWidth(80)
        clear_mag_search_btn.clicked.connect(lambda: self.mag_tree_search.clear())
        mag_search_layout.addWidget(clear_mag_search_btn)
        mag_tree_layout.addLayout(mag_search_layout)
        
        # Use QTreeWidget for editions with parent-child structure
        self.mag_tree = QTreeWidget()
        self.mag_tree.setColumnCount(2)
        self.mag_tree.setHeaderLabels(["Edition", "Sets"])
        self.mag_tree.setRootIsDecorated(True)
        self.mag_tree.setAlternatingRowColors(True)
        self.mag_tree.itemSelectionChanged.connect(self.on_magazine_select)
        self.mag_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.mag_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        mag_tree_layout.addWidget(self.mag_tree)
        mag_split.addWidget(mag_tree_card)

        # Right side - Question sets detail with tree
        detail_card = self._create_card()
        detail_layout = QVBoxLayout(detail_card)
        detail_layout.addWidget(self._create_label("Question Sets"))
        self.question_label = QLabel("Select an edition to view question sets")
        self.question_label.setObjectName("infoLabel")
        self.question_label.setStyleSheet(
            "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
        )
        detail_layout.addWidget(self.question_label)
        
        # Use QSplitter to add question text view
        mag_question_splitter = QSplitter(Qt.Vertical)
        
        # Use QTreeWidget instead of QListWidget for expandable question sets
        self.question_sets_tree = QTreeWidget()
        self.question_sets_tree.setHeaderLabels(["Question Set / Question", "Qno", "Page"])
        self.question_sets_tree.setAlternatingRowColors(True)
        self.question_sets_tree.setColumnWidth(0, 400)
        self.question_sets_tree.itemSelectionChanged.connect(self.on_mag_question_selected)
        self.question_sets_tree.setStyleSheet(
            """
            QTreeWidget {
                font-size: 13px;
            }
            QTreeWidget::item {
                padding: 6px;
                border-bottom: 1px solid #e2e8f0;
            }
            QTreeWidget::item:hover {
                background-color: #f8fafc;
            }
            """
        )
        mag_question_splitter.addWidget(self.question_sets_tree)
        
        # Question text view for magazine editions
        self.mag_question_text_view = QTextEdit()
        self.mag_question_text_view.setReadOnly(True)
        self.mag_question_text_view.setAcceptRichText(True)
        mag_question_splitter.addWidget(self.mag_question_text_view)
        
        mag_question_splitter.setStretchFactor(0, 3)
        mag_question_splitter.setStretchFactor(1, 1)
        mag_question_splitter.setSizes([400, 120])
        
        detail_layout.addWidget(mag_question_splitter)
        mag_split.addWidget(detail_card)
        
        mag_split.setSizes([500, 400])
        
        self.content_stack.addWidget(magazine_page)

    def _create_questions_page(self):
        """Create Question List page (index 2)."""
        questions_page = QWidget()
        questions_tab_layout = QVBoxLayout(questions_page)
        analysis_card = self._create_card()
        questions_tab_layout.addWidget(analysis_card)
        analysis_layout = QVBoxLayout(analysis_card)
        analysis_split = QSplitter(Qt.Horizontal)
        analysis_layout.addWidget(analysis_split, 1)

        chapter_card = self._create_card()
        chapter_layout = QVBoxLayout(chapter_card)
        chapter_layout.addWidget(self._create_label("Chapters"))
        self.chapter_view = ChapterCardView(self)
        self.chapter_view.chapter_selected.connect(self.on_chapter_selected)
        chapter_layout.addWidget(self.chapter_view)
        analysis_split.addWidget(chapter_card)

        question_card = self._create_card()
        question_layout = QVBoxLayout(question_card)
        
        # Search and action controls in single row
        search_layout = QHBoxLayout()
        search_layout.setSpacing(8)
        
        # Search controls container with visual grouping
        search_container = QWidget()
        search_container.setStyleSheet("""
            QWidget {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 6px;
            }
            QLabel {
                color: #1e40af;
                font-weight: 600;
                background: transparent;
            }
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 4px;
                padding: 4px 8px;
                color: #0f172a;
            }
            QLineEdit:focus {
                border: 1px solid #3b82f6;
            }
        """)
        search_container_layout = QHBoxLayout(search_container)
        search_container_layout.setContentsMargins(8, 4, 8, 4)
        search_container_layout.setSpacing(8)
        
        # Question Set search
        self.question_set_search = QLineEdit()
        self.question_set_search.setPlaceholderText("Search by question set name")
        self.question_set_search.setToolTip("Search by question set name")
        self.question_set_search.setMinimumHeight(28)
        self.question_set_search.setMaximumHeight(28)
        self.question_set_search.textChanged.connect(self.on_question_set_search_changed)
        search_container_layout.addWidget(self.question_set_search, 2)  # Stretch factor 2
        
        # Tags filter
        self.tag_filter_display = QLineEdit()
        self.tag_filter_display.setPlaceholderText("No tags selected")
        self.tag_filter_display.setToolTip("Selected tags for filtering")
        self.tag_filter_display.setReadOnly(True)
        self.tag_filter_display.setMinimumHeight(28)
        self.tag_filter_display.setMaximumHeight(28)
        self.tag_filter_display.setStyleSheet("""
            QLineEdit {
                background-color: #0f172a;
                color: #ffffff;
                border: 1px solid #334155;
                padding: 4px 8px;
                border-radius: 4px;
            }
        """)
        search_container_layout.addWidget(self.tag_filter_display, 2)  # Stretch factor 2
        
        # Tag filter button (same style as accordion tag button)
        self.tag_filter_btn = QPushButton("🏷️")
        self.tag_filter_btn.setToolTip("Select tags to filter questions")
        self.tag_filter_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: 1px solid #3b82f6;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 14px;
                min-width: 30px;
                min-height: 28px;
                max-height: 28px;
            }
            QPushButton:hover {
                background-color: #3b82f6;
            }
        """)
        self.tag_filter_btn.clicked.connect(self._show_tag_filter_dialog)
        search_container_layout.addWidget(self.tag_filter_btn)
        
        # Magazine search
        self.magazine_search = QLineEdit()
        self.magazine_search.setPlaceholderText("Search by magazine name")
        self.magazine_search.setToolTip("Search by magazine name")
        self.magazine_search.setMinimumHeight(28)
        self.magazine_search.setMaximumHeight(28)
        self.magazine_search.textChanged.connect(self.on_magazine_search_changed)
        search_container_layout.addWidget(self.magazine_search, 2)  # Stretch factor 2
        
        # Clear button with icon
        clear_search_btn = QPushButton("✕")
        clear_search_btn.setToolTip("Clear all search filters")
        clear_search_btn.setMinimumHeight(28)
        clear_search_btn.setMaximumHeight(28)
        clear_search_btn.setMinimumWidth(32)
        clear_search_btn.setMaximumWidth(32)
        clear_search_btn.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        clear_search_btn.clicked.connect(self.clear_question_search)
        search_container_layout.addWidget(clear_search_btn)
        
        search_layout.addWidget(search_container, 1)  # Stretch to fill space
        
        # Action buttons container with different visual style
        action_container = QWidget()
        action_container.setStyleSheet("""
            QWidget {
                background-color: #dbeafe;
                border: 1px solid #93c5fd;
                border-radius: 6px;
                padding: 6px;
            }
        """)
        action_layout = QHBoxLayout(action_container)
        action_layout.setContentsMargins(8, 4, 8, 4)
        action_layout.setSpacing(8)
        
        # Action buttons with icons
        add_to_list_btn = QPushButton()
        add_to_list_btn.setIcon(QIcon("icons/add.svg"))
        add_to_list_btn.setToolTip("Add selected questions to list")
        add_to_list_btn.setMinimumHeight(28)
        add_to_list_btn.setMaximumHeight(28)
        add_to_list_btn.setMinimumWidth(32)
        add_to_list_btn.setMaximumWidth(32)
        add_to_list_btn.clicked.connect(self.add_selected_to_list)
        action_layout.addWidget(add_to_list_btn)
        
        create_random_list_btn = QPushButton()
        create_random_list_btn.setIcon(QIcon("icons/random.png"))
        create_random_list_btn.setToolTip("Create random list from filtered questions")
        create_random_list_btn.setMinimumHeight(28)
        create_random_list_btn.setMaximumHeight(28)
        create_random_list_btn.setMinimumWidth(32)
        create_random_list_btn.setMaximumWidth(32)
        create_random_list_btn.clicked.connect(self.create_random_list_from_filtered)
        action_layout.addWidget(create_random_list_btn)
        
        # Copy mode selector
        action_layout.addSpacing(8)  # Visual separator
        self.copy_mode_combo = QComboBox()
        self.copy_mode_combo.addItems(["Copy: Text", "Copy: Metadata", "Copy: Both"])
        self.copy_mode_combo.setCurrentIndex(0)  # Default to "Copy: Text"
        self.copy_mode_combo.setToolTip("Select what to copy when double-clicking a question:\n"
                                        "• Text: Only the question content\n"
                                        "• Metadata: Only question number, page, chapter\n"
                                        "• Both: Question text + metadata")
        self.copy_mode_combo.setMinimumHeight(28)
        self.copy_mode_combo.setMaximumHeight(28)
        self.copy_mode_combo.setStyleSheet("""
            QComboBox {
                background-color: #ffffff;
                border: 1px solid #3b82f6;
                border-radius: 4px;
                padding: 4px 8px;
                color: #1e40af;
                font-weight: 600;
            }
            QComboBox:hover {
                background-color: #eff6ff;
            }
            QComboBox::drop-down {
                border: none;
                background: transparent;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                color: #1e40af;
                selection-background-color: #3b82f6;
                selection-color: #ffffff;
            }
        """)
        self.copy_mode_combo.currentTextChanged.connect(self._on_copy_mode_changed)
        action_layout.addWidget(self.copy_mode_combo)
        
        search_layout.addWidget(action_container)
        
        question_layout.addLayout(search_layout)
        
        # Drag-drop panel (initially hidden)
        from src.ui.widgets import DragDropQuestionPanel
        self.drag_drop_panel = DragDropQuestionPanel(self.question_lists, self)
        self.drag_drop_panel.save_clicked.connect(self._on_drag_drop_save)
        self.drag_drop_panel.cancel_clicked.connect(self._on_drag_drop_cancel)
        self.drag_drop_panel.setVisible(False)
        question_layout.addWidget(self.drag_drop_panel)
        
        # Replace traditional tree with card-based view
        self.question_card_view = QuestionListCardView(self)
        self.question_card_view.question_selected.connect(self.on_question_card_selected)
        
        # Enable keyboard focus for Ctrl detection
        self.question_card_view.setFocusPolicy(Qt.StrongFocus)
        
        # Keep old tree widget reference for compatibility (hidden)
        self.question_tree = QuestionTreeWidget(self)
        self.question_tree.setVisible(False)
        self.question_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.question_tree.customContextMenuRequested.connect(self._show_group_context_menu)
        
        # Add card view directly without the text viewer splitter
        question_layout.addWidget(self.question_card_view)
        
        # Create hidden text view for compatibility with other code
        self.question_text_view = QTextEdit()
        self.question_text_view.setVisible(False)
        self.question_text_view.setReadOnly(True)
        self.question_text_view.setAcceptRichText(True)
        
        analysis_split.addWidget(question_card)
        
        self.content_stack.addWidget(questions_page)

    def _create_grouping_page(self):
        """Create Chapter Grouping page (index 3)."""
        grouping_page = QWidget()
        grouping_layout = QVBoxLayout(grouping_page)
        grouping_card = self._create_card()
        grouping_layout.addWidget(grouping_card)
        grouping_card_layout = QHBoxLayout(grouping_card)

        self.group_list = GroupListWidget(self)
        self.group_list.itemSelectionChanged.connect(self.on_group_selected)
        grouping_card_layout.addWidget(self.group_list, 1)

        self.group_chapter_list = GroupingChapterListWidget(self)
        grouping_card_layout.addWidget(self.group_chapter_list, 2)

        group_controls = QVBoxLayout()
        group_controls.addWidget(self._create_label("Move chapter to group"))
        self.move_target_combo = QComboBox()
        group_controls.addWidget(self.move_target_combo)
        move_button = QPushButton("Move Chapter")
        move_button.clicked.connect(self.move_selected_chapter)
        group_controls.addWidget(move_button)
        group_controls.addStretch()
        grouping_card_layout.addLayout(group_controls)
        
        self.content_stack.addWidget(grouping_page)

    def _create_lists_page(self):
        """Create Custom Lists page (index 4)."""
        lists_page = QWidget()
        lists_tab_layout = QVBoxLayout(lists_page)
        lists_card = self._create_card()
        lists_tab_layout.addWidget(lists_card)
        lists_card_layout = QVBoxLayout(lists_card)
        
        lists_header_layout = QHBoxLayout()
        lists_card_layout.addWidget(self._create_label("Custom Lists"))
        lists_header_layout.addStretch()
        lists_card_layout.addLayout(lists_header_layout)
        
        lists_split = QSplitter(Qt.Horizontal)
        lists_card_layout.addWidget(lists_split, 1)
        
        # Left side - list of saved lists
        saved_lists_card = self._create_card()
        saved_lists_layout = QVBoxLayout(saved_lists_card)
        saved_lists_layout.addWidget(self._create_label("Saved Lists"))
        
        list_controls_layout = QHBoxLayout()
        new_list_btn = QPushButton("New List")
        new_list_btn.clicked.connect(self.create_new_question_list)
        rename_list_btn = QPushButton("Rename")
        rename_list_btn.clicked.connect(self.rename_question_list)
        delete_list_btn = QPushButton("Delete")
        delete_list_btn.clicked.connect(self.delete_question_list)
        list_controls_layout.addWidget(new_list_btn)
        list_controls_layout.addWidget(rename_list_btn)
        list_controls_layout.addWidget(delete_list_btn)
        saved_lists_layout.addLayout(list_controls_layout)
        
        self.saved_lists_widget = QListWidget()
        self.saved_lists_widget.itemSelectionChanged.connect(self.on_saved_list_selected)
        saved_lists_layout.addWidget(self.saved_lists_widget)
        lists_split.addWidget(saved_lists_card)
        
        # Right side - questions in selected list
        list_questions_card = self._create_card()
        list_questions_layout = QVBoxLayout(list_questions_card)
        self.list_name_label = QLabel("Select a list to view questions")
        self.list_name_label.setObjectName("headerLabel")
        list_questions_layout.addWidget(self.list_name_label)
        
        # Label to show filters applied to this list
        self.list_filters_label = QLabel("")
        self.list_filters_label.setObjectName("infoLabel")
        self.list_filters_label.setWordWrap(True)
        self.list_filters_label.setStyleSheet(
            "padding: 6px; background-color: #fef3c7; border-radius: 4px; color: #92400e; font-size: 12px;"
        )
        self.list_filters_label.setVisible(False)
        list_questions_layout.addWidget(self.list_filters_label)
        
        # Search and action controls in single row (similar to Question List tab)
        search_layout = QHBoxLayout()
        search_layout.setSpacing(8)
        
        # Search controls container with visual grouping
        search_container = QWidget()
        search_container.setStyleSheet("""
            QWidget {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 6px;
            }
            QLabel {
                color: #1e40af;
                font-weight: 600;
                background: transparent;
            }
            QLineEdit {
                background-color: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 4px;
                padding: 4px 8px;
                color: #0f172a;
            }
            QLineEdit:focus {
                border: 1px solid #3b82f6;
            }
        """)
        search_container_layout = QHBoxLayout(search_container)
        search_container_layout.setContentsMargins(8, 4, 8, 4)
        search_container_layout.setSpacing(8)
        
        # Question Set search
        self.list_question_set_search = QLineEdit()
        self.list_question_set_search.setPlaceholderText("Search by question set name")
        self.list_question_set_search.setToolTip("Search by question set name")
        self.list_question_set_search.setMinimumHeight(28)
        self.list_question_set_search.setMaximumHeight(28)
        self.list_question_set_search.textChanged.connect(self.on_list_question_set_search_changed)
        search_container_layout.addWidget(self.list_question_set_search, 1)
        
        # Clear button
        clear_search_btn = QPushButton("✕")
        clear_search_btn.setToolTip("Clear search filter")
        clear_search_btn.setMinimumHeight(28)
        clear_search_btn.setMaximumHeight(28)
        clear_search_btn.setMinimumWidth(32)
        clear_search_btn.setMaximumWidth(32)
        clear_search_btn.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        clear_search_btn.clicked.connect(self.clear_list_search)
        search_container_layout.addWidget(clear_search_btn)
        
        search_layout.addWidget(search_container, 1)
        
        # Action buttons container
        action_container = QWidget()
        action_container.setStyleSheet("""
            QWidget {
                background-color: #dbeafe;
                border: 1px solid #93c5fd;
                border-radius: 6px;
                padding: 6px;
            }
        """)
        action_layout = QHBoxLayout(action_container)
        action_layout.setContentsMargins(8, 4, 8, 4)
        action_layout.setSpacing(8)
        
        # Copy mode selector for custom lists
        self.list_copy_mode_combo = QComboBox()
        self.list_copy_mode_combo.addItems(["Copy: Text", "Copy: Metadata", "Copy: Both"])
        self.list_copy_mode_combo.setCurrentIndex(0)
        self.list_copy_mode_combo.setToolTip("Select what to copy when double-clicking a question:\n"
                                             "• Text: Only the question content\n"
                                             "• Metadata: Only question number, page, chapter\n"
                                             "• Both: Question text + metadata")
        self.list_copy_mode_combo.setMinimumHeight(28)
        self.list_copy_mode_combo.setMaximumHeight(28)
        self.list_copy_mode_combo.setStyleSheet("""
            QComboBox {
                background-color: #ffffff;
                border: 1px solid #3b82f6;
                border-radius: 4px;
                padding: 4px 8px;
                color: #1e40af;
                font-weight: 600;
            }
            QComboBox:hover {
                background-color: #eff6ff;
            }
            QComboBox::drop-down {
                border: none;
                background: transparent;
            }
            QComboBox QAbstractItemView {
                background-color: #ffffff;
                color: #1e40af;
                selection-background-color: #3b82f6;
                selection-color: #ffffff;
            }
        """)
        self.list_copy_mode_combo.currentTextChanged.connect(self._on_list_copy_mode_changed)
        action_layout.addWidget(self.list_copy_mode_combo)
        
        action_layout.addStretch()
        search_layout.addWidget(action_container)
        
        list_questions_layout.addLayout(search_layout)
        
        # Create a simple scroll area for 2-column card grid (not accordion)
        self.list_question_card_view = QScrollArea()
        self.list_question_card_view.setWidgetResizable(True)
        self.list_question_card_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_question_card_view.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.list_question_card_view.setStyleSheet("""
            QScrollArea {
                border: 1px solid #e2e8f0;
                background-color: #ffffff;
            }
        """)
        list_questions_layout.addWidget(self.list_question_card_view)
        
        lists_split.addWidget(list_questions_card)

        self.content_stack.addWidget(lists_page)

    def _create_import_page(self):
        """Create Data Import page (index 5)."""
        import_page = QWidget()
        import_layout = QVBoxLayout(import_page)

        import_form_card = self._create_card()
        import_form_layout = QVBoxLayout(import_form_card)
        self.input_edit = QLineEdit()
        self.input_edit.editingFinished.connect(self.refresh_file_list)
        input_row = QHBoxLayout()
        input_row.addWidget(self._create_label("Input folder"))
        input_row.addWidget(self.input_edit)
        browse_input = QPushButton("Browse…")
        browse_input.clicked.connect(self.select_input_folder)
        input_row.addWidget(browse_input)
        import_form_layout.addLayout(input_row)

        control_layout = QHBoxLayout()
        self.refresh_button = QPushButton("Refresh Files")
        self.refresh_button.clicked.connect(self.refresh_file_list)
        self.start_button = QPushButton("Start Watching")
        self.start_button.clicked.connect(self.start_watching)
        self.stop_button = QPushButton("Stop Watching")
        self.stop_button.clicked.connect(self.stop_watching)
        control_layout.addWidget(self.refresh_button)
        control_layout.addWidget(self.start_button)
        control_layout.addWidget(self.stop_button)
        control_layout.addStretch()
        import_form_layout.addLayout(control_layout)
        import_layout.addWidget(import_form_card)

        status_card = self._create_card()
        status_layout = QVBoxLayout(status_card)
        status_layout.addWidget(self._create_label("Upload Status"))
        self.file_table = QTableWidget(0, 3)
        self.file_table.setHorizontalHeaderLabels(["File", "Status", "Message"])
        self.file_table.horizontalHeader().setStretchLastSection(True)
        self.file_table.verticalHeader().setVisible(False)
        self.file_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.file_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.file_table.cellDoubleClicked.connect(self.on_file_double_clicked)
        status_layout.addWidget(self.file_table)
        import_layout.addWidget(status_card)

        self.content_stack.addWidget(import_page)

    def _create_jee_page(self):
        """Create JEE Main Papers page (index 6)."""
        jee_page = QWidget()
        jee_layout = QVBoxLayout(jee_page)
        jee_layout.setSpacing(8)
        jee_layout.setContentsMargins(10, 10, 10, 10)
        
        # Compact controls in single row
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(10)
        
        # File selection
        file_label = QLabel("File:")
        file_label.setMaximumWidth(35)
        controls_layout.addWidget(file_label)
        
        self.jee_file_edit = QLineEdit()
        self.jee_file_edit.setReadOnly(True)
        self.jee_file_edit.setMaximumHeight(26)
        controls_layout.addWidget(self.jee_file_edit, 1)  # Stretch factor 1
        
        browse_jee_btn = QPushButton("Browse…")
        browse_jee_btn.setMaximumHeight(26)
        browse_jee_btn.setMaximumWidth(75)
        browse_jee_btn.clicked.connect(self.select_jee_papers_file)
        controls_layout.addWidget(browse_jee_btn)
        
        # Spacer
        controls_layout.addSpacing(20)
        
        # Subject dropdown
        subject_label = QLabel("Subject:")
        subject_label.setMaximumWidth(50)
        controls_layout.addWidget(subject_label)
        
        self.jee_subject_combo = QComboBox()
        self.jee_subject_combo.setMinimumWidth(180)
        self.jee_subject_combo.setMaximumWidth(250)
        self.jee_subject_combo.setMaximumHeight(26)
        self.jee_subject_combo.currentTextChanged.connect(self.update_jee_tables)
        controls_layout.addWidget(self.jee_subject_combo)
        
        jee_layout.addLayout(controls_layout)
        
        # Split view: Chapters table on left, Questions table on right
        jee_splitter = QSplitter(Qt.Horizontal)
        
        # Left side - Chapters table
        chapters_widget = QWidget()
        chapters_layout = QVBoxLayout(chapters_widget)
        chapters_layout.setContentsMargins(0, 0, 0, 0)
        chapters_layout.setSpacing(4)
        
        chapters_header = QLabel("Chapters (by Question Count)")
        chapters_header.setStyleSheet("font-weight: bold; padding: 4px;")
        chapters_layout.addWidget(chapters_header)
        
        self.jee_chapters_table = QTableWidget(0, 2)
        self.jee_chapters_table.setHorizontalHeaderLabels(["Chapter", "Count"])
        self.jee_chapters_table.horizontalHeader().setStretchLastSection(False)
        self.jee_chapters_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.jee_chapters_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.jee_chapters_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.jee_chapters_table.setSelectionMode(QTableWidget.SingleSelection)
        self.jee_chapters_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.jee_chapters_table.setAlternatingRowColors(True)
        self.jee_chapters_table.itemSelectionChanged.connect(self.on_jee_chapter_selected)
        chapters_layout.addWidget(self.jee_chapters_table)
        
        jee_splitter.addWidget(chapters_widget)
        
        # Right side - Questions table
        questions_widget = QWidget()
        questions_layout = QVBoxLayout(questions_widget)
        questions_layout.setContentsMargins(0, 0, 0, 0)
        questions_layout.setSpacing(4)
        
        self.jee_questions_label = QLabel("Select a chapter to view questions")
        self.jee_questions_label.setStyleSheet("font-weight: bold; padding: 4px;")
        questions_layout.addWidget(self.jee_questions_label)
        
        self.jee_questions_table = QTableWidget(0, 0)
        self.jee_questions_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.jee_questions_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.jee_questions_table.setAlternatingRowColors(True)
        self.jee_questions_table.horizontalHeader().setStretchLastSection(True)
        questions_layout.addWidget(self.jee_questions_table)
        
        jee_splitter.addWidget(questions_widget)
        
        jee_splitter.setStretchFactor(0, 1)
        jee_splitter.setStretchFactor(1, 2)
        jee_splitter.setSizes([300, 600])
        
        jee_layout.addWidget(jee_splitter, 1)
        
        self.content_stack.addWidget(jee_page)

    def _apply_palette(self) -> None:
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor("#f4f5f9"))
        palette.setColor(QPalette.Base, QColor("#ffffff"))
        palette.setColor(QPalette.Text, QColor("#0f172a"))
        palette.setColor(QPalette.Button, QColor("#2563eb"))
        palette.setColor(QPalette.ButtonText, QColor("#ffffff"))
        self.setPalette(palette)
        self.setStyleSheet(
            """
            QWidget#card {
                background-color: #ffffff;
                border-radius: 14px;
                padding: 12px;
            }
            QPushButton {
                background-color: #2563eb;
                color: #ffffff;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
            }
            QPushButton:disabled {
                background-color: #94a3b8;
            }
            QLabel#headerLabel {
                font-weight: 600;
                color: #0f172a;
            }
            QLabel#infoLabel {
                color: #475569;
            }
            QTreeWidget, QTableWidget, QListWidget {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                background-color: #ffffff;
                color: #0f172a;
                alternate-background-color: #f8fafc;
            }
            QTreeWidget::item:selected, QTableWidget::item:selected, QListWidget::item:selected {
                background-color: #d0e2ff;
                color: #0f172a;
            }
            QTextEdit#logView {
                background-color: #0f172a;
                color: #e2e8f0;
                border-radius: 8px;
                padding: 8px;
                font-family: Consolas, 'Courier New', monospace;
            }
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                background-color: #ffffff;
            }
            QTabBar::tab {
                background-color: #f1f5f9;
                color: #475569;
                border: 1px solid #e2e8f0;
                border-bottom: none;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background-color: #ffffff;
                color: #0f172a;
                border-bottom: 2px solid #2563eb;
                font-weight: 600;
            }
            QTabBar::tab:hover:!selected {
                background-color: #e0e7ff;
                color: #1e40af;
            }
            """
        )

    def _create_card(self) -> QWidget:
        card = QWidget()
        card.setObjectName("card")
        return card

    def _create_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName("headerLabel")
        return label

    def toggle_log_visibility(self, checked: bool) -> None:
        self.log_card.setVisible(checked)
        self.log_toggle.setText("Hide Log" if checked else "Show Log")
    
    def _animate_status(self) -> None:
        """Animate the status spinner icon."""
        spinner_frames = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
        self.status_animation_frame = (self.status_animation_frame + 1) % len(spinner_frames)
        
        # Update the current status text with new spinner frame
        if hasattr(self, "_current_status_text") and hasattr(self, "_current_status_type"):
            if self._current_status_type in ("loading", "importing"):
                icon = spinner_frames[self.status_animation_frame]
                self.status_label.setText(f"{icon} {self._current_status_text}")
    
    def set_status(self, message: str, status_type: str = "info") -> None:
        """Set status bar message with appropriate icon and styling.
        
        Args:
            message: Status message to display
            status_type: Type of status - 'loading', 'importing', 'success', 'error', 'info'
        """
        self._current_status_text = message
        self._current_status_type = status_type
        
        # Stop any existing animation
        self.status_animation_timer.stop()
        
        if status_type == "loading":
            # Animated spinner for loading
            self.status_animation_timer.start(80)  # Update every 80ms
            icon = "⠋"
            self.status_label.setStyleSheet("""
                QLabel {
                    padding: 8px 16px;
                    background-color: #dbeafe;
                    border-radius: 6px;
                    color: #1e40af;
                    font-size: 13px;
                    font-weight: 500;
                }
            """)
        elif status_type == "importing":
            # Animated spinner for importing
            self.status_animation_timer.start(80)
            icon = "⠋"
            self.status_label.setStyleSheet("""
                QLabel {
                    padding: 8px 16px;
                    background-color: #fef3c7;
                    border-radius: 6px;
                    color: #92400e;
                    font-size: 13px;
                    font-weight: 500;
                }
            """)
        elif status_type == "success":
            icon = "✓"
            self.status_label.setStyleSheet("""
                QLabel {
                    padding: 8px 16px;
                    background-color: #d1fae5;
                    border-radius: 6px;
                    color: #065f46;
                    font-size: 13px;
                    font-weight: 500;
                }
            """)
        elif status_type == "error":
            icon = "✗"
            self.status_label.setStyleSheet("""
                QLabel {
                    padding: 8px 16px;
                    background-color: #fee2e2;
                    border-radius: 6px;
                    color: #991b1b;
                    font-size: 13px;
                    font-weight: 500;
                }
            """)
        else:  # info
            icon = "ℹ"
            self.status_label.setStyleSheet("""
                QLabel {
                    padding: 8px 16px;
                    background-color: #f1f5f9;
                    border-radius: 6px;
                    color: #475569;
                    font-size: 13px;
                    font-weight: 500;
                }
            """)
        
        self.status_label.setText(f"{icon} {message}")
        self.status_label.setVisible(True)
    
    def clear_status(self) -> None:
        """Clear status message but keep label visible to maintain layout."""
        self.status_animation_timer.stop()
        self.status_label.setText("")  # Clear text but keep label visible
        self._current_status_text = ""
        self._current_status_type = ""

    def _load_last_selection(self) -> None:
        if not LAST_SELECTION_FILE.exists():
            return
        try:
            data = json.loads(LAST_SELECTION_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return
        workbook_path = data.get("workbook_path", "")
        if workbook_path:
            self.output_edit.setText(workbook_path)
        input_folder = data.get("input_folder", "")
        if input_folder:
            self.input_edit.setText(input_folder)
            self._auto_start_watching()  # Auto-start watching if both paths are valid
        jee_papers_file = data.get("jee_papers_file", "")
        if jee_papers_file and Path(jee_papers_file).exists():
            self.jee_papers_file = Path(jee_papers_file)
            self.jee_file_edit.setText(jee_papers_file)
            self.load_jee_papers_data()

    def _save_last_selection(self) -> None:
        path_text = self.output_edit.text().strip()
        input_folder = self.input_edit.text().strip()
        if not path_text:
            return
        payload = {
            "workbook_path": path_text,
            "input_folder": input_folder if input_folder else "",
            "jee_papers_file": str(self.jee_papers_file) if self.jee_papers_file else "",
        }
        LAST_SELECTION_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _load_group_tags(self) -> None:
        """Load group tags from tags.cfg file."""
        if not TAGS_CONFIG_FILE.exists():
            return
        try:
            data = json.loads(TAGS_CONFIG_FILE.read_text(encoding="utf-8"))
            self.group_tags = data.get("group_tags", {})
            self.tag_colors = data.get("tag_colors", {})
        except json.JSONDecodeError:
            self.group_tags = {}
            self.tag_colors = {}

    def _save_group_tags(self) -> None:
        """Save group tags to tags.cfg file."""
        payload = {
            "group_tags": self.group_tags,
            "tag_colors": self.tag_colors,
        }
        TAGS_CONFIG_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _get_or_assign_tag_color(self, tag: str) -> str:
        """Get existing color for tag or assign a new one."""
        if tag not in self.tag_colors:
            # Cycle through available colors
            color_index = len(self.tag_colors) % len(self.available_tag_colors)
            self.tag_colors[tag] = self.available_tag_colors[color_index]
        return self.tag_colors[tag]

    def _load_canonical_chapters(self, grouping_file: Path) -> list[str]:
        """Load canonical chapters from the grouping JSON file."""
        if not grouping_file.exists():
            return []
        try:
            data = json.loads(grouping_file.read_text(encoding="utf-8"))
            return data.get("canonical_order", [])
        except json.JSONDecodeError:
            return []

    def _load_chapter_grouping(self, grouping_file: Path) -> dict[str, list[str]]:
        """Load chapter grouping from the specified JSON file."""
        if grouping_file.exists():
            try:
                data = json.loads(grouping_file.read_text(encoding="utf-8"))
                groups = data.get("groups", {})
            except json.JSONDecodeError:
                groups = {}
        else:
            groups = {}
        for group in self.canonical_chapters:
            groups.setdefault(group, [])
        groups.setdefault("Others", [])
        # Remove duplicates while preserving order
        for key, values in list(groups.items()):
            seen = set()
            unique = []
            for value in values:
                if value not in seen:
                    unique.append(value)
                    seen.add(value)
            groups[key] = unique
        self._rebuild_chapter_lookup(groups)
        return groups

    def _reload_grouping_for_magazine(self, magazine_name: str) -> None:
        """Reload chapter grouping based on magazine name."""
        if magazine_name == self.current_magazine_name:
            return  # Already loaded
        
        grouping_file = MAGAZINE_GROUPING_MAP.get(magazine_name, PHYSICS_GROUPING_FILE)
        self.current_magazine_name = magazine_name
        self.canonical_chapters = self._load_canonical_chapters(grouping_file)
        self.chapter_groups = self._load_chapter_grouping(grouping_file)
        self.log(f"Loaded chapter grouping for: {magazine_name or 'default (Physics)'}")

    def _save_chapter_grouping(self) -> None:
        """Save chapter grouping to the appropriate file based on current magazine."""
        data = {
            "canonical_order": self.canonical_chapters,
            "groups": self.chapter_groups,
        }
        grouping_file = MAGAZINE_GROUPING_MAP.get(self.current_magazine_name, PHYSICS_GROUPING_FILE)
        grouping_file.write_text(json.dumps(data, indent=2), encoding="utf-8")
        self._rebuild_chapter_lookup(self.chapter_groups)

    def _ordered_groups(self) -> list[str]:
        order = list(self.canonical_chapters)
        if "Others" not in order:
            order.append("Others")
        return order

    def _rebuild_chapter_lookup(self, groups: dict[str, list[str]]) -> None:
        lookup: dict[str, str] = {}
        for group, values in groups.items():
            norm_group = self._normalize_label(group)
            if norm_group:
                lookup[norm_group] = group
            for value in values:
                norm_value = self._normalize_label(value)
                if norm_value and norm_value not in lookup:
                    lookup[norm_value] = group
        self.chapter_lookup = lookup

    def _refresh_grouping_ui(self) -> None:
        if not hasattr(self, "group_list"):
            return
        self.group_list.blockSignals(True)
        self.group_list.clear()
        for group in self._ordered_groups():
            chapters = self.chapter_groups.get(group, [])
            item = QListWidgetItem(f"{group} ({len(chapters)})")
            item.setData(Qt.UserRole, group)
            self.group_list.addItem(item)
        self.group_list.blockSignals(False)
        if hasattr(self, "move_target_combo"):
            self.move_target_combo.clear()
            self.move_target_combo.addItems(self._ordered_groups())
        if hasattr(self, "group_chapter_list"):
            self.group_chapter_list.clear()

    def _select_group_in_ui(self, group: str) -> None:
        if not hasattr(self, "group_list"):
            return
        for row in range(self.group_list.count()):
            item = self.group_list.item(row)
            if (item.data(Qt.UserRole) or item.text()).startswith(group):
                self.group_list.setCurrentRow(row)
                break

    def _setup_timer(self) -> None:
        self.queue_timer = QTimer(self)
        self.queue_timer.setInterval(200)
        self.queue_timer.timeout.connect(self._process_queue)
        self.queue_timer.start()

    def update_row_count(self) -> None:
        workbook_path = Path(self.output_edit.text().strip())
        if not workbook_path.is_file():
            self._clear_all_question_data()
            self.row_count_label.setText("Total rows: N/A")
            self._set_magazine_summary("Magazines: N/A", "Missing ranges: N/A")
            self.question_label.setText("Select a workbook to display magazine editions.")
            self.current_workbook_path = None
            self.high_level_column_index = None
            if hasattr(self, 'dashboard_view'):
                return

        # Clear all existing data before loading new workbook
        self._clear_all_question_data()
        
        self.current_workbook_path = workbook_path
        self._save_last_selection()
        self.set_status(f'Loading workbook "{workbook_path.name}"', "loading")
        self.row_count_label.setText("Total rows: Loading...")
        self._set_magazine_summary("Magazines: Loading...", "Tracked editions: Loading...")
        self.question_label.setText("Loading editions...")
        self.metrics_request_id += 1
        request_id = self.metrics_request_id

        def worker(path: Path, req_id: int) -> None:
            try:
                # Pandas is used here to keep the UI responsive while gathering workbook metrics.
                df = pd.read_excel(path, sheet_name=0, dtype=object)
                
                # Cache the DataFrame for reuse throughout the application
                self.workbook_df = df
                
                row_count = self._compute_row_count_from_df(df)
                magazine_details, warnings = self._collect_magazine_details(df)
                
                # Detect magazine and load appropriate grouping BEFORE analyzing questions
                detected_magazine = self._detect_magazine_name(magazine_details)
                if detected_magazine != self.current_magazine_name:
                    grouping_file = MAGAZINE_GROUPING_MAP.get(detected_magazine, PHYSICS_GROUPING_FILE)
                    self.current_magazine_name = detected_magazine
                    self.canonical_chapters = self._load_canonical_chapters(grouping_file)
                    self.chapter_groups = self._load_chapter_grouping(grouping_file)
                
                # Now analyze questions with proper grouping loaded
                chapter_data, qa_warnings, question_col, raw_chapter_inputs = self._collect_question_analysis_data(df)
                warnings.extend(qa_warnings)
                self.event_queue.put(
                    (
                        "metrics",
                        req_id,
                        row_count,
                        magazine_details,
                        warnings,
                        chapter_data,
                        question_col,
                        raw_chapter_inputs,
                        detected_magazine,  # Pass magazine name to queue for logging
                    )
                )
            except Exception as exc:
                self.event_queue.put(("metrics_error", req_id, str(exc)))
                self.event_queue.put(("status_error", str(exc)))

        threading.Thread(target=worker, args=(workbook_path, request_id), daemon=True).start()

    def _set_magazine_summary(self, primary: str, secondary: str) -> None:
        self.mag_summary_label.setText(primary)
        self.mag_missing_label.setText(secondary)

    def _invalidate_workbook_cache(self) -> None:
        """Invalidate the workbook cache and reload data."""
        self.workbook_df = None
        self.update_row_count()
    
    def _clear_all_question_data(self) -> None:
        """Clear all question analysis data and UI elements."""
        # Clear data structures
        self.workbook_df = None  # Clear cached DataFrame
        self.chapter_questions.clear()
        self.current_questions.clear()
        self.all_questions.clear()
        self.question_set_search_term = ""
        self.magazine_search_term = ""
        
        # Clear UI elements
        self._populate_magazine_tree([])
        self._populate_question_sets([])
        self._populate_chapter_list({})
        
        # Clear search boxes
        if hasattr(self, "question_set_search"):
            self.question_set_search.clear()
        if hasattr(self, "magazine_search"):
            self.magazine_search.clear()
        if hasattr(self, "mag_tree_search"):
            self.mag_tree_search.clear()

    def _detect_magazine_name(self, details: list[dict]) -> str:
        """Detect magazine name from magazine details."""
        if not details:
            return ""
        # Get the first magazine's display name and normalize it
        display_name = details[0].get("display_name", "")
        normalized = self._normalize_label(display_name)
        # Check against known magazines
        for magazine_key in MAGAZINE_GROUPING_MAP.keys():
            if magazine_key in normalized:
                return magazine_key
        return ""

    def _collect_magazine_details(self, df: pd.DataFrame) -> tuple[list[dict], list[str]]:
        warnings: list[str] = []
        if df.empty:
            return [], warnings

        header_row = [None if pd.isna(col) else str(col) for col in df.columns]
        try:
            magazine_col = _find_magazine_column(header_row)
        except ValueError:
            warnings.append("Unable to determine magazine column for summary display.")
            return [], warnings

        question_set_col = None
        try:
            question_set_col = _find_question_set_column(header_row)
        except ValueError:
            warnings.append("Unable to determine question set column; question sets will not be listed.")

        coverage: dict[str, dict[str, object]] = {}
        magazine_series = df.iloc[:, magazine_col - 1]
        question_series = (
            df.iloc[:, question_set_col - 1] if question_set_col is not None else repeat(None, len(df))
        )
        for magazine_value, question_value in zip(magazine_series, question_series):
            if pd.isna(magazine_value):
                continue
            text = str(magazine_value).strip()
            if not text:
                continue
            display_parts = [part.strip() for part in text.split("|", 1)]
            display_name = display_parts[0] or "Unknown"
            display_edition = display_parts[1] if len(display_parts) > 1 else ""
            normalized = normalize_magazine_edition(text)
            norm_parts = normalized.split("|", 1)
            norm_name = norm_parts[0]
            norm_edition = norm_parts[1] if len(norm_parts) > 1 else ""
            entry = coverage.setdefault(
                norm_name,
                {
                    "display_name": display_name,
                    "editions": {},
                    "normalized_editions": set(),
                },
            )
            edition_label = display_edition or "(unspecified)"
            edition_entry = entry["editions"].setdefault(
                norm_edition,
                {
                    "display": edition_label,
                    "question_sets": set(),
                },
            )
            if question_set_col is not None and not pd.isna(question_value):
                q_text = str(question_value).strip()
                if q_text:
                    edition_entry["question_sets"].add(q_text)
            entry["normalized_editions"].add(norm_edition)

        if not coverage:
            return [], warnings

        details: list[dict] = []
        for norm_name in sorted(coverage.keys(), key=lambda key: str(coverage[key]["display_name"]).lower()):
            data = coverage[norm_name]
            edition_items = []
            for norm_edition, edition_data in sorted(
                data["editions"].items(),
                key=lambda item: self._edition_sort_key(item[0], item[1]["display"]),
            ):
                question_sets = sorted(edition_data["question_sets"], key=lambda value: value.lower())
                edition_items.append(
                    {
                        "display": edition_data["display"],
                        "normalized": norm_edition,
                        "question_sets": question_sets,
                    }
                )
            missing_ranges = self._compute_missing_ranges(data["normalized_editions"])
            details.append(
                {
                    "display_name": data["display_name"],
                    "missing_ranges": missing_ranges,
                    "editions": edition_items,
                }
            )
        return details, warnings

    def _collect_question_analysis_data(
        self, df: pd.DataFrame
    ) -> tuple[dict[str, list[dict]], list[str], int | None, list[str]]:
        warnings: list[str] = []
        if df.empty:
            return {}, warnings, None, []

        header_row = [None if pd.isna(col) else str(col) for col in df.columns]
        try:
            question_set_col = _find_high_level_chapter_column(header_row)
            qno_col = _find_qno_column(header_row)
            page_col = _find_page_column(header_row)
        except ValueError as exc:
            warnings.append(f"Question analysis unavailable: {exc}")
            return {}, warnings, None, []

        question_text_col = None
        try:
            question_text_col = _find_question_text_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        question_set_name_col = None
        try:
            question_set_name_col = _find_question_set_name_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        magazine_col = None
        try:
            magazine_col = _find_magazine_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        def normalize(value) -> str:
            if pd.isna(value):
                return ""
            text = str(value).strip()
            return text

        chapters: dict[str, list[dict]] = {}
        raw_inputs: set[str] = set()
        for row_number, row in enumerate(df.itertuples(index=False, name=None), start=2):
            values = list(row)
            raw_chapter_name = normalize(values[question_set_col - 1])
            if not raw_chapter_name:
                continue
            raw_inputs.add(raw_chapter_name)
            chapter_name = self._match_chapter_group(raw_chapter_name)
            qno_value = normalize(values[qno_col - 1])
            page_value = normalize(values[page_col - 1])
            question_text = normalize(values[question_text_col - 1]) if question_text_col else ""
            question_set_name = normalize(values[question_set_name_col - 1]) if question_set_name_col else raw_chapter_name
            magazine_value = normalize(values[magazine_col - 1]) if magazine_col else ""
            
            # Extract group_key for tag lookup (same key used for accordion headers)
            group_key = self._extract_group_key(question_set_name)

            chapters.setdefault(chapter_name, []).append(
                {
                    "group": chapter_name,
                    "question_set": raw_chapter_name,
                    "question_set_name": question_set_name,
                    "group_key": group_key,  # For tag color lookup in chips
                    "qno": qno_value,
                    "page": page_value,
                    "magazine": magazine_value,
                    "text": question_text,
                    "row_number": row_number,
                }
            )

        return chapters, warnings, question_set_col, sorted(raw_inputs)

    def _normalize_label(self, label: str) -> str:
        return re.sub(r"\s+", " ", label.strip().lower())
    
    def _extract_group_key(self, question_set_name: str) -> str:
        """Extract a group key from question set name for similarity grouping.
        
        Methodology:
        1. Normalize the name (lowercase, remove extra spaces)
        2. Extract significant tokens (ignore common words, years, numbers at end)
        3. Return first 2-3 significant words as group key
        
        Examples:
        - 'JEE Main 2023 Paper 1' -> 'jee main'
        - 'NEET-2024' -> 'neet'
        - 'Physics Olympiad 2023' -> 'physics olympiad'
        """
        if not question_set_name:
            return "ungrouped"
        
        # Normalize
        normalized = self._normalize_label(question_set_name)
        
        # Split into tokens
        tokens = normalized.split()
        
        # Remove common suffix patterns (years, paper numbers, etc.)
        significant_tokens = []
        for token in tokens:
            # Skip years (4 digits), paper numbers, common words
            if re.match(r'^(\d{4}|\d+|paper|set|part|section|test|exam)$', token):
                continue
            significant_tokens.append(token)
            # Take first 2-3 significant words
            if len(significant_tokens) >= 3:
                break
        
        if not significant_tokens:
            # If all tokens were filtered, use first 2 tokens
            return ' '.join(tokens[:2]) if len(tokens) >= 2 else normalized
        
        return ' '.join(significant_tokens)

    def _auto_assign_chapters(self, chapters: list[str]) -> None:
        changed = False
        existing = {
            self._normalize_label(ch)
            for values in self.chapter_groups.values()
            for ch in values
        }
        for chapter in chapters:
            normalized = self._normalize_label(chapter)
            if not normalized or normalized in existing:
                continue
            target = self._match_chapter_group(chapter)
            self.chapter_groups.setdefault(target, []).append(chapter)
            existing.add(normalized)
            changed = True
        if changed:
            self._save_chapter_grouping()
            self._refresh_grouping_ui()

    def _match_chapter_group(self, chapter: str) -> str:
        normalized = self._normalize_label(chapter)
        if not normalized:
            return "Others"
        direct = self.chapter_lookup.get(normalized)
        if direct:
            return direct
        for norm_value, group in self.chapter_lookup.items():
            if norm_value in normalized or normalized in norm_value:
                return group
        for group in self.canonical_chapters:
            norm_group = self._normalize_label(group)
            if normalized == norm_group:
                return group
        for group in self.canonical_chapters:
            norm_group = self._normalize_label(group)
            if norm_group in normalized or normalized in norm_group:
                return group
        return "Others"

    def _compute_row_count_from_df(self, df: pd.DataFrame) -> int:
        def row_has_value(row) -> bool:
            for value in row:
                if isinstance(value, str):
                    if value.strip():
                        return True
                    continue
                if pd.isna(value):
                    continue
                return True
            return False

        for idx in range(len(df) - 1, -1, -1):
            if row_has_value(df.iloc[idx]):
                return idx + 2  # DataFrame row index is zero-based, Excel rows start at 2 after header.
        return 1

    def _compute_missing_ranges(self, normalized_editions: set[str]) -> list[str]:
        monthly_tokens = sorted(
            {
                token
                for token in normalized_editions
                if re.fullmatch(r"\d{4}-\d{2}", token)
            }
        )
        if len(monthly_tokens) < 2:
            return []

        months = [dt.date(int(token[:4]), int(token[5:7]), 1) for token in monthly_tokens]
        present_keys = set(monthly_tokens)
        missing_ranges: list[str] = []
        current = months[0]
        last = months[-1]
        missing_start: dt.date | None = None

        while current <= last:
            key = current.strftime("%Y-%m")
            if key not in present_keys:
                if missing_start is None:
                    missing_start = current
            else:
                if missing_start is not None:
                    end = self._previous_month(current)
                    missing_ranges.append(self._format_range(missing_start, end))
                    missing_start = None
            current = self._add_month(current)

        if missing_start is not None:
            missing_ranges.append(self._format_range(missing_start, last))
        return missing_ranges

    def _add_month(self, date_value: dt.date) -> dt.date:
        if date_value.month == 12:
            return dt.date(date_value.year + 1, 1, 1)
        return dt.date(date_value.year, date_value.month + 1, 1)

    def _previous_month(self, date_value: dt.date) -> dt.date:
        if date_value.month == 1:
            return dt.date(date_value.year - 1, 12, 1)
        return dt.date(date_value.year, date_value.month - 1, 1)

    def _format_range(self, start: dt.date, end: dt.date) -> str:
        if start == end:
            return start.strftime("%b %Y")
        return f"{start.strftime('%b %Y')} - {end.strftime('%b %Y')}"

    def _parse_normalized_month(self, normalized: str) -> dt.date | None:
        if normalized and re.fullmatch(r"\d{4}-\d{2}", normalized):
            return dt.date(int(normalized[:4]), int(normalized[5:7]), 1)
        return None

    def _edition_sort_key(self, normalized: str, display_label: str) -> tuple:
        parsed = self._parse_normalized_month(normalized)
        if parsed:
            return (0, -parsed.toordinal(), display_label.lower())
        return (1, display_label.lower(), "")

    def _edition_sort_key(self, normalized: str, display_label: str) -> tuple:
        parsed = self._parse_normalized_month(normalized)
        if parsed:
            return (0, -parsed.toordinal(), display_label.lower())
        return (1, display_label.lower(), "")

    def _populate_magazine_tree(self, details: list[dict]) -> None:
        """Populate magazine editions tree grouped by year with counts."""
        if not hasattr(self, "mag_tree"):
            return
        self.mag_tree.clear()
        
        # Calculate statistics
        total_editions = 0
        total_sets = 0
        
        if not details:
            if hasattr(self, "mag_total_editions_label"):
                self.mag_total_editions_label.setText("Total Editions: 0")
            if hasattr(self, "mag_total_sets_label"):
                self.mag_total_sets_label.setText("Question Sets: 0")
            return

        for entry in details:
            magazine_name = entry["display_name"]
            
            # Create parent node for magazine
            magazine_parent = QTreeWidgetItem([magazine_name, ""])
            magazine_parent.setData(0, Qt.UserRole, {"type": "magazine", "display_name": magazine_name})
            font = QFont()
            font.setBold(True)
            magazine_parent.setFont(0, font)
            self.mag_tree.addTopLevelItem(magazine_parent)
            
            # Group editions by year
            editions_by_year = {}
            missing_by_year = {}
            
            # Process existing editions
            for edition in entry.get("editions", []):
                edition_label = edition["display"] or "(unspecified)"
                question_sets = edition["question_sets"]
                normalized = edition.get("normalized", "")
                
                parsed_date = self._parse_normalized_month(normalized)
                if parsed_date:
                    year = parsed_date.year
                    if year not in editions_by_year:
                        editions_by_year[year] = []
                    editions_by_year[year].append({
                        "date": parsed_date,
                        "label": edition_label,
                        "sets": question_sets,
                        "normalized": normalized
                    })
            
            # Process missing editions
            missing_ranges = entry.get("missing_ranges", [])
            if missing_ranges:
                missing_editions = self._expand_missing_ranges(missing_ranges)
                for missing_date in missing_editions:
                    year = missing_date.year
                    if year not in missing_by_year:
                        missing_by_year[year] = []
                    missing_by_year[year].append(missing_date)
            
            # Get all years and sort in descending order
            all_years = sorted(set(list(editions_by_year.keys()) + list(missing_by_year.keys())), reverse=True)
            
            # Create year nodes
            for year in all_years:
                year_editions = editions_by_year.get(year, [])
                year_missing = missing_by_year.get(year, [])
                year_count = len(year_editions)
                
                # Create year parent node with count
                year_parent = QTreeWidgetItem([f"{year} ({year_count})", ""])
                year_parent.setData(0, Qt.UserRole, {"type": "year", "year": year})
                font_year = QFont()
                font_year.setBold(True)
                year_parent.setFont(0, font_year)
                
                # Year-based background color
                bg_color = QColor(self._get_year_color(year))
                year_parent.setBackground(0, bg_color)
                year_parent.setBackground(1, bg_color)
                
                magazine_parent.addChild(year_parent)
                
                # Add existing editions for this year
                for edition_data in sorted(year_editions, key=lambda x: x["date"].toordinal(), reverse=True):
                    formatted_label = edition_data["date"].strftime("%b '%y")
                    sets_count = len(edition_data["sets"])
                    
                    child = QTreeWidgetItem([formatted_label, str(sets_count)])
                    child.setBackground(0, bg_color)
                    child.setBackground(1, bg_color)
                    
                    data = {
                        "type": "edition",
                        "display_name": magazine_name,
                        "edition_label": edition_data["label"],
                        "question_sets": edition_data["sets"],
                        "is_missing": False
                    }
                    child.setData(0, Qt.UserRole, data)
                    year_parent.addChild(child)
                    
                    total_editions += 1
                    total_sets += sets_count
                
                # Add missing editions for this year
                if year_missing:
                    missing_parent = QTreeWidgetItem(["❌ Missing Editions", ""])
                    missing_parent.setForeground(0, QColor("#dc2626"))
                    font_bold = QFont()
                    font_bold.setBold(True)
                    missing_parent.setFont(0, font_bold)
                    missing_parent.setData(0, Qt.UserRole, {"type": "missing_section"})
                    year_parent.addChild(missing_parent)
                    
                    for missing_date in sorted(year_missing, key=lambda d: d.toordinal(), reverse=True):
                        formatted_label = missing_date.strftime("%b '%y")
                        missing_child = QTreeWidgetItem([formatted_label, "-"])
                        
                        # RED BOLD for missing editions
                        font_red_bold = QFont()
                        font_red_bold.setBold(True)
                        missing_child.setFont(0, font_red_bold)
                        missing_child.setFont(1, font_red_bold)
                        missing_child.setForeground(0, QColor("#dc2626"))
                        missing_child.setForeground(1, QColor("#dc2626"))
                        
                        missing_child.setBackground(0, bg_color)
                        missing_child.setBackground(1, bg_color)
                        
                        missing_child.setData(0, Qt.UserRole, {"type": "missing", "is_missing": True})
                        missing_parent.addChild(missing_child)
                
                year_parent.setExpanded(True)
            
            magazine_parent.setExpanded(True)
        
        # Update summary statistics
        if hasattr(self, "mag_total_editions_label"):
            self.mag_total_editions_label.setText(f"Total Editions: {total_editions}")
        if hasattr(self, "mag_total_sets_label"):
            self.mag_total_sets_label.setText(f"Question Sets: {total_sets}")
    
    def _expand_missing_ranges(self, missing_ranges: list[str]) -> list[dt.date]:
        """Expand missing ranges into individual dates."""
        missing_dates = []
        
        for range_str in missing_ranges:
            if " - " in range_str:
                # Range like "Jan 2023 - Mar 2023"
                start_str, end_str = range_str.split(" - ")
                start_date = dt.datetime.strptime(start_str, "%b %Y").date()
                end_date = dt.datetime.strptime(end_str, "%b %Y").date()
                
                current = start_date
                while current <= end_date:
                    missing_dates.append(current)
                    current = self._add_month(current)
            else:
                # Single month like "Jan 2023"
                missing_dates.append(dt.datetime.strptime(range_str, "%b %Y").date())
        
        return missing_dates
    
    def _get_year_color(self, year: int) -> str:
        """Get a consistent color for a given year."""
        year_colors = {
            2020: "#fef3c7",  # Light amber
            2021: "#dbeafe",  # Light blue
            2022: "#d1fae5",  # Light green
            2023: "#fce7f3",  # Light pink
            2024: "#e0e7ff",  # Light indigo
            2025: "#fed7aa",  # Light orange
            2026: "#e9d5ff",  # Light purple
            2027: "#ccfbf1",  # Light teal
            2028: "#fef08a",  # Light yellow
            2029: "#fbcfe8",  # Light rose
        }
        # For years not in the map, generate a color based on year
        if year not in year_colors:
            # Generate pastel colors using year as seed
            hue = (year * 137) % 360
            return f"hsl({hue}, 70%, 90%)"
        return year_colors[year]

    def _populate_question_sets(self, question_sets: list[str] | None, magazine_name: str = "", edition_label: str = "") -> None:
        """Populate question sets tree with questions as children."""
        if not hasattr(self, "question_sets_tree"):
            return
        self.question_sets_tree.clear()
        if not question_sets or self.workbook_df is None:
            return
        
        # Use cached DataFrame for much better performance
        try:
            df = self.workbook_df
            self.log(f"Using cached DataFrame ({len(df)} rows) to populate question sets")
            header_row = [None if pd.isna(col) else str(col) for col in df.columns]
            
            # Find required columns
            try:
                qno_col = _find_qno_column(header_row)
                page_col = _find_page_column(header_row)
                question_set_col = _find_question_set_column(header_row)
                magazine_col = _find_magazine_column(header_row)
            except ValueError:
                # If columns not found, just show question sets without children
                for name in question_sets:
                    parent_item = QTreeWidgetItem([f"📝 {name}", "", ""])
                    self.question_sets_tree.addTopLevelItem(parent_item)
                return
            
            # Collect questions by question set for this specific magazine edition
            questions_by_set: dict[str, list[dict]] = {}
            normalized_target = normalize_magazine_edition(f"{magazine_name}|{edition_label}")
            
            # Use pandas vectorized operations for better performance
            magazine_series = df.iloc[:, magazine_col - 1]
            question_set_series = df.iloc[:, question_set_col - 1]
            qno_series = df.iloc[:, qno_col - 1]
            page_series = df.iloc[:, page_col - 1]
            
            for idx, (mag_value, qset_value, qno_value, page_value) in enumerate(zip(
                magazine_series, question_set_series, qno_series, page_series
            )):
                # Skip if magazine value is missing
                if pd.isna(mag_value):
                    continue
                
                # Check if magazine matches
                normalized_mag = normalize_magazine_edition(str(mag_value))
                if normalized_mag != normalized_target:
                    continue
                
                # Get question set
                if pd.isna(qset_value):
                    continue
                    
                qset_name = str(qset_value).strip()
                if qset_name not in question_sets:
                    continue
                
                # Get question details
                qno_str = str(qno_value).strip() if not pd.isna(qno_value) else ""
                page_str = str(page_value).strip() if not pd.isna(page_value) else ""
                
                questions_by_set.setdefault(qset_name, []).append({
                    "qno": qno_str,
                    "page": page_str,
                    "row": idx + 2  # DataFrame is 0-based, Excel rows start at 2
                })
            
            # Create tree items
            for qset_name in question_sets:
                questions = questions_by_set.get(qset_name, [])
                parent_item = QTreeWidgetItem([f"📝 {qset_name}", "", f"({len(questions)} questions)"])
                parent_item.setData(0, Qt.UserRole, {"type": "question_set", "name": qset_name})
                
                # Add questions as children
                for q in questions:
                    child_item = QTreeWidgetItem([f"  Q{q['qno']}", q['qno'], q['page']])
                    child_item.setData(0, Qt.UserRole, {"type": "question", "qno": q['qno'], "page": q['page'], "row": q['row']})
                    parent_item.addChild(child_item)
                
                self.question_sets_tree.addTopLevelItem(parent_item)
                
        except Exception as e:
            self.log(f"Error loading questions for question sets: {e}")
            # Fallback to simple list
            for name in question_sets:
                parent_item = QTreeWidgetItem([f"📝 {name}", "", ""])
                self.question_sets_tree.addTopLevelItem(parent_item)

    def _populate_chapter_list(self, chapters: dict[str, list[dict]]) -> None:
        if not hasattr(self, "chapter_view"):
            return
        self.chapter_questions = chapters or {}
        if hasattr(self, "question_tree"):
            self.question_tree.clear()
        self.question_text_view.clear()
        if not self.chapter_questions:
            self.chapter_view.clear_chapters()
            return
        
        # Sort chapters by question count (descending) then by name (ascending)
        sorted_chapters = sorted(
            self.chapter_questions.items(),
            key=lambda kv: (-len(kv[1]), kv[0].lower()),
        )
        
        # Clear and populate chapter view
        self.chapter_view.clear_chapters()
        
        for chapter_key, questions in sorted_chapters:
            self.chapter_view.add_chapter(
                chapter_name=chapter_key,
                chapter_key=chapter_key,
                question_count=len(questions)
            )
        
        # Select first chapter
        if sorted_chapters:
            first_chapter_key = sorted_chapters[0][0]
            self.chapter_view._on_card_clicked(first_chapter_key)

    def _populate_question_table(self, questions: list[dict]) -> None:
        if not hasattr(self, "question_tree"):
            return
        self.all_questions = questions or []
        self._apply_question_search()

    def _apply_question_search(self, preserve_scroll: bool = False) -> None:
        """Apply normalized search terms and update the question card view."""
        # Save scroll position if requested
        scroll_value = None
        if preserve_scroll and hasattr(self, "question_card_view"):
            scrollbar = self.question_card_view.verticalScrollBar()
            if scrollbar:
                scroll_value = scrollbar.value()
        
        filtered = self.all_questions
        
        # Apply Question Set search (normalized)
        if self.question_set_search_term:
            normalized_search = self._normalize_label(self.question_set_search_term)
            filtered = [
                q for q in filtered
                if normalized_search in self._normalize_label(q.get("question_set_name", ""))
            ]
        
        # Apply Magazine search (normalized)
        if self.magazine_search_term:
            normalized_search = self._normalize_label(self.magazine_search_term)
            filtered = [
                q for q in filtered
                if normalized_search in self._normalize_label(q.get("magazine", ""))
            ]
        
        # Update card view with grouping
        self.current_questions = filtered
        
        # Clear both views
        if hasattr(self, "question_card_view"):
            self.question_card_view.clear()
        if hasattr(self, "question_tree"):
            self.question_tree.clear()
        self.question_text_view.clear()
        
        if not filtered:
            return
        
        # Group questions by similar question set names
        groups = {}
        for question in filtered:
            question_set_name = question.get("question_set_name", "Unknown")
            group_key = self._extract_group_key(question_set_name)
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append(question)
        
        # Apply tag filtering (multiple tags)
        if self.selected_tag_filters:
            filtered_groups = {}
            for group_key, group_questions in groups.items():
                tags = self.group_tags.get(group_key, [])
                # Check if any of the selected filter tags match any of the group's tags
                if any(filter_tag in tags for filter_tag in self.selected_tag_filters):
                    filtered_groups[group_key] = group_questions
            groups = filtered_groups
        
        # Sort groups by number of questions (descending) then by name
        sorted_groups = sorted(groups.items(), key=lambda x: (-len(x[1]), x[0]))
        
        # Populate card view with accordion groups
        if hasattr(self, "question_card_view"):
            for group_key, group_questions in sorted_groups:
                tags = self.group_tags.get(group_key, [])
                self.question_card_view.add_group(group_key, group_questions, tags, self.tag_colors)
        
        # Restore scroll position if it was saved
        if scroll_value is not None and hasattr(self, "question_card_view"):
            scrollbar = self.question_card_view.verticalScrollBar()
            if scrollbar:
                scrollbar.setValue(scroll_value)

    def on_question_set_search_changed(self, text: str) -> None:
        """Handle Question Set search box text change."""
        self.question_set_search_term = text.strip()
        self._apply_question_search()

    def on_magazine_search_changed(self, text: str) -> None:
        """Handle Magazine search box text change."""
        self.magazine_search_term = text.strip()
        self._apply_question_search()

    def clear_question_search(self) -> None:
        """Clear all search terms."""
        self.question_set_search_term = ""
        self.magazine_search_term = ""
        self.selected_tag_filters = []
        if hasattr(self, "question_set_search"):
            self.question_set_search.clear()
        if hasattr(self, "tag_filter_display"):
            self.tag_filter_display.clear()
        if hasattr(self, "magazine_search"):
            self.magazine_search.clear()
        self._apply_question_search()

    def _show_tag_filter_dialog(self) -> None:
        """Show dialog to select multiple tags for filtering."""
        # Get all existing tags across all groups
        all_tags = sorted(set(tag for tags in self.group_tags.values() for tag in tags))
        
        if not all_tags:
            QMessageBox.information(self, "No Tags", "No tags have been created yet. Assign tags to groups first.")
            return
        
        # Show multi-select dialog
        dialog = MultiSelectTagDialog(
            all_tags, 
            self.selected_tag_filters, 
            "Filter by Tags",
            self.tag_colors,
            self.available_tag_colors,
            self
        )
        
        if dialog.exec() == QDialog.Accepted:
            self.selected_tag_filters = dialog.get_selected_tags()
            # Update label
            if self.selected_tag_filters:
                self.tag_filter_display.setText(", ".join(self.selected_tag_filters))
            else:
                self.tag_filter_display.clear()
            self._apply_question_search()

    def on_magazine_search_changed(self, text: str) -> None:
        """Handle Magazine search box text change."""
        self.magazine_search_term = text.strip()
        self._apply_question_search()

    def on_mag_tree_search_changed(self, text: str) -> None:
        """Filter magazine tree based on search text (supports 3-level hierarchy)."""
        search_term = text.strip().lower()
        
        if not hasattr(self, "mag_tree"):
            return
        
        # Show all if search is empty
        if not search_term:
            for i in range(self.mag_tree.topLevelItemCount()):
                magazine = self.mag_tree.topLevelItem(i)
                magazine.setHidden(False)
                for j in range(magazine.childCount()):
                    year = magazine.child(j)
                    year.setHidden(False)
                    for k in range(year.childCount()):
                        year.child(k).setHidden(False)
            return
        
        # Filter tree items (magazine -> year -> edition)
        for i in range(self.mag_tree.topLevelItemCount()):
            magazine = self.mag_tree.topLevelItem(i)
            magazine_text = magazine.text(0).lower()
            magazine_matches = search_term in magazine_text
            
            visible_years = 0
            for j in range(magazine.childCount()):
                year = magazine.child(j)
                year_text = year.text(0).lower()
                year_matches = search_term in year_text
                
                visible_editions = 0
                for k in range(year.childCount()):
                    edition = year.child(k)
                    edition_text = f"{edition.text(0)} {edition.text(1)}".lower()
                    edition_matches = search_term in edition_text
                    
                    # Check grandchildren (missing editions under missing section)
                    visible_missing = 0
                    if edition.childCount() > 0:
                        for m in range(edition.childCount()):
                            missing_edition = edition.child(m)
                            missing_text = f"{missing_edition.text(0)} {missing_edition.text(1)}".lower()
                            missing_matches = search_term in missing_text
                            missing_edition.setHidden(not (magazine_matches or year_matches or edition_matches or missing_matches))
                            if magazine_matches or year_matches or edition_matches or missing_matches:
                                visible_missing += 1
                    
                    # Edition visible if it or its missing children match
                    edition.setHidden(not (magazine_matches or year_matches or edition_matches or visible_missing > 0))
                    if magazine_matches or year_matches or edition_matches or visible_missing > 0:
                        visible_editions += 1
                
                year.setHidden(visible_editions == 0)
                if visible_editions > 0:
                    visible_years += 1
            
            magazine.setHidden(visible_years == 0)
    
    def on_magazine_select(self) -> None:
        """Handle magazine edition selection from tree."""
        if not hasattr(self, "mag_tree"):
            return
        
        selected_items = self.mag_tree.selectedItems()
        if not selected_items:
            self._populate_question_sets([])
            self.question_label.setText("Select an edition to view question sets")
            self.question_label.setStyleSheet(
                "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
            )
            return
        
        item = selected_items[0]
        data = item.data(0, Qt.UserRole)
        
        if not isinstance(data, dict):
            self._populate_question_sets([])
            self.question_label.setText("Select an edition to view question sets")
            self.question_label.setStyleSheet(
                "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
            )
            return
        
        # Don't show question sets for missing editions or magazine/missing section/year headers
        if data.get("type") in ["magazine", "missing_section", "missing", "year"] or data.get("is_missing", False):
            self._populate_question_sets([])
            if data.get("type") == "missing" or data.get("is_missing", False):
                self.question_label.setText("❌ Missing edition - no data available")
                self.question_label.setStyleSheet(
                    "padding: 8px; background-color: #fee2e2; border-radius: 6px; color: #dc2626;"
                )
            else:
                self.question_label.setText("Select an edition to view question sets")
                self.question_label.setStyleSheet(
                    "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
                )
            return
        
        # Reset label style for valid editions
        self.question_label.setStyleSheet(
            "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
        )

        question_sets = data.get("question_sets", [])
        magazine_name = data.get("display_name", "")
        edition_label = data.get("edition_label", "")
        label = f"{magazine_name} - {edition_label}"
        
        if question_sets:
            self.question_label.setText(f"✓ {label} • {len(question_sets)} question set(s)")
            self._populate_question_sets(question_sets, magazine_name, edition_label)
        else:
            self.question_label.setText(f"⚠ {label} • No question sets found")
            self._populate_question_sets([])

    def on_group_selected(self) -> None:
        if not hasattr(self, "group_list"):
            return
        item = self.group_list.currentItem()
        self.group_chapter_list.clear()
        if not item:
            return
        group = item.data(Qt.UserRole) or item.text()
        for chapter in sorted(self.chapter_groups.get(group, []), key=lambda value: value.lower()):
            chapter_item = QListWidgetItem(chapter)
            chapter_item.setData(Qt.UserRole, chapter)
            self.group_chapter_list.addItem(chapter_item)

    def move_selected_chapter(self) -> None:
        group_item = getattr(self, "group_list", None)
        chapter_list = getattr(self, "group_chapter_list", None)
        target_combo = getattr(self, "move_target_combo", None)
        if (
            not group_item
            or not chapter_list
            or not target_combo
            or not group_item.currentItem()
            or not chapter_list.currentItem()
        ):
            return
        current_group = group_item.currentItem().data(Qt.UserRole)
        chapter_name = chapter_list.currentItem().data(Qt.UserRole)
        target_group = target_combo.currentText()
        if not chapter_name or not target_group or current_group == target_group:
            return
        self.move_chapter_to_group(chapter_name, target_group)

    def move_chapter_to_group(self, chapter_name: str, target_group: str, stay_on_group: str | None = None) -> None:
        if not chapter_name or not target_group:
            return
        for group, values in self.chapter_groups.items():
            if chapter_name in values:
                values.remove(chapter_name)
                break
        self.chapter_groups.setdefault(target_group, [])
        if chapter_name not in self.chapter_groups[target_group]:
            self.chapter_groups[target_group].append(chapter_name)
        self._save_chapter_grouping()
        self._refresh_grouping_ui()
        self._select_group_in_ui(stay_on_group or target_group)
        self.on_group_selected()

    def reassign_question(self, question: dict, target_group: str) -> None:
        """Reassign a single question to a different chapter."""
        if not question or not target_group:
            return
        if self.current_workbook_path is None or self.high_level_column_index is None:
            QMessageBox.warning(self, "Unavailable", "Workbook must be loaded before regrouping questions.")
            return
        old_group = question.get("group")
        if old_group == target_group:
            return
        row_number = question.get("row_number")
        if not isinstance(row_number, int):
            QMessageBox.warning(self, "Unavailable", "Question row information is missing.")
            return
        qno = question.get("qno", "")
        prompt = f"Move question '{qno}' to '{target_group}'?"
        if QMessageBox.question(self, "Confirm Reassignment", prompt) != QMessageBox.Yes:
            return
        try:
            workbook = load_workbook(self.current_workbook_path)
            sheet = workbook[workbook.sheetnames[0]]
            sheet.cell(row=row_number, column=self.high_level_column_index, value=target_group)
            workbook.save(self.current_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Update Failed", f"Unable to update workbook: {exc}")
            return
        self.log(f"Question '{qno}' moved to '{target_group}'. Reloading data...")
        self.update_row_count()

    def reassign_questions(self, questions: list[dict], target_group: str) -> None:
        """Reassign multiple questions to a different chapter in bulk."""
        if not questions or not target_group:
            return
        if self.current_workbook_path is None or self.high_level_column_index is None:
            QMessageBox.warning(self, "Unavailable", "Workbook must be loaded before regrouping questions.")
            return
        
        # Filter out questions already in target group and validate row numbers
        questions_to_move = []
        for question in questions:
            old_group = question.get("group")
            if old_group == target_group:
                continue
            row_number = question.get("row_number")
            if not isinstance(row_number, int):
                continue
            questions_to_move.append(question)
        
        if not questions_to_move:
            return
        
        count = len(questions_to_move)
        qnos = ", ".join(q.get("qno", "") for q in questions_to_move[:3])
        if count > 3:
            qnos += f" and {count - 3} more"
        
        prompt = f"Move {count} question(s) ({qnos}) to '{target_group}'?"
        if QMessageBox.question(self, "Confirm Reassignment", prompt) != QMessageBox.Yes:
            return
        
        try:
            workbook = load_workbook(self.current_workbook_path)
            sheet = workbook[workbook.sheetnames[0]]
            
            # Update all questions in one go
            for question in questions_to_move:
                row_number = question.get("row_number")
                sheet.cell(row=row_number, column=self.high_level_column_index, value=target_group)
            
            workbook.save(self.current_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Update Failed", f"Unable to update workbook: {exc}")
            return
        
        self.log(f"{count} question(s) moved to '{target_group}'. Reloading data...")
        self.update_row_count()

    def on_chapter_selected(self, chapter_key: str) -> None:
        if not chapter_key:
            self._populate_question_table([])
            self.current_selected_chapter = None
            return
        self.current_selected_chapter = chapter_key
        questions = self.chapter_questions.get(chapter_key, [])
        self._populate_question_table(questions)

    def on_question_selected(self) -> None:
        if not hasattr(self, "question_tree"):
            return
        selected_items = self.question_tree.selectedItems()
        if not selected_items:
            self.question_text_view.clear()
            return
        
        # Get the first selected item
        item = selected_items[0]
        
        # Skip if it's a group header (has children)
        if item.childCount() > 0:
            self.question_text_view.clear()
            return
        
        # Get question data from UserRole
        question = item.data(0, Qt.UserRole)
        if question:
            html = (
                f"<div style='background-color: #0f172a; color: #e2e8f0; font-family: Arial, sans-serif; padding: 10px;'>"
                f"<span style='color: #60a5fa; font-weight: bold;'>Qno:</span> <span style='color: #cbd5e1;'>{question.get('qno','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Page No:</span> <span style='color: #cbd5e1;'>{question.get('page','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Question Set:</span> <span style='color: #cbd5e1;'>{question.get('question_set_name','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Magazine Edition:</span> <span style='color: #cbd5e1;'>{question.get('magazine','')}</span><br/>"
                f"<hr style='border: none; border-top: 1px solid #334155; margin: 12px 0;'/>"
                f"<div style='color: #e2e8f0; line-height: 1.7; font-size: 14px;'>{question.get('text','').replace(chr(10), '<br/>')}</div>"
                f"</div>"
            )
            self.question_text_view.setHtml(html)
    
    def _on_copy_mode_changed(self, mode: str) -> None:
        """Handle copy mode selection change."""
        self.copy_mode = mode
    
    def on_question_card_selected(self, question: dict) -> None:
        """Handle question card click in card view."""
        if not question:
            self.question_text_view.clear()
            return
        
        # Display question details in the bottom panel
        html = (
            f"<div style='background-color: #0f172a; color: #e2e8f0; font-family: Arial, sans-serif; padding: 10px;'>"
            f"<span style='color: #60a5fa; font-weight: bold;'>Qno:</span> <span style='color: #cbd5e1;'>{question.get('qno','')}</span> &nbsp;&nbsp;"
            f"<span style='color: #60a5fa; font-weight: bold;'>Page No:</span> <span style='color: #cbd5e1;'>{question.get('page','')}</span> &nbsp;&nbsp;"
            f"<span style='color: #60a5fa; font-weight: bold;'>Question Set:</span> <span style='color: #cbd5e1;'>{question.get('question_set_name','')}</span> &nbsp;&nbsp;"
            f"<span style='color: #60a5fa; font-weight: bold;'>Magazine Edition:</span> <span style='color: #cbd5e1;'>{question.get('magazine','')}</span><br/>"
            f"<span style='color: #60a5fa; font-weight: bold;'>Chapter:</span> <span style='color: #cbd5e1;'>{question.get('group','')}</span><br/>"
            f"<hr style='border: none; border-top: 1px solid #334155; margin: 12px 0;'/>"
            f"<div style='color: #e2e8f0; line-height: 1.7; font-size: 14px;'>{question.get('text','').replace(chr(10), '<br/>')}</div>"
            f"</div>"
        )
        self.question_text_view.setHtml(html)
    
    def show_group_context_menu_for_card(self, group_key: str, position):
        """Show context menu for accordion group (from card view)."""
        menu = QMenu(self)
        
        # Add tag action
        add_tag_action = menu.addAction("🏷️ Assign Tag to Group")
        
        # Remove tag submenu (if tags exist)
        existing_tags = self.group_tags.get(group_key, [])
        if existing_tags:
            remove_tag_menu = menu.addMenu("🗑️ Remove Tag from Group")
            for tag in existing_tags:
                remove_action = remove_tag_menu.addAction(tag)
                remove_action.setData(("remove", group_key, tag))
        
        # Show menu and handle action
        action = menu.exec(position)
        
        if action == add_tag_action:
            self._assign_tag_to_group(group_key)
        elif action and action.data():
            action_type, grp_key, tag = action.data()
            if action_type == "remove":
                self._remove_tag_from_group(grp_key, tag)
    
    def on_mag_question_selected(self) -> None:
        """Handle question selection in Magazine Editions tab."""
        if not hasattr(self, "question_sets_tree"):
            return
        selected_items = self.question_sets_tree.selectedItems()
        if not selected_items:
            self.mag_question_text_view.clear()
            return
        
        # Get the first selected item
        item = selected_items[0]
        
        # Skip if it's a parent item (question set)
        data = item.data(0, Qt.UserRole)
        if not data or data.get("type") != "question":
            self.mag_question_text_view.clear()
            return
        
        # Get question details from the tree item data
        qno = data.get("qno", "")
        page = data.get("page", "")
        row_num = data.get("row", 0)
        
        # Get full question data from cached DataFrame
        if self.workbook_df is not None and row_num > 0:
            try:
                # Row number is 1-based Excel row, DataFrame is 0-based
                df_row_idx = row_num - 2  # Subtract 2 (1 for header, 1 for 0-based)
                
                if 0 <= df_row_idx < len(self.workbook_df):
                    row_data = self.workbook_df.iloc[df_row_idx]
                    
                    # Get column names
                    header_row = [None if pd.isna(col) else str(col) for col in self.workbook_df.columns]
                    
                    # Find columns
                    try:
                        question_text_col = _find_question_text_column(header_row)
                        question_text = str(row_data.iloc[question_text_col - 1]) if not pd.isna(row_data.iloc[question_text_col - 1]) else ""
                    except ValueError:
                        question_text = ""
                    
                    try:
                        question_set_name_col = _find_question_set_name_column(header_row)
                        question_set_name = str(row_data.iloc[question_set_name_col - 1]) if not pd.isna(row_data.iloc[question_set_name_col - 1]) else ""
                    except ValueError:
                        question_set_name = ""
                    
                    try:
                        magazine_col = _find_magazine_column(header_row)
                        magazine = str(row_data.iloc[magazine_col - 1]) if not pd.isna(row_data.iloc[magazine_col - 1]) else ""
                    except ValueError:
                        magazine = ""
                    
                    # Display question details in same format as Question List tab
                    html = (
                        f"<div style='background-color: #0f172a; color: #e2e8f0; font-family: Arial, sans-serif; padding: 10px;'>"
                        f"<span style='color: #60a5fa; font-weight: bold;'>Qno:</span> <span style='color: #cbd5e1;'>{qno}</span> &nbsp;&nbsp;"
                        f"<span style='color: #60a5fa; font-weight: bold;'>Page No:</span> <span style='color: #cbd5e1;'>{page}</span> &nbsp;&nbsp;"
                        f"<span style='color: #60a5fa; font-weight: bold;'>Question Set:</span> <span style='color: #cbd5e1;'>{question_set_name}</span> &nbsp;&nbsp;"
                        f"<span style='color: #60a5fa; font-weight: bold;'>Magazine Edition:</span> <span style='color: #cbd5e1;'>{magazine}</span><br/>"
                        f"<hr style='border: none; border-top: 1px solid #334155; margin: 12px 0;'/>"
                        f"<div style='color: #e2e8f0; line-height: 1.7; font-size: 14px;'>{question_text.replace(chr(10), '<br/>')}</div>"
                        f"</div>"
                    )
                    self.mag_question_text_view.setHtml(html)
                else:
                    self.mag_question_text_view.clear()
            except Exception as e:
                self.log(f"Error loading question text: {e}")
                self.mag_question_text_view.clear()
        else:
            self.mag_question_text_view.clear()

    def _on_tag_badge_clicked(self, tag: str) -> None:
        """Handle tag badge click - add to filter list."""
        if tag not in self.selected_tag_filters:
            self.selected_tag_filters.append(tag)
            if hasattr(self, "tag_filter_display"):
                self.tag_filter_display.setText(", ".join(self.selected_tag_filters))
            self._apply_question_search()

    def _show_group_context_menu(self, position) -> None:
        """Show context menu for group items."""
        item = self.question_tree.itemAt(position)
        if not item:
            return
        
        # Only show menu for group items (items with children)
        if item.childCount() == 0:
            return
        
        group_key = item.data(0, Qt.UserRole + 1)
        if not group_key:
            return
        
        menu = QMenu(self.question_tree)
        
        # Add tag action
        add_tag_action = menu.addAction("Add Tag")
        
        # Remove tag actions
        existing_tags = self.group_tags.get(group_key, [])
        if existing_tags:
            remove_menu = menu.addMenu("Remove Tag")
            for tag in existing_tags:
                remove_menu.addAction(tag)
        
        action = menu.exec(self.question_tree.viewport().mapToGlobal(position))
        
        if action:
            if action.text() == "Add Tag":
                self._assign_tag_to_group(group_key)
            else:
                # Remove tag
                self._remove_tag_from_group(group_key, action.text())

    def _assign_tag_to_group(self, group_key: str) -> None:
        """Show dialog to assign multiple tags to a group."""
        # Get all existing tags across all groups
        all_tags = sorted(set(tag for tags in self.group_tags.values() for tag in tags))
        
        # Get currently assigned tags for this group
        current_tags = self.group_tags.get(group_key, [])
        
        # Show multi-select dialog
        dialog = MultiSelectTagDialog(
            all_tags, 
            current_tags, 
            f"Add Tags to '{group_key.title()}'",
            self.tag_colors,
            self.available_tag_colors,
            self
        )
        
        if dialog.exec() == QDialog.Accepted:
            selected_tags = dialog.get_selected_tags()
            if selected_tags:
                self.group_tags[group_key] = selected_tags
                self._save_group_tags()
                self._apply_question_search(preserve_scroll=True)  # Refresh tree to show new tags

    def _remove_tag_from_group(self, group_key: str, tag: str) -> None:
        """Remove a tag from a group."""
        if group_key in self.group_tags and tag in self.group_tags[group_key]:
            self.group_tags[group_key].remove(tag)
            if not self.group_tags[group_key]:
                # Remove empty tag list
                del self.group_tags[group_key]
            self._save_group_tags()
            self._apply_question_search(preserve_scroll=True)  # Refresh tree to update display
        else:
            self.question_text_view.clear()

    def _load_saved_question_lists(self) -> None:
        """Load all saved question lists from QuestionList folder."""
        if not hasattr(self, "saved_lists_widget"):
            return
        
        self.saved_lists_widget.clear()
        self.question_lists.clear()
        
        if not QUESTION_LIST_DIR.exists():
            return
        
        for json_file in sorted(QUESTION_LIST_DIR.glob("*.json")):
            try:
                data = json.loads(json_file.read_text(encoding="utf-8"))
                list_name = data.get("name", json_file.stem)
                questions = data.get("questions", [])
                magazine = data.get("magazine", "")  # Load magazine name
                filters = data.get("filters", {})  # Load filters
                
                self.question_lists[list_name] = questions
                self.question_lists_metadata[list_name] = {
                    "magazine": magazine,
                    "filters": filters
                }
                
                # Display list name with question count and magazine
                display_text = f"{list_name} ({len(questions)})"
                if magazine:
                    display_text += f" - {magazine}"
                
                item = QListWidgetItem(display_text)
                item.setData(Qt.UserRole, list_name)
                self.saved_lists_widget.addItem(item)
            except Exception as exc:
                self.log(f"Error loading list {json_file.name}: {exc}")
        
        # Trigger selection of first item if any lists loaded
        if self.saved_lists_widget.count() > 0:
            self.saved_lists_widget.setCurrentRow(0)
        
        # Update drag-drop panel dropdown with loaded lists
        self.drag_drop_panel.update_list_selector(self.question_lists)
    
    def _save_question_list(self, list_name: str, save_filters: bool = False) -> None:
        """Save a question list to file.
        
        Args:
            list_name: Name of the list to save
            save_filters: If True, save current active filters with the list
        """
        if list_name not in self.question_lists:
            return
        
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in list_name)
        file_path = QUESTION_LIST_DIR / f"{safe_name}.json"
        
        # Get or create metadata
        metadata = self.question_lists_metadata.get(list_name, {})
        
        data = {
            "name": list_name,
            "magazine": metadata.get("magazine", self.current_magazine_name),
            "questions": self.question_lists[list_name],
        }
        
        # Save filters if requested or if they already exist
        if save_filters or metadata.get("filters"):
            filters = self._get_active_filters() if save_filters else metadata.get("filters", {})
            if filters:
                data["filters"] = filters
        
        try:
            file_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
            self.log(f"Saved question list: {list_name}")
        except Exception as exc:
            QMessageBox.critical(self, "Save Error", f"Failed to save list: {exc}")
    
    def create_new_question_list(self) -> None:
        """Create a new question list."""
        from PySide6.QtWidgets import QInputDialog
        
        list_name, ok = QInputDialog.getText(self, "New Question List", "Enter list name:")
        if not ok or not list_name.strip():
            return
        
        list_name = list_name.strip()
        if list_name in self.question_lists:
            QMessageBox.warning(self, "Duplicate Name", f"A list named '{list_name}' already exists.")
            return
        
        self.question_lists[list_name] = []
        self._save_question_list(list_name)
        self._load_saved_question_lists()
        self.drag_drop_panel.update_list_selector(self.question_lists)
        self.log(f"Created new question list: {list_name}")
    
    def rename_question_list(self) -> None:
        """Rename selected question list."""
        from PySide6.QtWidgets import QInputDialog
        
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            QMessageBox.information(self, "No Selection", "Please select a list to rename.")
            return
        
        old_name = current_item.data(Qt.UserRole)
        new_name, ok = QInputDialog.getText(self, "Rename List", "Enter new name:", text=old_name)
        if not ok or not new_name.strip():
            return
        
        new_name = new_name.strip()
        if new_name == old_name:
            return
        
        if new_name in self.question_lists:
            QMessageBox.warning(self, "Duplicate Name", f"A list named '{new_name}' already exists.")
            return
        
        # Delete old file
        safe_old_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in old_name)
        old_file = QUESTION_LIST_DIR / f"{safe_old_name}.json"
        old_file.unlink(missing_ok=True)
        
        # Update data structure
        self.question_lists[new_name] = self.question_lists.pop(old_name)
        self._save_question_list(new_name)
        self._load_saved_question_lists()
        self.drag_drop_panel.update_list_selector(self.question_lists)
        self.log(f"Renamed list from '{old_name}' to '{new_name}'")
    
    def delete_question_list(self) -> None:
        """Delete selected question list."""
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            QMessageBox.information(self, "No Selection", "Please select a list to delete.")
            return
        
        list_name = current_item.data(Qt.UserRole)
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete the list '{list_name}'?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        
        # Delete file
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in list_name)
        file_path = QUESTION_LIST_DIR / f"{safe_name}.json"
        file_path.unlink(missing_ok=True)
        
        # Update data structure
        del self.question_lists[list_name]
        self._load_saved_question_lists()
        self.drag_drop_panel.update_list_selector(self.question_lists)
        self.list_question_table.setRowCount(0)
        self.list_name_label.setText("Select a list to view questions")
        self.log(f"Deleted question list: {list_name}")
    
    def add_selected_to_list(self) -> None:
        """Toggle the drag-drop panel for adding questions to a list."""
        if not self.question_lists:
            reply = QMessageBox.question(
                self,
                "No Lists",
                "No question lists exist. Create a new list?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                self.create_new_question_list()
            return
        
        # Toggle drag-drop panel visibility
        self.drag_drop_panel.setVisible(not self.drag_drop_panel.isVisible())
        
        # Update panel with current question lists
        if self.drag_drop_panel.isVisible():
            from src.ui.widgets import DragDropQuestionPanel
            # Recreate panel with updated question lists
            old_panel = self.drag_drop_panel
            self.drag_drop_panel = DragDropQuestionPanel(self.question_lists, self)
            self.drag_drop_panel.save_clicked.connect(self._on_drag_drop_save)
            self.drag_drop_panel.cancel_clicked.connect(self._on_drag_drop_cancel)
            
            # Replace in layout
            question_layout = old_panel.parent().layout()
            for i in range(question_layout.count()):
                if question_layout.itemAt(i).widget() == old_panel:
                    old_panel.deleteLater()
                    question_layout.insertWidget(i, self.drag_drop_panel)
                    break
    
    def _on_drag_drop_save(self, list_name: str, questions: list):
        """Handle save from drag-drop panel."""
        added_count = 0
        for question in questions:
            # Check for duplicates based on row_number
            if not any(q.get("row_number") == question.get("row_number") for q in self.question_lists[list_name]):
                self.question_lists[list_name].append(question.copy())
                added_count += 1
        
        self._save_question_list(list_name)
        self._load_saved_question_lists()
        self.log(f"Added {added_count} question(s) to list '{list_name}'")
        QMessageBox.information(self, "Success", f"Added {added_count} question(s) to '{list_name}'.")
        
        # Hide panel
        self.drag_drop_panel.setVisible(False)
    
    def _on_drag_drop_cancel(self):
        """Handle cancel from drag-drop panel."""
        self.drag_drop_panel.setVisible(False)
    
    def _highlight_question_card(self, question_data: dict):
        """Highlight the corresponding question card with yellow background."""
        from PySide6.QtCore import QTimer
        
        # Find and highlight the card in the question card view
        for group in self.question_card_view.accordion_groups:
            for card in group.get_all_cards():
                if card.question_data.get("row_number") == question_data.get("row_number"):
                    # Expand accordion group if collapsed
                    if not group.is_expanded:
                        group.toggle_expanded()
                    
                    # Temporarily highlight with yellow
                    card.setStyleSheet("""
                        QLabel {
                            background-color: #fef08a;
                            border: 3px solid #eab308;
                            border-radius: 8px;
                            padding: 12px;
                            margin: 4px;
                        }
                    """)
                    
                    # Reset after 2 seconds
                    def reset_card_style():
                        if card.is_selected:
                            card.set_selected(True)
                        else:
                            card.setStyleSheet(card.original_stylesheet)
                    
                    QTimer.singleShot(2000, reset_card_style)
                    
                    # Scroll to make the card visible
                    self.question_card_view.ensureWidgetVisible(card)
                    return
    
    def _remove_question_from_list(self, question: dict) -> None:
        """Remove a single question from current list (called from remove button on card)."""
        if not self.current_list_name:
            QMessageBox.information(self, "No List Selected", "Please select a list first.")
            return
        
        # Find and remove the question by qno
        qno = question.get('qno')
        questions = self.question_lists[self.current_list_name]
        
        for idx, q in enumerate(questions):
            if q.get('qno') == qno:
                del questions[idx]
                self._save_question_list(self.current_list_name)
                self.on_saved_list_selected()  # Refresh display
                self.log(f"Removed question Q{qno} from '{self.current_list_name}'")
                return
        
        QMessageBox.warning(self, "Not Found", f"Question Q{qno} not found in list.")
    
    def _populate_list_card_view(self, questions: list[dict]) -> None:
        """Populate card view with 2-column grid of question cards from custom list using QuestionCardWithRemoveButton wrapper."""
        if not questions:
            return
        
        # Clear existing layout
        if not hasattr(self, '_list_card_grid_widget'):
            # Create a new grid widget for 2-column layout
            self._list_card_grid_widget = QWidget()
            self._list_card_grid_layout = QGridLayout(self._list_card_grid_widget)
            self._list_card_grid_layout.setSpacing(12)
            self._list_card_grid_layout.setContentsMargins(0, 0, 0, 0)
            
            # Set the grid widget in scroll area
            self.list_question_card_view.setWidget(self._list_card_grid_widget)
        else:
            # Clear existing cards from grid
            while self._list_card_grid_layout.count():
                item = self._list_card_grid_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
        
        # Add question cards in 2-column grid using QuestionCardWithRemoveButton wrapper
        for idx, question in enumerate(questions):
            card_wrapper = QuestionCardWithRemoveButton(question, self)
            card_wrapper.clicked.connect(lambda q=question: self.on_list_question_card_selected(q))
            card_wrapper.remove_requested.connect(lambda q=question: self._remove_question_from_list(q))
            row = idx // 2
            col = idx % 2
            self._list_card_grid_layout.addWidget(card_wrapper, row, col)
        
        # Add stretch to push cards to top
        self._list_card_grid_layout.addItem(QSpacerItem(0, 0, QSizePolicy.Minimum, QSizePolicy.Expanding), 
                                           len(questions) // 2 + 1, 0, 1, 2)
    
    def _create_list_question_card(self, question: dict) -> QWidget:
        """Create a single question card widget for custom list (deprecated - use QuestionCardWidget directly)."""
        # This method is deprecated - use QuestionCardWidget from widgets.py instead
        pass
    
    def on_list_question_set_search_changed(self, text: str) -> None:
        """Handle question set search change in custom list."""
        self.list_question_set_search_term = text.strip()
        self._apply_list_search()
    
    def on_list_tag_filter_changed(self) -> None:
        """Handle tag filter change in custom list."""
        self._apply_list_search()
    
    def _apply_list_search(self) -> None:
        """Apply search/filter to custom list questions."""
        if not self.current_list_name or not hasattr(self, '_list_card_grid_layout'):
            return
        
        questions = self.question_lists[self.current_list_name]
        if not questions:
            return
        
        filtered = questions
        
        # Apply question set search
        if self.list_question_set_search_term:
            normalized_search = self._normalize_label(self.list_question_set_search_term)
            filtered = [
                q for q in filtered
                if normalized_search in self._normalize_label(q.get("question_set_name", ""))
            ]
        
        # Clear and repopulate card view
        while self._list_card_grid_layout.count():
            item = self._list_card_grid_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if filtered:
            self._populate_list_card_view(filtered)
    
    def clear_list_search(self) -> None:
        """Clear search in custom list."""
        self.list_question_set_search.clear()
        self.list_question_set_search_term = ""
        self._apply_list_search()
    
    def _on_list_copy_mode_changed(self, mode: str) -> None:
        """Handle copy mode selection change in custom list."""
        self.list_copy_mode = mode
    
    def on_saved_list_selected(self) -> None:
        """Handle selection of a saved question list."""
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            self.list_question_table.setRowCount(0)
            self.list_name_label.setText("Select a list to view questions")
            self.current_list_name = None
            self.drag_drop_panel.display_existing_questions([])
            # Hide panel if no list selected
            self.drag_drop_panel.setVisible(False)
            return
        
        list_name = current_item.data(Qt.UserRole)
        self.current_list_name = list_name
        self._populate_list_question_table(list_name)
        
        # Display existing questions in drag-drop panel
        if list_name in self.question_lists:
            questions = self.question_lists[list_name]
            self.drag_drop_panel.display_existing_questions(questions)
            # Show the panel to display existing questions
            self.drag_drop_panel.setVisible(True)
    
    def _populate_list_question_table(self, list_name: str) -> None:
        """Populate the list question card view with questions from the selected list."""
        if list_name not in self.question_lists:
            return
        
        questions = self.question_lists[list_name]
        self.current_list_questions = questions
        self.list_name_label.setText(f"{list_name} ({len(questions)} questions)")
        
        # Show filters if they exist
        metadata = self.question_lists_metadata.get(list_name, {})
        filters = metadata.get("filters", {})
        if filters:
            filter_parts = []
            if filters.get("selected_chapter"):
                filter_parts.append(f"Chapter: '{filters['selected_chapter']}'")
            if filters.get("question_set_search"):
                filter_parts.append(f"Question Set: '{filters['question_set_search']}'")
            if filters.get("selected_tags"):
                filter_parts.append(f"Tags: {', '.join(filters['selected_tags'])}")
            if filters.get("selected_magazine"):
                filter_parts.append(f"Magazine: {filters['selected_magazine']}")
            
            if filter_parts:
                self.list_filters_label.setText(f"🔍 Filters: {' | '.join(filter_parts)}")
                self.list_filters_label.setVisible(True)
            else:
                self.list_filters_label.setVisible(False)
        else:
            self.list_filters_label.setVisible(False)
        
        # Clear existing card grid
        if hasattr(self, '_list_card_grid_layout'):
            while self._list_card_grid_layout.count():
                item = self._list_card_grid_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
        
        # Reset search state for custom lists
        self.list_question_set_search_term = ""
        self.list_question_set_search.clear()
        
        if not questions:
            return
        
        # Populate card view with 2-column grid
        self._populate_list_card_view(questions)
        
        # Show drag-drop panel with existing questions
        if list_name in self.question_lists:
            self.drag_drop_panel.display_existing_questions(questions)
            self.drag_drop_panel.setVisible(True)
    
    def on_list_question_selected(self) -> None:
        """Handle selection of a question in the list question table (old method - kept for compatibility)."""
        # This method is no longer used with card view, but kept for reference
        pass
    
    def on_list_question_card_selected(self, question: dict) -> None:
        """Handle question card click in custom list card view."""
        # Card selection used for information only - no separate text view needed
        pass
    
    def _get_active_filters(self) -> dict:
        """Get currently active filters in Question Analysis tab."""
        filters = {}
        
        # Selected chapter filter (from left table)
        if self.current_selected_chapter:
            filters["selected_chapter"] = self.current_selected_chapter
        
        # Question Set search filter
        if self.question_set_search.text().strip():
            filters["question_set_search"] = self.question_set_search.text().strip()
        
        # Tag filters
        if self.selected_tag_filters:
            filters["selected_tags"] = self.selected_tag_filters.copy()
        
        # Magazine filter (from selected magazine edition)
        if self.magazine_search_term:
            filters["selected_magazine"] = self.magazine_search_term
        
        return filters
    
    def create_random_list_from_filtered(self) -> None:
        """Create a new question list with random questions from currently filtered list."""
        from PySide6.QtWidgets import QInputDialog, QDialog, QVBoxLayout, QLabel, QSpinBox, QDialogButtonBox
        
        # Get all visible questions from current filtered view
        visible_questions = []
        
        # Collect questions from card view instead of tree
        if hasattr(self, 'question_card_view'):
            for group in self.question_card_view.accordion_groups:
                for card in group.get_all_cards():
                    question_data = card.question_data
                    if question_data:
                        visible_questions.append(question_data)
        else:
            # Fallback to old tree method if card view doesn't exist
            root = self.question_tree.invisibleRootItem()
            
            def collect_visible_questions(parent_item):
                for i in range(parent_item.childCount()):
                    child = parent_item.child(i)
                    if not child.isHidden():
                        if child.childCount() == 0:  # Leaf node (question)
                            question_data = child.data(0, Qt.UserRole)
                            if question_data:
                                visible_questions.append(question_data)
                        else:  # Group node, recurse
                            collect_visible_questions(child)
            
            collect_visible_questions(root)
        
        if not visible_questions:
            QMessageBox.warning(self, "No Questions", "No questions available in the current filtered view.")
            return
        
        # Create dialog for list name and question count
        dialog = QDialog(self)
        dialog.setWindowTitle("Create Random Question List")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        
        # List name input
        layout.addWidget(QLabel("List Name:"))
        name_input = QLineEdit()
        layout.addWidget(name_input)
        
        # Question count input
        layout.addWidget(QLabel(f"Number of Questions (max {len(visible_questions)}):"))
        count_spinner = QSpinBox()
        count_spinner.setMinimum(1)
        count_spinner.setMaximum(len(visible_questions))
        count_spinner.setValue(min(30, len(visible_questions)))  # Default 30
        layout.addWidget(count_spinner)
        
        # Filter info
        active_filters = self._get_active_filters()
        if active_filters:
            filter_info = QLabel("<b>Active filters will be saved with this list:</b>")
            layout.addWidget(filter_info)
            
            filter_parts = []
            if active_filters.get("selected_chapter"):
                filter_parts.append(f"• Chapter: '{active_filters['selected_chapter']}'")
            if active_filters.get("question_set_search"):
                filter_parts.append(f"• Question Set: '{active_filters['question_set_search']}'")
            if active_filters.get("selected_tags"):
                filter_parts.append(f"• Tags: {', '.join(active_filters['selected_tags'])}")
            if active_filters.get("selected_magazine"):
                filter_parts.append(f"• Magazine: {active_filters['selected_magazine']}")
            
            if filter_parts:
                filter_label = QLabel("\n".join(filter_parts))
                filter_label.setStyleSheet("padding: 8px; background-color: #fef3c7; border-radius: 4px; color: #92400e;")
                layout.addWidget(filter_label)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        if dialog.exec() != QDialog.Accepted:
            return
        
        list_name = name_input.text().strip()
        if not list_name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a list name.")
            return
        
        if list_name in self.question_lists:
            reply = QMessageBox.question(
                self,
                "List Exists",
                f"List '{list_name}' already exists. Overwrite?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
        
        # Select random questions
        import random
        num_questions = count_spinner.value()
        selected_questions = random.sample(visible_questions, num_questions)
        
        # Create the list
        self.question_lists[list_name] = [q.copy() for q in selected_questions]
        self.question_lists_metadata[list_name] = {
            "magazine": self.current_magazine_name,
            "filters": active_filters
        }
        
        self._save_question_list(list_name, save_filters=True)
        self._load_saved_question_lists()
        
        self.log(f"Created random list '{list_name}' with {num_questions} questions from filtered view")
        QMessageBox.information(
            self,
            "Success",
            f"Created list '{list_name}' with {num_questions} random questions from filtered view."
        )

    def log(self, message: str) -> None:
        timestamp = time.strftime("%H:%M:%S")
        self.log_view.append(f"[{timestamp}] {message}")
        self.log_view.moveCursor(QTextCursor.End)

    def select_input_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.input_edit.setText(folder)
            self._save_last_selection()
            self.refresh_file_list()
            self._auto_start_watching()  # Auto-start watching

    def select_output_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel Workbook",
            "",
            "Excel files (*.xlsx)",
        )
        if file_path:
            self.output_edit.setText(file_path)
            self._save_last_selection()
            self.update_row_count()

    def refresh_file_list(self) -> None:
        input_path = Path(self.input_edit.text().strip())
        if not input_path.exists():
            self.log("Input folder does not exist.")
            return

        tsv_files = sorted(input_path.glob("*.tsv"))
        for tsv_file in tsv_files:
            self._ensure_row(tsv_file.name)
        self.log(f"Found {len(tsv_files)} TSV file(s) in input folder.")

    def _ensure_row(self, filename: str) -> int:
        row = self.file_rows.get(filename)
        if row is not None:
            return row
        row = self.file_table.rowCount()
        self.file_table.insertRow(row)
        self.file_table.setItem(row, 0, QTableWidgetItem(filename))
        self.file_table.setItem(row, 1, QTableWidgetItem("Pending"))
        self.file_table.setItem(row, 2, QTableWidgetItem("Awaiting processing"))
        self.file_rows[filename] = row
        return row

    def update_file_status(self, filename: str, status: str, message: str) -> None:
        row = self._ensure_row(filename)
        for col, value in enumerate((filename, status, message)):
            item = self.file_table.item(row, col)
            if item is None:
                item = QTableWidgetItem(value)
                self.file_table.setItem(row, col, item)
            else:
                item.setText(value)
        if status.lower() == "error":
            self.file_errors[filename] = message
        else:
            self.file_errors.pop(filename, None)

    def on_file_double_clicked(self, row: int, column: int) -> None:  # noqa: ARG002
        filename_item = self.file_table.item(row, 0)
        status_item = self.file_table.item(row, 1)
        message_item = self.file_table.item(row, 2)
        if not filename_item or not status_item:
            return
        if status_item.text().lower() != "error":
            return
        filename = filename_item.text()
        detail = self.file_errors.get(filename, message_item.text() if message_item else "")
        QMessageBox.critical(self, "Validation Error", f"{filename}\n\n{detail}")

    def start_watching(self) -> None:
        if self.watch_thread and self.watch_thread.is_alive():
            QMessageBox.information(self, "Watcher", "Watcher is already running.")
            return

        input_path = Path(self.input_edit.text().strip())
        workbook_path = Path(self.output_edit.text().strip())

        if not input_path.is_dir():
            QMessageBox.critical(self, "Invalid Input", "Please select a valid input folder.")
            return
        if not workbook_path.is_file():
            QMessageBox.critical(self, "Invalid Workbook", "Please select an existing Excel workbook.")
            return

        self.stop_event.clear()
        self.watch_thread = threading.Thread(
            target=self._watch_loop, args=(input_path, workbook_path), daemon=True
        )
        self.watch_thread.start()
        self.start_button.setEnabled(False)
        self.log("Started watching for TSV files.")
    
    def _auto_start_watching(self) -> None:
        """Automatically start watching if both input folder and workbook are valid."""
        if self.watch_thread and self.watch_thread.is_alive():
            return  # Already watching
        
        input_path = Path(self.input_edit.text().strip())
        workbook_path = Path(self.output_edit.text().strip())
        
        if input_path.is_dir() and workbook_path.is_file():
            self.stop_event.clear()
            self.watch_thread = threading.Thread(
                target=self._watch_loop, args=(input_path, workbook_path), daemon=True
            )
            self.watch_thread.start()
            self.start_button.setEnabled(False)
            self.log("Auto-started watching for TSV files.")

    def stop_watching(self) -> None:
        self.stop_event.set()
        if self.watch_thread and self.watch_thread.is_alive():
            self.watch_thread.join(timeout=0.1)
        self.start_button.setEnabled(True)
        self.log("Stopped watching.")

    def _watch_loop(self, input_dir: Path, workbook_path: Path) -> None:
        poll_interval = 3.0
        while not self.stop_event.is_set():
            tsv_files = sorted(input_dir.glob("*.tsv"))
            if not tsv_files:
                time.sleep(poll_interval)
                continue

            for tsv_file in tsv_files:
                if self.stop_event.is_set():
                    break
                self.event_queue.put(("status", tsv_file.name, "Processing", "Validating..."))
                self.event_queue.put(("status_importing", tsv_file.name))
                try:
                    result_message = process_tsv(tsv_file, workbook_path, self.current_magazine_name)
                except Exception as exc:
                    self.event_queue.put(("status", tsv_file.name, "Error", str(exc)))
                    self.event_queue.put(("log", f"Error processing {tsv_file.name}: {exc}"))
                    self.event_queue.put(("status_error", f"Failed to import {tsv_file.name}: {exc}"))
                else:
                    self.event_queue.put(("status", tsv_file.name, "Completed", result_message))
                    self.event_queue.put(("log", f"Processed {tsv_file.name}: {result_message}"))
                    self.event_queue.put(("rowcount",))
                    self.event_queue.put(("status_success", tsv_file.name, result_message))

            time.sleep(poll_interval)

    def _process_queue(self) -> None:
        while True:
            try:
                event = self.event_queue.get_nowait()
            except queue.Empty:
                break

            event_type = event[0]
            if event_type == "status":
                _, filename, status, message = event
                self.update_file_status(filename, status, message)
            elif event_type == "log":
                _, message = event
                self.log(message)
            elif event_type == "metrics":
                _, req_id, row_count, details, warnings, chapter_data, question_col, raw_chapter_inputs, detected_magazine = event
                if req_id != self.metrics_request_id:
                    continue
                self.high_level_column_index = question_col
                
                # Magazine grouping already loaded in worker thread, just log it
                if detected_magazine:
                    self.log(f"Loaded chapter grouping for: {detected_magazine or 'default (Physics)'}")
                
                self.row_count_label.setText(f"Total rows: {row_count}")
                # Show success status after workbook loads
                self.set_status(f'Loaded workbook "{self.current_workbook_path.name}"', "success")
                total_editions = sum(len(entry["editions"]) for entry in details)
                self._set_magazine_summary(
                    f"Magazines loaded: {len(details)}",
                    f"Tracked editions: {total_editions}",
                )
                self._populate_magazine_tree(details)
                self._populate_question_sets([])
                missing_qset_warning = next(
                    (msg for msg in warnings if "question set" in msg.lower()), None
                )
                if missing_qset_warning:
                    label_message = missing_qset_warning
                elif details:
                    label_message = "Select an edition to view question sets."
                else:
                    label_message = warnings[0] if warnings else "No magazine editions found."
                self.question_label.setText(label_message)
                self._auto_assign_chapters(raw_chapter_inputs)
                self._populate_chapter_list(chapter_data)
                self._refresh_grouping_ui()
                
                # Update dashboard with statistics
                if hasattr(self, 'dashboard_view') and hasattr(self, 'workbook_df'):
                    self.dashboard_view.update_dashboard_data(self.workbook_df, self.chapter_groups)
                
                for warning in warnings:
                    self.log(warning)
            elif event_type == "metrics_error":
                _, req_id, error_message = event
                if req_id != self.metrics_request_id:
                    continue
                self.high_level_column_index = None
                self.row_count_label.setText("Total rows: Error")
                self._set_magazine_summary("Magazines: Error", "Missing ranges: Error")
                self._populate_magazine_tree([])
                self._populate_question_sets([])
                self._populate_chapter_list({})
                self._refresh_grouping_ui()
                self.question_label.setText("Unable to load editions.")
                self.log(f"Unable to read workbook rows: {error_message}")
            elif event_type == "status_error":
                _, error_msg = event
                self.set_status(f"Error: {error_msg}", "error")
            elif event_type == "status_importing":
                _, filename = event
                self.set_status(f'Importing data from "{filename}"', "importing")
            elif event_type == "status_success":
                _, filename, result_msg = event
                self.set_status(f'Data imported from "{filename}" • {result_msg}', "success")
            elif event_type == "rowcount":
                self.update_row_count()
        # Timer will trigger this method again; no manual reschedule needed.

    # ============================================================================
    # JEE Main Papers Methods
    # ============================================================================
    
    def select_jee_papers_file(self) -> None:
        """Open file dialog to select JEE papers Excel file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select JEE Papers File",
            "",
            "Excel files (*.xlsx *.xls);;All files (*.*)",
        )
        if file_path:
            self.jee_papers_file = Path(file_path)
            self.jee_file_edit.setText(file_path)
            self._save_last_selection()
            self.load_jee_papers_data()
    
    def load_jee_papers_data(self) -> None:
        """Load and process JEE papers data from Excel file."""
        if not self.jee_papers_file or not self.jee_papers_file.exists():
            self.log("JEE papers file not found.")
            return
        
        try:
            # Read Excel file from "Data" sheet, columns A:E
            self.jee_papers_df = pd.read_excel(
                self.jee_papers_file, 
                sheet_name='Data',
                usecols='A:E'
            )
            
            # Validate required columns
            required_columns = ['Question Number', 'JEE Main Session', 'Year', 'Subject', 'Chapter']
            missing_columns = [col for col in required_columns if col not in self.jee_papers_df.columns]
            
            if missing_columns:
                self.log(f"Error: Missing required columns: {', '.join(missing_columns)}")
                QMessageBox.warning(
                    self,
                    "Invalid File Format",
                    f"Missing required columns: {', '.join(missing_columns)}\nExpected columns in sheet 'Data': {', '.join(required_columns)}"
                )
                return
            
            # Populate subject dropdown
            subjects = sorted(self.jee_papers_df['Subject'].unique())
            self.jee_subject_combo.clear()
            self.jee_subject_combo.addItems(subjects)
            
            self.log(f"Loaded {len(self.jee_papers_df)} questions from JEE papers Excel file (Sheet: 'Data', Columns: A:E).")
            
            # Update tables if a subject is selected
            if subjects:
                self.update_jee_tables()
                
        except ValueError as e:
            # Handle missing sheet or invalid range
            error_msg = str(e)
            if 'Data' in error_msg:
                self.log(f"Error: Sheet 'Data' not found in Excel file.")
                QMessageBox.critical(self, "Error", "Sheet 'Data' not found in the Excel file.\nPlease ensure the file has a sheet named 'Data'.")
            else:
                self.log(f"Error loading JEE papers file: {error_msg}")
                QMessageBox.critical(self, "Error", f"Failed to load file: {error_msg}")
        except Exception as e:
            self.log(f"Error loading JEE papers file: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load file: {e}")
    
    def update_jee_tables(self) -> None:
        """Update the chapters table based on selected subject."""
        if self.jee_papers_df is None or self.jee_papers_df.empty:
            return
        
        subject = self.jee_subject_combo.currentText()
        if not subject:
            return
        
        try:
            # Filter by subject
            subject_df = self.jee_papers_df[self.jee_papers_df['Subject'] == subject]
            
            if subject_df.empty:
                self.log(f"No data found for subject: {subject}")
                return
            
            # Count questions per chapter
            chapter_counts = subject_df['Chapter'].value_counts().sort_values(ascending=False)
            
            # Populate chapters table
            self.jee_chapters_table.setRowCount(len(chapter_counts))
            for row, (chapter, count) in enumerate(chapter_counts.items()):
                chapter_item = QTableWidgetItem(chapter)
                count_item = QTableWidgetItem(str(count))
                count_item.setTextAlignment(Qt.AlignCenter)
                self.jee_chapters_table.setItem(row, 0, chapter_item)
                self.jee_chapters_table.setItem(row, 1, count_item)
            
            # Clear questions table
            self.jee_questions_table.setRowCount(0)
            self.jee_questions_table.setColumnCount(0)
            self.jee_questions_label.setText(f"Select a chapter to view questions ({subject})")
            
            self.log(f"Updated tables for {subject}: {len(chapter_counts)} chapters, {subject_df.shape[0]} total questions")
            
        except Exception as e:
            self.log(f"Error updating JEE tables: {e}")
    
    def on_jee_chapter_selected(self) -> None:
        """Handle chapter selection to show questions for that chapter."""
        if self.jee_papers_df is None or self.jee_papers_df.empty:
            return
        
        selected_items = self.jee_chapters_table.selectedItems()
        if not selected_items:
            return
        
        # Get selected chapter name
        row = selected_items[0].row()
        chapter_item = self.jee_chapters_table.item(row, 0)
        if not chapter_item:
            return
        
        chapter_name = chapter_item.text()
        subject = self.jee_subject_combo.currentText()
        
        try:
            # Filter questions by subject and chapter
            filtered_df = self.jee_papers_df[
                (self.jee_papers_df['Subject'] == subject) & 
                (self.jee_papers_df['Chapter'] == chapter_name)
            ]
            
            if filtered_df.empty:
                self.jee_questions_label.setText(f"No questions found for {chapter_name}")
                self.jee_questions_table.setRowCount(0)
                return
            
            # Update label
            self.jee_questions_label.setText(f"{chapter_name} - {len(filtered_df)} Questions")
            
            # Setup columns (all columns from Excel)
            columns = list(filtered_df.columns)
            self.jee_questions_table.setColumnCount(len(columns))
            self.jee_questions_table.setHorizontalHeaderLabels(columns)
            
            # Populate rows
            self.jee_questions_table.setRowCount(len(filtered_df))
            for row_idx, (_, row_data) in enumerate(filtered_df.iterrows()):
                for col_idx, col_name in enumerate(columns):
                    value = str(row_data[col_name]) if pd.notna(row_data[col_name]) else ""
                    item = QTableWidgetItem(value)
                    self.jee_questions_table.setItem(row_idx, col_idx, item)
            
            # Auto-resize columns to content
            self.jee_questions_table.resizeColumnsToContents()
            
            self.log(f"Showing {len(filtered_df)} questions for {chapter_name}")
            
        except Exception as e:
            self.log(f"Error displaying chapter questions: {e}")

    def closeEvent(self, event) -> None:
        self.stop_watching()
        super().closeEvent(event)


