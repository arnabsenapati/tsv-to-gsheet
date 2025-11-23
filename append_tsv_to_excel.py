from __future__ import annotations

import csv
import json
import sys
import datetime as dt
import queue
import re
import threading
import time
from decimal import Decimal, InvalidOperation
from itertools import repeat
from pathlib import Path
from PySide6.QtCore import Qt, QMimeData, QTimer, Signal
from PySide6.QtGui import QColor, QPalette, QDrag, QDragEnterEvent, QDropEvent, QTextCursor
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QAbstractItemView,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QHeaderView,
)

from openpyxl import load_workbook
import pandas as pd

BASE_DIR = Path(__file__).resolve().parent
PHYSICS_CHAPTER_FILE = BASE_DIR / "physicsCHapters.txt"
PHYSICS_GROUPING_FILE = BASE_DIR / "PhysicsChapterGrouping.json"
CHEMISTRY_GROUPING_FILE = BASE_DIR / "ChemistryChapterGrouping.json"
MATHEMATICS_GROUPING_FILE = BASE_DIR / "MathematicsChapterGrouping.json"
LAST_SELECTION_FILE = BASE_DIR / "last_selection.json"
QUESTION_LIST_DIR = BASE_DIR / "QuestionList"
TAGS_CONFIG_FILE = BASE_DIR / "tags.cfg"

# Magazine name to grouping file mapping
MAGAZINE_GROUPING_MAP = {
    "chemistry today": CHEMISTRY_GROUPING_FILE,
    "physics for you": PHYSICS_GROUPING_FILE,
    "mathematics today": MATHEMATICS_GROUPING_FILE,
}


MONTH_ALIASES = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def validate_tsv(tsv_path: Path) -> None:
    with tsv_path.open("r", encoding="utf-8", newline="") as tsv_file:
        reader = csv.reader(tsv_file, delimiter="\t")
        header = next(reader, None)
        if header is None:
            raise ValueError(f"{tsv_path.name} is empty or missing a header row.")

        column_count = len(header)
        for line_no, row in enumerate(reader, start=2):
            if len(row) != column_count:
                raise ValueError(
                    f"{tsv_path.name} line {line_no}: expected {column_count} columns but found {len(row)}."
                )
            for col_idx, value in enumerate(row, start=1):
                if value.strip() == "":
                    raise ValueError(
                        f"{tsv_path.name} line {line_no} column {col_idx}: value must not be blank."
                    )


def _find_qno_column(header_row: list[str]) -> int:
    for idx, value in enumerate(header_row, start=1):
        if value is None:
            continue
        if str(value).strip().lower() == "qno":
            return idx
    raise ValueError("Could not find 'Qno' column in Excel header.")


def _match_column(header_row: list[str], keyword_groups: list[tuple[str, ...]], friendly_name: str) -> int:
    normalized_headers = []
    for idx, value in enumerate(header_row, start=1):
        text = "" if value is None else str(value).lower()
        normalized_headers.append((idx, text))

    for keywords in keyword_groups:
        for idx, text in normalized_headers:
            if all(keyword in text for keyword in keywords):
                return idx
    raise ValueError(f"Unable to locate column for {friendly_name}. Please ensure the header contains {keyword_groups[0]}.")


def _find_magazine_column(header_row: list[str]) -> int:
    keyword_groups = [
        ("magazine", "edition"),
        ("magazine", "issue"),
        ("magazine",),
        ("edition",),
    ]
    return _match_column(header_row, keyword_groups, "Magazine Edition")


def _find_question_set_column(header_row: list[str]) -> int:
    keyword_groups = [
        ("question", "set"),
        ("question", "paper"),
        ("set", "name"),
        ("set",),
    ]
    return _match_column(header_row, keyword_groups, "Question Set")


def _find_high_level_chapter_column(header_row: list[str]) -> int:
    keyword_groups = [
        ("high", "level", "chapter"),
        ("high", "level"),
        ("chapter",),
    ]
    return _match_column(header_row, keyword_groups, "High Level Chapter")


def _find_question_set_name_column(header_row: list[str]) -> int:
    for idx, value in enumerate(header_row, start=1):
        if value is None:
            continue
        text = str(value).strip().lower()
        if text == "name of question set" or text == "name of the question set":
            return idx
    raise ValueError("Unable to locate 'Name of Question set' column.")


def _find_page_column(header_row: list[str]) -> int:
    keyword_groups = [
        ("page", "no"),
        ("page", "number"),
        ("page",),
        ("pg",),
    ]
    return _match_column(header_row, keyword_groups, "Page Number")


def _find_question_text_column(header_row: list[str]) -> int:
    for idx, value in enumerate(header_row, start=1):
        if value is None:
            continue
        text = str(value).strip().lower()
        if not text:
            continue
        if "question" in text and not any(
            keyword in text for keyword in ("set", "qno", "number", "no", "id")
        ):
            return idx
    raise ValueError("Unable to locate column containing question text.")


def _find_insert_row(worksheet) -> int:
    """Locate the next empty row by scanning for the last row that contains any value."""

    def _has_value(value) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True

    for row_idx in range(worksheet.max_row, 1, -1):
        for cell in worksheet[row_idx]:
            if _has_value(cell.value):
                return row_idx + 1
    return 2


def infer_column_types(worksheet, num_columns: int) -> dict[int, type]:
    column_types: dict[int, type] = {}
    for col_idx in range(1, num_columns + 1):
        for row_idx in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                column_types[col_idx] = type(value)
                break
    return column_types


def convert_value_for_column(value: str, target_type: type | None, header_row: list[str], col_idx: int):
    header_label = header_row[col_idx - 1] if col_idx - 1 < len(header_row) else f"Column {col_idx}"
    if target_type is None or target_type is str:
        return value

    stripped = value.strip()
    if target_type is bool:
        lowered = stripped.lower()
        if lowered in {"true", "1", "yes"}:
            return True
        if lowered in {"false", "0", "no"}:
            return False
        raise ValueError(f"Value '{value}' cannot be interpreted as boolean for column '{header_label}'.")
    if target_type is int:
        try:
            return int(stripped)
        except ValueError:
            try:
                return int(float(stripped))
            except ValueError as exc:
                raise ValueError(f"Value '{value}' cannot be interpreted as integer for column '{header_label}'.") from exc
    if target_type is float:
        try:
            return float(stripped)
        except ValueError as exc:
            raise ValueError(f"Value '{value}' cannot be interpreted as float for column '{header_label}'.") from exc
    if target_type is Decimal:
        try:
            return Decimal(stripped)
        except InvalidOperation as exc:
            raise ValueError(f"Value '{value}' cannot be interpreted as decimal for column '{header_label}'.") from exc
    if isinstance(target_type, type) and issubclass(target_type, dt.datetime):
        try:
            return dt.datetime.fromisoformat(stripped)
        except ValueError as exc:
            raise ValueError(f"Value '{value}' is not a valid datetime for column '{header_label}'.") from exc
    if isinstance(target_type, type) and issubclass(target_type, dt.date):
        try:
            return dt.date.fromisoformat(stripped)
        except ValueError as exc:
            raise ValueError(f"Value '{value}' is not a valid date for column '{header_label}'.") from exc

    return value


def _normalize_text(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", " ", value.lower())
    return " ".join(cleaned.split())


def _normalize_month_year(value: str) -> str:
    lower = value.lower()
    month = None
    for alias, number in MONTH_ALIASES.items():
        if alias in lower:
            month = number
            break

    year_match = re.search(r"(20\d{2}|19\d{2}|'\d{2})", lower)
    year = None
    if year_match:
        token = year_match.group(0)
        if token.startswith("'"):
            year = 2000 + int(token.strip("'"))
        else:
            year = int(token)

    if month and year:
        return f"{year:04d}-{month:02d}"
    if year and not month:
        return str(year)
    return _normalize_text(value)


def normalize_magazine_edition(value: str) -> str:
    if not value:
        return ""
    parts = value.split("|", 1)
    magazine_name = parts[0].strip()
    edition_part = parts[1].strip() if len(parts) > 1 else ""
    normalized_mag_name = _normalize_text(magazine_name)
    normalized_edition = _normalize_month_year(edition_part or magazine_name)
    return f"{normalized_mag_name}|{normalized_edition}"


def normalize_question_set(value: str) -> str:
    return _normalize_text(value)


def normalize_qno(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        try:
            return str(int(round(value)))
        except (TypeError, ValueError):
            pass
    text = str(value).strip()
    if text.isdigit():
        return str(int(text))
    cleaned = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return " ".join(cleaned.split())


def normalize_page(value) -> str:
    if value is None:
        return ""
    return _normalize_text(str(value))


class GroupingChapterListWidget(QListWidget):
    def __init__(self, parent_window):
        super().__init__()
        self.parent_window = parent_window
        self.setSelectionMode(QListWidget.SingleSelection)
        self.setDragEnabled(True)

    def startDrag(self, supportedActions) -> None:
        item = self.currentItem()
        if not item:
            return
        chapter = item.data(Qt.UserRole) or item.text()
        mime = QMimeData()
        mime.setText(chapter)
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.MoveAction)


class GroupListWidget(QListWidget):
    def __init__(self, parent_window):
        super().__init__()
        self.parent_window = parent_window
        self.setSelectionMode(QListWidget.SingleSelection)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)

    def dropEvent(self, event: QDropEvent) -> None:
        if not event.mimeData().hasText():
            super().dropEvent(event)
            return
        chapter = event.mimeData().text().strip()
        if not chapter:
            return
        position = event.position().toPoint() if hasattr(event, "position") else event.pos()
        target_item = self.itemAt(position)
        if not target_item:
            target_item = self.currentItem()
        if not target_item:
            return
        group = target_item.data(Qt.UserRole) or target_item.text()
        current_item = self.parent_window.group_list.currentItem()
        source_group = current_item.data(Qt.UserRole) if current_item else None
        self.parent_window.move_chapter_to_group(chapter, group, stay_on_group=source_group)
        event.acceptProposedAction()


class QuestionTableWidget(QTableWidget):
    MIME_TYPE = "application/x-question-row"

    def __init__(self, parent_window):
        super().__init__(0, 4)
        self.parent_window = parent_window
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragOnly)

    def startDrag(self, supportedActions) -> None:
        row = self.currentRow()
        if row < 0 or row >= len(self.parent_window.current_questions):
            return
        question = self.parent_window.current_questions[row]
        payload = json.dumps(
            {
                "row_number": question.get("row_number"),
                "qno": question.get("qno"),
                "question_set": question.get("question_set"),
                "group": question.get("group"),
            }
        )
        mime = QMimeData()
        mime.setData(self.MIME_TYPE, payload.encode("utf-8"))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.MoveAction)


class TagBadge(QLabel):
    """Colored tag badge that can be clicked to filter."""
    
    clicked = Signal(str)  # Emits tag name when clicked
    
    def __init__(self, tag: str, color: str, parent=None):
        super().__init__(parent)
        self.tag = tag
        self.setText(f" {tag} ")
        self.setStyleSheet(f"""
            QLabel {{
                background-color: {color};
                color: white;
                border-radius: 2px;
                padding: 1px 4px;
                font-size: 9px;
                font-weight: bold;
            }}
            QLabel:hover {{
                opacity: 0.8;
            }}
        """)
        self.setCursor(Qt.PointingHandCursor)
    
    def mousePressEvent(self, event):
        """Handle click to filter by this tag."""
        if event.button() == Qt.LeftButton:
            self.clicked.emit(self.tag)
        super().mousePressEvent(event)


class ClickableTagBadge(QPushButton):
    """Beautiful clickable tag badge with +/✓ toggle."""
    
    def __init__(self, tag: str, color: str, is_selected: bool = False, parent=None):
        super().__init__(parent)
        self.tag = tag
        self.color = color
        self.is_selected = is_selected
        self.setFixedHeight(32)
        self.setCursor(Qt.PointingHandCursor)
        self.clicked.connect(self._toggle_selection)
        self._update_style()
    
    def _toggle_selection(self):
        """Toggle selection state."""
        self.is_selected = not self.is_selected
        self._update_style()
    
    def _update_style(self):
        """Update button appearance based on selection state."""
        icon = "✓" if self.is_selected else "+"
        if self.is_selected:
            # Selected state - solid color
            self.setStyleSheet(f"""
                QPushButton {{
                    background-color: {self.color};
                    color: white;
                    border: 2px solid {self.color};
                    border-radius: 6px;
                    padding: 6px 12px;
                    font-size: 13px;
                    font-weight: bold;
                    text-align: left;
                }}
                QPushButton:hover {{
                    background-color: {self._darken_color(self.color)};
                    border-color: {self._darken_color(self.color)};
                }}
            """)
        else:
            # Unselected state - outline only
            self.setStyleSheet(f"""
                QPushButton {{
                    background-color: white;
                    color: {self.color};
                    border: 2px solid {self.color};
                    border-radius: 6px;
                    padding: 6px 12px;
                    font-size: 13px;
                    font-weight: bold;
                    text-align: left;
                }}
                QPushButton:hover {{
                    background-color: {self._lighten_color(self.color)};
                }}
            """)
        self.setText(f"{icon}  {self.tag}")
    
    def _darken_color(self, color: str) -> str:
        """Darken a hex color by 20%."""
        color = color.lstrip('#')
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
        r, g, b = int(r * 0.8), int(g * 0.8), int(b * 0.8)
        return f"#{r:02x}{g:02x}{b:02x}"
    
    def _lighten_color(self, color: str) -> str:
        """Lighten a hex color by adding 90% opacity."""
        return f"{color}20"  # Add alpha for transparency


class MultiSelectTagDialog(QDialog):
    """Beautiful dialog for selecting multiple tags with colored badges."""
    
    def __init__(self, existing_tags: list[str], selected_tags: list[str] = None, 
                 title: str = "Select Tags", tag_colors: dict[str, str] = None, 
                 available_colors: list[str] = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        
        self.tag_colors = tag_colors or {}
        self.available_colors = available_colors or ["#2563eb", "#10b981", "#f59e0b", "#ef4444", "#8b5cf6"]
        self.tag_badges = {}
        self.selected_tags_list = selected_tags or []
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Title and instruction
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b;")
        layout.addWidget(title_label)
        
        instruction = QLabel("Click on tags to select/deselect them:")
        instruction.setStyleSheet("font-size: 12px; color: #64748b; margin-bottom: 5px;")
        layout.addWidget(instruction)
        
        # Scrollable area for tag badges
        scroll_area = QWidget()
        self.tags_layout = QVBoxLayout(scroll_area)
        self.tags_layout.setSpacing(8)
        self.tags_layout.setAlignment(Qt.AlignTop)
        
        # Add existing tags as clickable badges
        if existing_tags:
            self._add_tag_badges(existing_tags)
        else:
            no_tags_label = QLabel("No existing tags. Create new ones below.")
            no_tags_label.setStyleSheet("color: #94a3b8; font-style: italic; padding: 20px;")
            self.tags_layout.addWidget(no_tags_label)
        
        layout.addWidget(scroll_area, 1)
        
        # Separator
        separator = QWidget()
        separator.setFixedHeight(1)
        separator.setStyleSheet("background-color: #e2e8f0;")
        layout.addWidget(separator)
        
        # Input field for new tags
        new_tag_label = QLabel("Create new tag:")
        new_tag_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #475569;")
        layout.addWidget(new_tag_label)
        
        new_tag_layout = QHBoxLayout()
        self.new_tag_input = QLineEdit()
        self.new_tag_input.setPlaceholderText("Type tag name and press Enter or click Add")
        self.new_tag_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #cbd5e1;
                border-radius: 6px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
            }
        """)
        self.new_tag_input.returnPressed.connect(self._add_new_tag)
        new_tag_layout.addWidget(self.new_tag_input)
        
        add_btn = QPushButton("Add")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 20px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        add_btn.clicked.connect(self._add_new_tag)
        new_tag_layout.addWidget(add_btn)
        layout.addLayout(new_tag_layout)
        
        # Dialog buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.button(QDialogButtonBox.Ok).setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        button_box.button(QDialogButtonBox.Cancel).setStyleSheet("""
            QPushButton {
                background-color: #f1f5f9;
                color: #475569;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                padding: 8px 24px;
                font-size: 13px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e2e8f0;
            }
        """)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def _add_tag_badges(self, tags: list[str]):
        """Add tag badges in a flowing layout."""
        row_layout = None
        tags_per_row = 0
        max_tags_per_row = 3
        
        for tag in sorted(tags):
            if tags_per_row == 0:
                row_layout = QHBoxLayout()
                row_layout.setSpacing(10)
                self.tags_layout.addLayout(row_layout)
            
            color = self._get_tag_color(tag)
            is_selected = tag in self.selected_tags_list
            badge = ClickableTagBadge(tag, color, is_selected)
            self.tag_badges[tag] = badge
            row_layout.addWidget(badge)
            
            tags_per_row += 1
            if tags_per_row >= max_tags_per_row:
                row_layout.addStretch()
                tags_per_row = 0
        
        if row_layout and tags_per_row > 0:
            row_layout.addStretch()
    
    def _get_tag_color(self, tag: str) -> str:
        """Get color for a tag."""
        if tag not in self.tag_colors:
            color_index = len(self.tag_colors) % len(self.available_colors)
            self.tag_colors[tag] = self.available_colors[color_index]
        return self.tag_colors[tag]
    
    def _add_new_tag(self):
        """Add a new tag to the list."""
        tag = self.new_tag_input.text().strip()
        if not tag:
            return
        
        # Check if tag already exists
        if tag in self.tag_badges:
            # Already exists, just select it
            self.tag_badges[tag].is_selected = True
            self.tag_badges[tag]._update_style()
            self.new_tag_input.clear()
            return
        
        # Add new tag badge
        color = self._get_tag_color(tag)
        badge = ClickableTagBadge(tag, color, is_selected=True)
        self.tag_badges[tag] = badge
        
        # Add to layout (create new row if needed)
        last_layout = None
        for i in range(self.tags_layout.count()):
            item = self.tags_layout.itemAt(i)
            if isinstance(item, QHBoxLayout):
                last_layout = item
        
        if not last_layout or last_layout.count() >= 4:  # 3 badges + 1 stretch
            new_row = QHBoxLayout()
            new_row.setSpacing(10)
            new_row.addWidget(badge)
            new_row.addStretch()
            self.tags_layout.addLayout(new_row)
        else:
            # Remove stretch and add badge
            stretch_item = last_layout.takeAt(last_layout.count() - 1)
            last_layout.addWidget(badge)
            if stretch_item:
                last_layout.addItem(stretch_item)
        
        self.new_tag_input.clear()
    
    def get_selected_tags(self) -> list[str]:
        """Get list of selected tags."""
        return [tag for tag, badge in self.tag_badges.items() if badge.is_selected]


class QuestionTreeWidget(QTreeWidget):
    MIME_TYPE = "application/x-question-row"

    def __init__(self, parent_window):
        super().__init__()
        self.parent_window = parent_window
        self.setColumnCount(4)
        self.setHeaderLabels(["Question No", "Page", "Question Set Name", "Magazine"])
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragOnly)
        self.setAlternatingRowColors(True)
        self.itemSelectionChanged.connect(self.parent_window.on_question_selected)
        self.header().setStretchLastSection(True)

    def startDrag(self, supportedActions) -> None:
        current = self.currentItem()
        if not current:
            return
        # Only allow dragging leaf items (actual questions)
        if current.childCount() > 0:
            return
        question_data = current.data(0, Qt.UserRole)
        if not question_data:
            return
        payload = json.dumps(
            {
                "row_number": question_data.get("row_number"),
                "qno": question_data.get("qno"),
                "question_set": question_data.get("question_set"),
                "group": question_data.get("group"),
            }
        )
        mime = QMimeData()
        mime.setData(self.MIME_TYPE, payload.encode("utf-8"))
        drag = QDrag(self)
        drag.setMimeData(mime)
        drag.exec(Qt.MoveAction)


class ChapterTableWidget(QTableWidget):
    def __init__(self, parent_window):
        super().__init__(0, 2)
        self.parent_window = parent_window
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setAcceptDrops(True)
        self.setDragDropMode(QAbstractItemView.DropOnly)
        self.setDragEnabled(False)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasFormat(QuestionTableWidget.MIME_TYPE):
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasFormat(QuestionTableWidget.MIME_TYPE):
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)

    def dropEvent(self, event: QDropEvent) -> None:
        if not event.mimeData().hasFormat(QuestionTableWidget.MIME_TYPE):
            super().dropEvent(event)
            return
        try:
            payload = bytes(event.mimeData().data(QuestionTableWidget.MIME_TYPE)).decode("utf-8")
            question = json.loads(payload)
        except Exception:
            return

        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        index = self.indexAt(pos)
        if not index.isValid():
            return
        chapter_item = self.item(index.row(), 0)
        if not chapter_item:
            return
        target_group = chapter_item.data(Qt.UserRole) or chapter_item.text()
        self.parent_window.reassign_question(question, target_group)
        event.acceptProposedAction()

def read_tsv_rows(tsv_path: Path) -> list[list[str]]:
    with tsv_path.open("r", encoding="utf-8", newline="") as tsv_file:
        reader = csv.reader(tsv_file, delimiter="\t")
        next(reader, None)  # Skip header
        return [row for row in reader]


def collect_existing_triplets(
    worksheet, magazine_col: int, qno_col: int, page_col: int
) -> dict[tuple[str, str, str], tuple[str, str, str]]:
    triplets: dict[tuple[str, str, str], tuple[str, str, str]] = {}
    for row_idx in range(2, worksheet.max_row + 1):
        magazine_value = worksheet.cell(row=row_idx, column=magazine_col).value
        if not magazine_value:
            continue
        qno_value = worksheet.cell(row=row_idx, column=qno_col).value
        page_value = worksheet.cell(row=row_idx, column=page_col).value
        normalized_magazine = normalize_magazine_edition(str(magazine_value))
        normalized_qno = normalize_qno(qno_value)
        normalized_page = normalize_page(page_value)
        if normalized_qno and normalized_page:
            key = (normalized_magazine, normalized_qno, normalized_page)
            triplets.setdefault(
                key,
                (
                    str(magazine_value),
                    str(qno_value) if qno_value is not None else "",
                    str(page_value) if page_value is not None else "",
                ),
            )
    return triplets


def extract_file_metadata(
    rows: list[list[str]],
    magazine_col: int,
    question_set_col: int,
    qno_col: int,
    page_col: int,
) -> tuple[str, list[tuple[str, str, str, str, str, str, str]]]:
    magazine_identifier = None
    row_signatures: list[tuple[str, str, str, str, str, str, str]] = []
    seen_row_signatures: set[tuple[str, str, str]] = set()
    for row in rows:
        # Guard against short rows
        required_columns = max(magazine_col, question_set_col, qno_col, page_col)
        if len(row) < required_columns:
            raise ValueError("TSV row does not contain all required columns.")
        magazine_value = row[magazine_col - 1].strip()
        if not magazine_value:
            raise ValueError("Magazine edition must be provided for every row.")
        normalized_magazine = normalize_magazine_edition(magazine_value)
        if magazine_identifier is None:
            magazine_identifier = normalized_magazine
        elif magazine_identifier != normalized_magazine:
            raise ValueError(
                "All rows in the TSV must belong to the same magazine edition. "
                "Please split files by edition before importing."
            )

        question_value = row[question_set_col - 1].strip()
        if not question_value:
            raise ValueError("Question set must be provided for every row.")

        qno_value = row[qno_col - 1].strip()
        if not qno_value:
            raise ValueError("Question number must be provided for every row.")
        normalized_qno = normalize_qno(qno_value)
        if not normalized_qno:
            raise ValueError(f"Unable to normalize question number '{qno_value}'.")

        page_value = row[page_col - 1].strip()
        if not page_value:
            raise ValueError("Page number must be provided for every row.")
        normalized_page = normalize_page(page_value)
        if not normalized_page:
            raise ValueError(f"Unable to normalize page number '{page_value}'.")

        combo_signature = (magazine_identifier, normalized_qno, normalized_page)
        if combo_signature in seen_row_signatures:
            raise ValueError(
                "Duplicate question/page detected within TSV for magazine edition "
                f"'{magazine_value}', question number '{qno_value}', page '{page_value}'."
            )
        seen_row_signatures.add(combo_signature)
        row_signatures.append(
            (
                normalized_magazine,
                normalized_qno,
                normalized_page,
                magazine_value,
                qno_value,
                page_value,
            )
        )

    if magazine_identifier is None:
        raise ValueError("Unable to identify magazine edition in the TSV file.")

    return magazine_identifier, row_signatures


def append_rows_to_excel(
    workbook_path: Path,
    worksheet,
    header_row: list[str],
    column_types: dict[int, type],
    rows: list[list[str]],
    insert_row: int,
    page_col: int,
) -> tuple[str, str]:
    """Append rows to Excel and return (status_message, page_range)."""
    appended_rows = 0
    page_numbers = []
    
    for row in rows:
        # Extract page number for tracking
        if page_col - 1 < len(row):
            page_value = row[page_col - 1].strip()
            if page_value:
                page_numbers.append(page_value)
        
        for col_idx, value in enumerate(row, start=1):
            target_type = column_types.get(col_idx)
            converted = convert_value_for_column(value, target_type, header_row, col_idx)
            worksheet.cell(row=insert_row, column=col_idx, value=converted)
        insert_row += 1
        appended_rows += 1

    worksheet.parent.save(workbook_path)
    
    # Calculate page range
    page_range = "N/A"
    if page_numbers:
        try:
            # Try to extract numeric values for range
            numeric_pages = []
            for page in page_numbers:
                # Extract first number from page string
                match = re.search(r'\d+', page)
                if match:
                    numeric_pages.append(int(match.group()))
            
            if numeric_pages:
                min_page = min(numeric_pages)
                max_page = max(numeric_pages)
                if min_page == max_page:
                    page_range = str(min_page)
                else:
                    page_range = f"{min_page}-{max_page}"
            else:
                # If no numeric pages, just use first and last
                if len(page_numbers) == 1:
                    page_range = page_numbers[0]
                else:
                    page_range = f"{page_numbers[0]}-{page_numbers[-1]}"
        except Exception:
            # Fallback to simple first-last format
            if len(page_numbers) == 1:
                page_range = page_numbers[0]
            else:
                page_range = f"{page_numbers[0]}-{page_numbers[-1]}"
    
    status_msg = f"Appended {appended_rows} rows to '{worksheet.title}'"
    return status_msg, page_range


def process_tsv(tsv_path: Path, workbook_path: Path) -> str:
    try:
        validate_tsv(tsv_path)
        rows = read_tsv_rows(tsv_path)

        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        header_row = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if not header_row:
            raise ValueError("Worksheet header row is empty.")

        qno_column = _find_qno_column(header_row)
        magazine_col = _find_magazine_column(header_row)
        question_set_col = _find_question_set_column(header_row)
        page_col = _find_page_column(header_row)
        insert_row = _find_insert_row(worksheet)

        column_types = infer_column_types(worksheet, len(header_row))
        existing_triplets = collect_existing_triplets(worksheet, magazine_col, qno_column, page_col)

        magazine_identifier, row_signatures = extract_file_metadata(
            rows, magazine_col, question_set_col, qno_column, page_col
        )

        duplicates = []
        for normalized_magazine, normalized_qno, normalized_page, original_mag, original_qno, original_page in row_signatures:
            combo = (normalized_magazine, normalized_qno, normalized_page)
            if combo in existing_triplets:
                existing_mag, existing_qno, existing_page = existing_triplets[combo]
                duplicates.append(
                    (
                        original_mag or existing_mag,
                        original_qno,
                        original_page,
                        existing_qno,
                        existing_page,
                    )
                )

        if duplicates:
            readable = "; ".join(
                f"Magazine '{mag}' Question '{qno}' Page '{page}' already exists (Workbook has Qno '{ex_qno}', Page '{ex_page}')"
                for mag, qno, page, ex_qno, ex_page in duplicates
            )
            raise ValueError(
                "Duplicate questions detected: "
                f"{readable}. Remove or update these entries before importing."
            )

        status_message, page_range = append_rows_to_excel(
            workbook_path=workbook_path,
            worksheet=worksheet,
            header_row=header_row,
            column_types=column_types,
            rows=rows,
            insert_row=insert_row,
            page_col=page_col,
        )
        
        # Add page range to status message
        status_message = f"{status_message}, Pages: {page_range}"
        
    except ValueError as exc:
        tsv_path.unlink(missing_ok=True)
        raise ValueError(f"{exc} (file removed due to validation failure)") from exc

    tsv_path.unlink()
    return status_message + " (file removed)"


class TSVWatcherWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("TSV to Excel Watcher")
        self.resize(1200, 820)

        self.event_queue: queue.Queue[tuple] = queue.Queue()
        self.watch_thread: threading.Thread | None = None
        self.stop_event = threading.Event()
        self.file_rows: dict[str, int] = {}
        self.file_errors: dict[str, str] = {}
        self.metrics_request_id = 0
        self.chapter_questions: dict[str, list[dict[str, str]]] = {}
        self.current_questions: list[dict[str, str]] = []
        self.all_questions: list[dict[str, str]] = []  # Unfiltered questions
        self.question_set_search_term: str = ""
        self.magazine_search_term: str = ""
        self.tag_filter_term: str = ""
        self.selected_tag_filters: list[str] = []  # Multiple selected tags for filtering
        self.current_magazine_name: str = ""  # Track current magazine for grouping
        self.canonical_chapters: list[str] = []
        self.chapter_lookup: dict[str, str] = {}
        self.chapter_groups: dict[str, list[str]] = {}
        self.current_workbook_path: Path | None = None
        self.high_level_column_index: int | None = None
        self.question_lists: dict[str, list[dict]] = {}  # name -> list of questions
        self.current_list_name: str | None = None
        self.current_list_questions: list[dict] = []  # Questions in currently selected list
        self.group_tags: dict[str, list[str]] = {}  # group_key -> list of tags
        self.tag_colors: dict[str, str] = {}  # tag -> color
        
        # Predefined color palette for tags (20 colors)
        self.available_tag_colors = [
            "#2563eb",  # Blue
            "#10b981",  # Green
            "#f59e0b",  # Amber
            "#ef4444",  # Red
            "#8b5cf6",  # Purple
            "#06b6d4",  # Cyan
            "#ec4899",  # Pink
            "#14b8a6",  # Teal
            "#f97316",  # Orange
            "#6366f1",  # Indigo
            "#84cc16",  # Lime
            "#f43f5e",  # Rose
            "#0ea5e9",  # Sky Blue
            "#a855f7",  # Violet
            "#22c55e",  # Green Light
            "#eab308",  # Yellow
            "#d946ef",  # Fuchsia
            "#3b82f6",  # Blue Light
            "#fb923c",  # Orange Light
            "#38bdf8",  # Sky
        ]

        # Ensure QuestionList directory exists
        QUESTION_LIST_DIR.mkdir(exist_ok=True)
        
        self._load_group_tags()

        self._build_ui()
        self._load_last_selection()
        self._setup_timer()
        # Delay initial load to ensure UI is ready and timer is running
        QTimer.singleShot(100, self.update_row_count)

    def _build_ui(self) -> None:
        self._apply_palette()
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(18, 18, 18, 18)
        root_layout.setSpacing(14)

        top_card = self._create_card()
        top_layout = QVBoxLayout(top_card)
        top_layout.setSpacing(10)

        self.output_edit = QLineEdit()
        self.output_edit.editingFinished.connect(self.update_row_count)
        output_row = QHBoxLayout()
        output_row.addWidget(self._create_label("Workbook"))
        output_row.addWidget(self.output_edit)
        browse_output = QPushButton("Browse…")
        browse_output.clicked.connect(self.select_output_file)
        output_row.addWidget(browse_output)
        top_layout.addLayout(output_row)

        info_row = QHBoxLayout()
        self.row_count_label = QLabel("Total rows: N/A")
        self.row_count_label.setObjectName("headerLabel")
        self.mag_summary_label = QLabel("Magazines: N/A")
        self.mag_summary_label.setObjectName("infoLabel")
        self.mag_missing_label = QLabel("Missing ranges: N/A")
        self.mag_missing_label.setObjectName("infoLabel")
        info_row.addWidget(self.row_count_label)
        info_row.addStretch()
        info_row.addWidget(self.mag_summary_label)
        info_row.addWidget(self.mag_missing_label)
        top_layout.addLayout(info_row)
        root_layout.addWidget(top_card)

        tab_widget = QTabWidget()
        root_layout.addWidget(tab_widget, 1)

        qa_tab = QWidget()
        qa_layout = QVBoxLayout(qa_tab)
        qa_tabs = QTabWidget()
        qa_layout.addWidget(qa_tabs)

        magazine_tab = QWidget()
        magazine_tab_layout = QVBoxLayout(magazine_tab)
        
        # Top summary card
        mag_summary_card = self._create_card()
        mag_summary_layout = QHBoxLayout(mag_summary_card)
        self.mag_total_editions_label = QLabel("Total Editions: 0")
        self.mag_total_editions_label.setObjectName("headerLabel")
        self.mag_total_sets_label = QLabel("Question Sets: 0")
        self.mag_total_sets_label.setObjectName("infoLabel")
        mag_summary_layout.addWidget(self.mag_total_editions_label)
        mag_summary_layout.addStretch()
        mag_summary_layout.addWidget(self.mag_total_sets_label)
        magazine_tab_layout.addWidget(mag_summary_card)
        
        mag_split = QSplitter(Qt.Horizontal)
        magazine_tab_layout.addWidget(mag_split, 1)
        
        # Left side - Magazine tree with search
        mag_tree_card = self._create_card()
        mag_tree_layout = QVBoxLayout(mag_tree_card)
        mag_tree_layout.addWidget(self._create_label("Magazine Editions"))
        
        # Search box for magazine tree
        mag_search_layout = QHBoxLayout()
        mag_search_layout.addWidget(QLabel("Search:"))
        self.mag_tree_search = QLineEdit()
        self.mag_tree_search.setPlaceholderText("Filter magazines or editions...")
        self.mag_tree_search.textChanged.connect(self.on_mag_tree_search_changed)
        mag_search_layout.addWidget(self.mag_tree_search)
        clear_mag_search_btn = QPushButton("Clear")
        clear_mag_search_btn.setMaximumWidth(80)
        clear_mag_search_btn.clicked.connect(lambda: self.mag_tree_search.clear())
        mag_search_layout.addWidget(clear_mag_search_btn)
        mag_tree_layout.addLayout(mag_search_layout)
        
        self.mag_tree = QTreeWidget()
        self.mag_tree.setColumnCount(3)
        self.mag_tree.setHeaderLabels(["Magazine", "Edition", "Missing Ranges"])
        self.mag_tree.setRootIsDecorated(False)
        self.mag_tree.setAlternatingRowColors(True)
        self.mag_tree.itemSelectionChanged.connect(self.on_magazine_select)
        self.mag_tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.mag_tree.header().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.mag_tree.header().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        mag_tree_layout.addWidget(self.mag_tree)
        mag_split.addWidget(mag_tree_card)

        # Right side - Question sets detail
        detail_card = self._create_card()
        detail_layout = QVBoxLayout(detail_card)
        detail_layout.addWidget(self._create_label("Question Sets"))
        self.question_label = QLabel("Select an edition to view question sets")
        self.question_label.setObjectName("infoLabel")
        self.question_label.setStyleSheet(
            "padding: 8px; background-color: #f1f5f9; border-radius: 6px; color: #475569;"
        )
        detail_layout.addWidget(self.question_label)
        self.question_list = QListWidget()
        self.question_list.setAlternatingRowColors(True)
        self.question_list.setStyleSheet(
            """
            QListWidget {
                font-size: 13px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #e2e8f0;
            }
            QListWidget::item:hover {
                background-color: #f8fafc;
            }
            """
        )
        detail_layout.addWidget(self.question_list)
        mag_split.addWidget(detail_card)
        
        mag_split.setSizes([500, 400])
        qa_tabs.addTab(magazine_tab, "Magazine Editions")

        questions_tab = QWidget()
        questions_tab_layout = QVBoxLayout(questions_tab)
        analysis_card = self._create_card()
        questions_tab_layout.addWidget(analysis_card)
        analysis_layout = QVBoxLayout(analysis_card)
        analysis_split = QSplitter(Qt.Horizontal)
        analysis_layout.addWidget(analysis_split, 1)

        chapter_card = self._create_card()
        chapter_layout = QVBoxLayout(chapter_card)
        chapter_layout.addWidget(self._create_label("Chapters"))
        self.chapter_table = ChapterTableWidget(self)
        self.chapter_table.setHorizontalHeaderLabels(["Chapter", "Questions"])
        self.chapter_table.horizontalHeader().setStretchLastSection(False)
        self.chapter_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.chapter_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.chapter_table.verticalHeader().setVisible(False)
        self.chapter_table.itemSelectionChanged.connect(self.on_chapter_selected)
        chapter_layout.addWidget(self.chapter_table)
        analysis_split.addWidget(chapter_card)

        question_card = self._create_card()
        question_layout = QVBoxLayout(question_card)
        question_layout.addWidget(self._create_label("Questions"))
        
        # Add search controls
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        
        search_layout.addWidget(QLabel("Question Set:"))
        self.question_set_search = QLineEdit()
        self.question_set_search.setPlaceholderText("Type to search...")
        self.question_set_search.textChanged.connect(self.on_question_set_search_changed)
        search_layout.addWidget(self.question_set_search)
        
        search_layout.addWidget(QLabel("Tags:"))
        self.tag_filter_label = QLabel("None")
        self.tag_filter_label.setStyleSheet("padding: 4px; background-color: #f1f5f9; border-radius: 4px; min-width: 100px;")
        search_layout.addWidget(self.tag_filter_label)
        self.tag_filter_btn = QPushButton("Select Tags")
        self.tag_filter_btn.clicked.connect(self._show_tag_filter_dialog)
        search_layout.addWidget(self.tag_filter_btn)
        
        search_layout.addWidget(QLabel("Magazine:"))
        self.magazine_search = QLineEdit()
        self.magazine_search.setPlaceholderText("Type to search...")
        self.magazine_search.textChanged.connect(self.on_magazine_search_changed)
        search_layout.addWidget(self.magazine_search)
        
        clear_search_btn = QPushButton("Clear Search")
        clear_search_btn.clicked.connect(self.clear_question_search)
        search_layout.addWidget(clear_search_btn)
        search_layout.addStretch()
        
        question_layout.addLayout(search_layout)
        
        # Add question list controls
        list_control_layout = QHBoxLayout()
        add_to_list_btn = QPushButton("Add Selected to List")
        add_to_list_btn.clicked.connect(self.add_selected_to_list)
        list_control_layout.addWidget(add_to_list_btn)
        list_control_layout.addStretch()
        question_layout.addLayout(list_control_layout)
        
        self.question_tree = QuestionTreeWidget(self)
        self.question_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.question_tree.customContextMenuRequested.connect(self._show_group_context_menu)
        question_splitter = QSplitter(Qt.Vertical)
        question_splitter.addWidget(self.question_tree)
        self.question_text_view = QTextEdit()
        self.question_text_view.setReadOnly(True)
        self.question_text_view.setAcceptRichText(True)
        question_splitter.addWidget(self.question_text_view)
        question_splitter.setStretchFactor(0, 3)
        question_splitter.setStretchFactor(1, 1)
        question_splitter.setSizes([400, 120])
        question_layout.addWidget(question_splitter)
        analysis_split.addWidget(question_card)
        qa_tabs.addTab(questions_tab, "Question List")

        grouping_tab = QWidget()
        grouping_layout = QVBoxLayout(grouping_tab)
        grouping_card = self._create_card()
        grouping_layout.addWidget(grouping_card)
        grouping_card_layout = QHBoxLayout(grouping_card)

        self.group_list = GroupListWidget(self)
        self.group_list.itemSelectionChanged.connect(self.on_group_selected)
        grouping_card_layout.addWidget(self.group_list, 1)

        self.group_chapter_list = GroupingChapterListWidget(self)
        grouping_card_layout.addWidget(self.group_chapter_list, 2)

        group_controls = QVBoxLayout()
        group_controls.addWidget(self._create_label("Move chapter to group"))
        self.move_target_combo = QComboBox()
        group_controls.addWidget(self.move_target_combo)
        move_button = QPushButton("Move Chapter")
        move_button.clicked.connect(self.move_selected_chapter)
        group_controls.addWidget(move_button)
        group_controls.addStretch()
        grouping_card_layout.addLayout(group_controls)
        qa_tabs.addTab(grouping_tab, "Chapter Grouping")
        
        # Question Lists Tab
        lists_tab = QWidget()
        lists_tab_layout = QVBoxLayout(lists_tab)
        lists_card = self._create_card()
        lists_tab_layout.addWidget(lists_card)
        lists_card_layout = QVBoxLayout(lists_card)
        
        lists_header_layout = QHBoxLayout()
        lists_card_layout.addWidget(self._create_label("Custom Lists"))
        lists_header_layout.addStretch()
        lists_card_layout.addLayout(lists_header_layout)
        
        lists_split = QSplitter(Qt.Horizontal)
        lists_card_layout.addWidget(lists_split, 1)
        
        # Left side - list of saved lists
        saved_lists_card = self._create_card()
        saved_lists_layout = QVBoxLayout(saved_lists_card)
        saved_lists_layout.addWidget(self._create_label("Saved Lists"))
        
        list_controls_layout = QHBoxLayout()
        new_list_btn = QPushButton("New List")
        new_list_btn.clicked.connect(self.create_new_question_list)
        rename_list_btn = QPushButton("Rename")
        rename_list_btn.clicked.connect(self.rename_question_list)
        delete_list_btn = QPushButton("Delete")
        delete_list_btn.clicked.connect(self.delete_question_list)
        list_controls_layout.addWidget(new_list_btn)
        list_controls_layout.addWidget(rename_list_btn)
        list_controls_layout.addWidget(delete_list_btn)
        saved_lists_layout.addLayout(list_controls_layout)
        
        self.saved_lists_widget = QListWidget()
        self.saved_lists_widget.itemSelectionChanged.connect(self.on_saved_list_selected)
        saved_lists_layout.addWidget(self.saved_lists_widget)
        lists_split.addWidget(saved_lists_card)
        
        # Right side - questions in selected list
        list_questions_card = self._create_card()
        list_questions_layout = QVBoxLayout(list_questions_card)
        self.list_name_label = QLabel("Select a list to view questions")
        self.list_name_label.setObjectName("headerLabel")
        list_questions_layout.addWidget(self.list_name_label)
        
        remove_from_list_btn = QPushButton("Remove Selected from List")
        remove_from_list_btn.clicked.connect(self.remove_selected_from_list)
        list_questions_layout.addWidget(remove_from_list_btn)
        
        list_question_splitter = QSplitter(Qt.Vertical)
        self.list_question_table = QTableWidget(0, 4)
        self.list_question_table.setHorizontalHeaderLabels(["Question No", "Page", "Question Set Name", "Magazine"])
        self.list_question_table.horizontalHeader().setStretchLastSection(True)
        self.list_question_table.verticalHeader().setVisible(False)
        self.list_question_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.list_question_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.list_question_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.list_question_table.itemSelectionChanged.connect(self.on_list_question_selected)
        list_question_splitter.addWidget(self.list_question_table)
        
        self.list_question_text_view = QTextEdit()
        self.list_question_text_view.setReadOnly(True)
        self.list_question_text_view.setAcceptRichText(True)
        list_question_splitter.addWidget(self.list_question_text_view)
        list_question_splitter.setStretchFactor(0, 3)
        list_question_splitter.setStretchFactor(1, 1)
        list_question_splitter.setSizes([400, 120])
        
        list_questions_layout.addWidget(list_question_splitter)
        lists_split.addWidget(list_questions_card)
        
        qa_tabs.addTab(lists_tab, "Custom Lists")
        tab_widget.addTab(qa_tab, "Question Analysis")

        import_tab = QWidget()
        import_layout = QVBoxLayout(import_tab)

        import_form_card = self._create_card()
        import_form_layout = QVBoxLayout(import_form_card)
        self.input_edit = QLineEdit()
        self.input_edit.editingFinished.connect(self.refresh_file_list)
        input_row = QHBoxLayout()
        input_row.addWidget(self._create_label("Input folder"))
        input_row.addWidget(self.input_edit)
        browse_input = QPushButton("Browse…")
        browse_input.clicked.connect(self.select_input_folder)
        input_row.addWidget(browse_input)
        import_form_layout.addLayout(input_row)

        control_layout = QHBoxLayout()
        self.refresh_button = QPushButton("Refresh Files")
        self.refresh_button.clicked.connect(self.refresh_file_list)
        self.start_button = QPushButton("Start Watching")
        self.start_button.clicked.connect(self.start_watching)
        self.stop_button = QPushButton("Stop Watching")
        self.stop_button.clicked.connect(self.stop_watching)
        control_layout.addWidget(self.refresh_button)
        control_layout.addWidget(self.start_button)
        control_layout.addWidget(self.stop_button)
        control_layout.addStretch()
        import_form_layout.addLayout(control_layout)
        import_layout.addWidget(import_form_card)

        status_card = self._create_card()
        status_layout = QVBoxLayout(status_card)
        status_layout.addWidget(self._create_label("Upload Status"))
        self.file_table = QTableWidget(0, 3)
        self.file_table.setHorizontalHeaderLabels(["File", "Status", "Message"])
        self.file_table.horizontalHeader().setStretchLastSection(True)
        self.file_table.verticalHeader().setVisible(False)
        self.file_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.file_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.file_table.cellDoubleClicked.connect(self.on_file_double_clicked)
        status_layout.addWidget(self.file_table)
        import_layout.addWidget(status_card)

        tab_widget.addTab(import_tab, "Data Import")

        self.log_toggle = QPushButton("Show Log")
        self.log_toggle.setCheckable(True)
        self.log_toggle.toggled.connect(self.toggle_log_visibility)
        root_layout.addWidget(self.log_toggle, 0, alignment=Qt.AlignLeft)

        self.log_card = self._create_card()
        log_layout = QVBoxLayout(self.log_card)
        log_layout.addWidget(self._create_label("Log"))
        self.log_view = QTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setObjectName("logView")
        log_layout.addWidget(self.log_view)
        root_layout.addWidget(self.log_card, 0)
        self.log_card.setVisible(False)
        self._refresh_grouping_ui()
        self._load_saved_question_lists()

    def _apply_palette(self) -> None:
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor("#f4f5f9"))
        palette.setColor(QPalette.Base, QColor("#ffffff"))
        palette.setColor(QPalette.Text, QColor("#0f172a"))
        palette.setColor(QPalette.Button, QColor("#2563eb"))
        palette.setColor(QPalette.ButtonText, QColor("#ffffff"))
        self.setPalette(palette)
        self.setStyleSheet(
            """
            QWidget#card {
                background-color: #ffffff;
                border-radius: 14px;
                padding: 12px;
            }
            QPushButton {
                background-color: #2563eb;
                color: #ffffff;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
            }
            QPushButton:disabled {
                background-color: #94a3b8;
            }
            QLabel#headerLabel {
                font-weight: 600;
                color: #0f172a;
            }
            QLabel#infoLabel {
                color: #475569;
            }
            QTreeWidget, QTableWidget, QListWidget {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                background-color: #ffffff;
                color: #0f172a;
                alternate-background-color: #f8fafc;
            }
            QTreeWidget::item:selected, QTableWidget::item:selected, QListWidget::item:selected {
                background-color: #d0e2ff;
                color: #0f172a;
            }
            QTextEdit#logView {
                background-color: #0f172a;
                color: #e2e8f0;
                border-radius: 8px;
                padding: 8px;
                font-family: Consolas, 'Courier New', monospace;
            }
            """
        )

    def _create_card(self) -> QWidget:
        card = QWidget()
        card.setObjectName("card")
        return card

    def _create_label(self, text: str) -> QLabel:
        label = QLabel(text)
        label.setObjectName("headerLabel")
        return label

    def toggle_log_visibility(self, checked: bool) -> None:
        self.log_card.setVisible(checked)
        self.log_toggle.setText("Hide Log" if checked else "Show Log")

    def _load_last_selection(self) -> None:
        if not LAST_SELECTION_FILE.exists():
            return
        try:
            data = json.loads(LAST_SELECTION_FILE.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return
        workbook_path = data.get("workbook_path", "")
        if workbook_path:
            self.output_edit.setText(workbook_path)
        input_folder = data.get("input_folder", "")
        if input_folder:
            self.input_edit.setText(input_folder)

    def _save_last_selection(self) -> None:
        path_text = self.output_edit.text().strip()
        input_folder = self.input_edit.text().strip()
        if not path_text:
            return
        payload = {
            "workbook_path": path_text,
            "input_folder": input_folder if input_folder else "",
        }
        LAST_SELECTION_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _load_group_tags(self) -> None:
        """Load group tags from tags.cfg file."""
        if not TAGS_CONFIG_FILE.exists():
            return
        try:
            data = json.loads(TAGS_CONFIG_FILE.read_text(encoding="utf-8"))
            self.group_tags = data.get("group_tags", {})
            self.tag_colors = data.get("tag_colors", {})
        except json.JSONDecodeError:
            self.group_tags = {}
            self.tag_colors = {}

    def _save_group_tags(self) -> None:
        """Save group tags to tags.cfg file."""
        payload = {
            "group_tags": self.group_tags,
            "tag_colors": self.tag_colors,
        }
        TAGS_CONFIG_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    def _get_or_assign_tag_color(self, tag: str) -> str:
        """Get existing color for tag or assign a new one."""
        if tag not in self.tag_colors:
            # Cycle through available colors
            color_index = len(self.tag_colors) % len(self.available_tag_colors)
            self.tag_colors[tag] = self.available_tag_colors[color_index]
        return self.tag_colors[tag]

    def _load_canonical_chapters(self, grouping_file: Path) -> list[str]:
        """Load canonical chapters from the grouping JSON file."""
        if not grouping_file.exists():
            return []
        try:
            data = json.loads(grouping_file.read_text(encoding="utf-8"))
            return data.get("canonical_order", [])
        except json.JSONDecodeError:
            return []

    def _load_chapter_grouping(self, grouping_file: Path) -> dict[str, list[str]]:
        """Load chapter grouping from the specified JSON file."""
        if grouping_file.exists():
            try:
                data = json.loads(grouping_file.read_text(encoding="utf-8"))
                groups = data.get("groups", {})
            except json.JSONDecodeError:
                groups = {}
        else:
            groups = {}
        for group in self.canonical_chapters:
            groups.setdefault(group, [])
        groups.setdefault("Others", [])
        # Remove duplicates while preserving order
        for key, values in list(groups.items()):
            seen = set()
            unique = []
            for value in values:
                if value not in seen:
                    unique.append(value)
                    seen.add(value)
            groups[key] = unique
        self._rebuild_chapter_lookup(groups)
        return groups

    def _reload_grouping_for_magazine(self, magazine_name: str) -> None:
        """Reload chapter grouping based on magazine name."""
        if magazine_name == self.current_magazine_name:
            return  # Already loaded
        
        grouping_file = MAGAZINE_GROUPING_MAP.get(magazine_name, PHYSICS_GROUPING_FILE)
        self.current_magazine_name = magazine_name
        self.canonical_chapters = self._load_canonical_chapters(grouping_file)
        self.chapter_groups = self._load_chapter_grouping(grouping_file)
        self.log(f"Loaded chapter grouping for: {magazine_name or 'default (Physics)'}")

    def _save_chapter_grouping(self) -> None:
        """Save chapter grouping to the appropriate file based on current magazine."""
        data = {
            "canonical_order": self.canonical_chapters,
            "groups": self.chapter_groups,
        }
        grouping_file = MAGAZINE_GROUPING_MAP.get(self.current_magazine_name, PHYSICS_GROUPING_FILE)
        grouping_file.write_text(json.dumps(data, indent=2), encoding="utf-8")
        self._rebuild_chapter_lookup(self.chapter_groups)

    def _ordered_groups(self) -> list[str]:
        order = list(self.canonical_chapters)
        if "Others" not in order:
            order.append("Others")
        return order

    def _rebuild_chapter_lookup(self, groups: dict[str, list[str]]) -> None:
        lookup: dict[str, str] = {}
        for group, values in groups.items():
            norm_group = self._normalize_label(group)
            if norm_group:
                lookup[norm_group] = group
            for value in values:
                norm_value = self._normalize_label(value)
                if norm_value and norm_value not in lookup:
                    lookup[norm_value] = group
        self.chapter_lookup = lookup

    def _refresh_grouping_ui(self) -> None:
        if not hasattr(self, "group_list"):
            return
        self.group_list.blockSignals(True)
        self.group_list.clear()
        for group in self._ordered_groups():
            chapters = self.chapter_groups.get(group, [])
            item = QListWidgetItem(f"{group} ({len(chapters)})")
            item.setData(Qt.UserRole, group)
            self.group_list.addItem(item)
        self.group_list.blockSignals(False)
        if hasattr(self, "move_target_combo"):
            self.move_target_combo.clear()
            self.move_target_combo.addItems(self._ordered_groups())
        if hasattr(self, "group_chapter_list"):
            self.group_chapter_list.clear()

    def _select_group_in_ui(self, group: str) -> None:
        if not hasattr(self, "group_list"):
            return
        for row in range(self.group_list.count()):
            item = self.group_list.item(row)
            if (item.data(Qt.UserRole) or item.text()).startswith(group):
                self.group_list.setCurrentRow(row)
                break

    def _setup_timer(self) -> None:
        self.queue_timer = QTimer(self)
        self.queue_timer.setInterval(200)
        self.queue_timer.timeout.connect(self._process_queue)
        self.queue_timer.start()

    def update_row_count(self) -> None:
        workbook_path = Path(self.output_edit.text().strip())
        if not workbook_path.is_file():
            self._clear_all_question_data()
            self.row_count_label.setText("Total rows: N/A")
            self._set_magazine_summary("Magazines: N/A", "Missing ranges: N/A")
            self.question_label.setText("Select a workbook to display magazine editions.")
            self.current_workbook_path = None
            self.high_level_column_index = None
            return

        # Clear all existing data before loading new workbook
        self._clear_all_question_data()
        
        self.current_workbook_path = workbook_path
        self._save_last_selection()
        self.row_count_label.setText("Total rows: Loading...")
        self._set_magazine_summary("Magazines: Loading...", "Tracked editions: Loading...")
        self.question_label.setText("Loading editions...")
        self.metrics_request_id += 1
        request_id = self.metrics_request_id

        def worker(path: Path, req_id: int) -> None:
            try:
                # Pandas is used here to keep the UI responsive while gathering workbook metrics.
                df = pd.read_excel(path, sheet_name=0, dtype=object)
                row_count = self._compute_row_count_from_df(df)
                magazine_details, warnings = self._collect_magazine_details(df)
                
                # Detect magazine and load appropriate grouping BEFORE analyzing questions
                detected_magazine = self._detect_magazine_name(magazine_details)
                if detected_magazine != self.current_magazine_name:
                    grouping_file = MAGAZINE_GROUPING_MAP.get(detected_magazine, PHYSICS_GROUPING_FILE)
                    self.current_magazine_name = detected_magazine
                    self.canonical_chapters = self._load_canonical_chapters(grouping_file)
                    self.chapter_groups = self._load_chapter_grouping(grouping_file)
                
                # Now analyze questions with proper grouping loaded
                chapter_data, qa_warnings, question_col, raw_chapter_inputs = self._collect_question_analysis_data(df)
                warnings.extend(qa_warnings)
                self.event_queue.put(
                    (
                        "metrics",
                        req_id,
                        row_count,
                        magazine_details,
                        warnings,
                        chapter_data,
                        question_col,
                        raw_chapter_inputs,
                        detected_magazine,  # Pass magazine name to queue for logging
                    )
                )
            except Exception as exc:
                self.event_queue.put(("metrics_error", req_id, str(exc)))

        threading.Thread(target=worker, args=(workbook_path, request_id), daemon=True).start()

    def _set_magazine_summary(self, primary: str, secondary: str) -> None:
        self.mag_summary_label.setText(primary)
        self.mag_missing_label.setText(secondary)

    def _clear_all_question_data(self) -> None:
        """Clear all question analysis data and UI elements."""
        # Clear data structures
        self.chapter_questions.clear()
        self.current_questions.clear()
        self.all_questions.clear()
        self.question_set_search_term = ""
        self.magazine_search_term = ""
        
        # Clear UI elements
        self._populate_magazine_tree([])
        self._populate_question_sets([])
        self._populate_chapter_list({})
        
        # Clear search boxes
        if hasattr(self, "question_set_search"):
            self.question_set_search.clear()
        if hasattr(self, "magazine_search"):
            self.magazine_search.clear()
        if hasattr(self, "mag_tree_search"):
            self.mag_tree_search.clear()

    def _detect_magazine_name(self, details: list[dict]) -> str:
        """Detect magazine name from magazine details."""
        if not details:
            return ""
        # Get the first magazine's display name and normalize it
        display_name = details[0].get("display_name", "")
        normalized = self._normalize_label(display_name)
        # Check against known magazines
        for magazine_key in MAGAZINE_GROUPING_MAP.keys():
            if magazine_key in normalized:
                return magazine_key
        return ""

    def _collect_magazine_details(self, df: pd.DataFrame) -> tuple[list[dict], list[str]]:
        warnings: list[str] = []
        if df.empty:
            return [], warnings

        header_row = [None if pd.isna(col) else str(col) for col in df.columns]
        try:
            magazine_col = _find_magazine_column(header_row)
        except ValueError:
            warnings.append("Unable to determine magazine column for summary display.")
            return [], warnings

        question_set_col = None
        try:
            question_set_col = _find_question_set_column(header_row)
        except ValueError:
            warnings.append("Unable to determine question set column; question sets will not be listed.")

        coverage: dict[str, dict[str, object]] = {}
        magazine_series = df.iloc[:, magazine_col - 1]
        question_series = (
            df.iloc[:, question_set_col - 1] if question_set_col is not None else repeat(None, len(df))
        )
        for magazine_value, question_value in zip(magazine_series, question_series):
            if pd.isna(magazine_value):
                continue
            text = str(magazine_value).strip()
            if not text:
                continue
            display_parts = [part.strip() for part in text.split("|", 1)]
            display_name = display_parts[0] or "Unknown"
            display_edition = display_parts[1] if len(display_parts) > 1 else ""
            normalized = normalize_magazine_edition(text)
            norm_parts = normalized.split("|", 1)
            norm_name = norm_parts[0]
            norm_edition = norm_parts[1] if len(norm_parts) > 1 else ""
            entry = coverage.setdefault(
                norm_name,
                {
                    "display_name": display_name,
                    "editions": {},
                    "normalized_editions": set(),
                },
            )
            edition_label = display_edition or "(unspecified)"
            edition_entry = entry["editions"].setdefault(
                norm_edition,
                {
                    "display": edition_label,
                    "question_sets": set(),
                },
            )
            if question_set_col is not None and not pd.isna(question_value):
                q_text = str(question_value).strip()
                if q_text:
                    edition_entry["question_sets"].add(q_text)
            entry["normalized_editions"].add(norm_edition)

        if not coverage:
            return [], warnings

        details: list[dict] = []
        for norm_name in sorted(coverage.keys(), key=lambda key: str(coverage[key]["display_name"]).lower()):
            data = coverage[norm_name]
            edition_items = []
            for norm_edition, edition_data in sorted(
                data["editions"].items(),
                key=lambda item: self._edition_sort_key(item[0], item[1]["display"]),
            ):
                question_sets = sorted(edition_data["question_sets"], key=lambda value: value.lower())
                edition_items.append(
                    {
                        "display": edition_data["display"],
                        "normalized": norm_edition,
                        "question_sets": question_sets,
                    }
                )
            missing_ranges = self._compute_missing_ranges(data["normalized_editions"])
            details.append(
                {
                    "display_name": data["display_name"],
                    "missing_ranges": missing_ranges,
                    "editions": edition_items,
                }
            )
        return details, warnings

    def _collect_question_analysis_data(
        self, df: pd.DataFrame
    ) -> tuple[dict[str, list[dict]], list[str], int | None, list[str]]:
        warnings: list[str] = []
        if df.empty:
            return {}, warnings, None, []

        header_row = [None if pd.isna(col) else str(col) for col in df.columns]
        try:
            question_set_col = _find_high_level_chapter_column(header_row)
            qno_col = _find_qno_column(header_row)
            page_col = _find_page_column(header_row)
        except ValueError as exc:
            warnings.append(f"Question analysis unavailable: {exc}")
            return {}, warnings, None, []

        question_text_col = None
        try:
            question_text_col = _find_question_text_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        question_set_name_col = None
        try:
            question_set_name_col = _find_question_set_name_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        magazine_col = None
        try:
            magazine_col = _find_magazine_column(header_row)
        except ValueError as exc:
            warnings.append(str(exc))

        def normalize(value) -> str:
            if pd.isna(value):
                return ""
            text = str(value).strip()
            return text

        chapters: dict[str, list[dict]] = {}
        raw_inputs: set[str] = set()
        for row_number, row in enumerate(df.itertuples(index=False, name=None), start=2):
            values = list(row)
            raw_chapter_name = normalize(values[question_set_col - 1])
            if not raw_chapter_name:
                continue
            raw_inputs.add(raw_chapter_name)
            chapter_name = self._match_chapter_group(raw_chapter_name)
            qno_value = normalize(values[qno_col - 1])
            page_value = normalize(values[page_col - 1])
            question_text = normalize(values[question_text_col - 1]) if question_text_col else ""
            question_set_name = normalize(values[question_set_name_col - 1]) if question_set_name_col else raw_chapter_name
            magazine_value = normalize(values[magazine_col - 1]) if magazine_col else ""

            chapters.setdefault(chapter_name, []).append(
                {
                    "group": chapter_name,
                    "question_set": raw_chapter_name,
                    "question_set_name": question_set_name,
                    "qno": qno_value,
                    "page": page_value,
                    "magazine": magazine_value,
                    "text": question_text,
                    "row_number": row_number,
                }
            )

        return chapters, warnings, question_set_col, sorted(raw_inputs)

    def _normalize_label(self, label: str) -> str:
        return re.sub(r"\s+", " ", label.strip().lower())
    
    def _extract_group_key(self, question_set_name: str) -> str:
        """Extract a group key from question set name for similarity grouping.
        
        Methodology:
        1. Normalize the name (lowercase, remove extra spaces)
        2. Extract significant tokens (ignore common words, years, numbers at end)
        3. Return first 2-3 significant words as group key
        
        Examples:
        - 'JEE Main 2023 Paper 1' -> 'jee main'
        - 'NEET-2024' -> 'neet'
        - 'Physics Olympiad 2023' -> 'physics olympiad'
        """
        if not question_set_name:
            return "ungrouped"
        
        # Normalize
        normalized = self._normalize_label(question_set_name)
        
        # Split into tokens
        tokens = normalized.split()
        
        # Remove common suffix patterns (years, paper numbers, etc.)
        significant_tokens = []
        for token in tokens:
            # Skip years (4 digits), paper numbers, common words
            if re.match(r'^(\d{4}|\d+|paper|set|part|section|test|exam)$', token):
                continue
            significant_tokens.append(token)
            # Take first 2-3 significant words
            if len(significant_tokens) >= 3:
                break
        
        if not significant_tokens:
            # If all tokens were filtered, use first 2 tokens
            return ' '.join(tokens[:2]) if len(tokens) >= 2 else normalized
        
        return ' '.join(significant_tokens)

    def _auto_assign_chapters(self, chapters: list[str]) -> None:
        changed = False
        existing = {
            self._normalize_label(ch)
            for values in self.chapter_groups.values()
            for ch in values
        }
        for chapter in chapters:
            normalized = self._normalize_label(chapter)
            if not normalized or normalized in existing:
                continue
            target = self._match_chapter_group(chapter)
            self.chapter_groups.setdefault(target, []).append(chapter)
            existing.add(normalized)
            changed = True
        if changed:
            self._save_chapter_grouping()
            self._refresh_grouping_ui()

    def _match_chapter_group(self, chapter: str) -> str:
        normalized = self._normalize_label(chapter)
        if not normalized:
            return "Others"
        direct = self.chapter_lookup.get(normalized)
        if direct:
            return direct
        for norm_value, group in self.chapter_lookup.items():
            if norm_value in normalized or normalized in norm_value:
                return group
        for group in self.canonical_chapters:
            norm_group = self._normalize_label(group)
            if normalized == norm_group:
                return group
        for group in self.canonical_chapters:
            norm_group = self._normalize_label(group)
            if norm_group in normalized or normalized in norm_group:
                return group
        return "Others"

    def _compute_row_count_from_df(self, df: pd.DataFrame) -> int:
        def row_has_value(row) -> bool:
            for value in row:
                if isinstance(value, str):
                    if value.strip():
                        return True
                    continue
                if pd.isna(value):
                    continue
                return True
            return False

        for idx in range(len(df) - 1, -1, -1):
            if row_has_value(df.iloc[idx]):
                return idx + 2  # DataFrame row index is zero-based, Excel rows start at 2 after header.
        return 1

    def _compute_missing_ranges(self, normalized_editions: set[str]) -> list[str]:
        monthly_tokens = sorted(
            {
                token
                for token in normalized_editions
                if re.fullmatch(r"\d{4}-\d{2}", token)
            }
        )
        if len(monthly_tokens) < 2:
            return []

        months = [dt.date(int(token[:4]), int(token[5:7]), 1) for token in monthly_tokens]
        present_keys = set(monthly_tokens)
        missing_ranges: list[str] = []
        current = months[0]
        last = months[-1]
        missing_start: dt.date | None = None

        while current <= last:
            key = current.strftime("%Y-%m")
            if key not in present_keys:
                if missing_start is None:
                    missing_start = current
            else:
                if missing_start is not None:
                    end = self._previous_month(current)
                    missing_ranges.append(self._format_range(missing_start, end))
                    missing_start = None
            current = self._add_month(current)

        if missing_start is not None:
            missing_ranges.append(self._format_range(missing_start, last))
        return missing_ranges

    def _add_month(self, date_value: dt.date) -> dt.date:
        if date_value.month == 12:
            return dt.date(date_value.year + 1, 1, 1)
        return dt.date(date_value.year, date_value.month + 1, 1)

    def _previous_month(self, date_value: dt.date) -> dt.date:
        if date_value.month == 1:
            return dt.date(date_value.year - 1, 12, 1)
        return dt.date(date_value.year, date_value.month - 1, 1)

    def _format_range(self, start: dt.date, end: dt.date) -> str:
        if start == end:
            return start.strftime("%b %Y")
        return f"{start.strftime('%b %Y')} - {end.strftime('%b %Y')}"

    def _parse_normalized_month(self, normalized: str) -> dt.date | None:
        if normalized and re.fullmatch(r"\d{4}-\d{2}", normalized):
            return dt.date(int(normalized[:4]), int(normalized[5:7]), 1)
        return None

    def _edition_sort_key(self, normalized: str, display_label: str) -> tuple:
        parsed = self._parse_normalized_month(normalized)
        if parsed:
            return (0, -parsed.toordinal(), display_label.lower())
        return (1, display_label.lower(), "")

    def _populate_magazine_tree(self, details: list[dict]) -> None:
        if not hasattr(self, "mag_tree"):
            return
        self.mag_tree.clear()
        
        # Calculate statistics
        total_editions = 0
        total_sets = 0
        
        if not details:
            if hasattr(self, "mag_total_editions_label"):
                self.mag_total_editions_label.setText("Total Editions: 0")
            if hasattr(self, "mag_total_sets_label"):
                self.mag_total_sets_label.setText("Question Sets: 0")
            return

        for entry in details:
            total_editions += len(entry.get("editions", []))
            for edition in entry.get("editions", []):
                total_sets += len(edition.get("question_sets", []))
            missing_text = ", ".join(entry["missing_ranges"]) if entry["missing_ranges"] else "None"
            parent = QTreeWidgetItem([entry["display_name"], "", missing_text])
            parent.setData(0, Qt.UserRole, {"type": "magazine", "display_name": entry["display_name"]})
            self.mag_tree.addTopLevelItem(parent)
            for edition in entry["editions"]:
                edition_label = edition["display"] or "(unspecified)"
                question_sets = edition["question_sets"]
                edition_value = (
                    f"{edition_label} ({len(question_sets)} set(s))" if question_sets else edition_label
                )
                child = QTreeWidgetItem(["", edition_value, ""])
                data = {
                    "type": "edition",
                    "display_name": entry["display_name"],
                    "edition_label": edition_label,
                    "question_sets": question_sets,
                }
                child.setData(0, Qt.UserRole, data)
                
                # Color code editions based on question set count
                if len(question_sets) == 0:
                    child.setForeground(2, QColor("#94a3b8"))  # Gray for no sets
                elif len(question_sets) < 5:
                    # Yellow background for editions with less than 5 sets
                    child.setBackground(0, QColor("#fef9c3"))
                    child.setBackground(1, QColor("#fef9c3"))
                    child.setBackground(2, QColor("#fef9c3"))
                elif len(question_sets) >= 5:
                    child.setForeground(1, QColor("#10b981"))  # Green for many sets
                
                parent.addChild(child)
        
        self.mag_tree.expandAll()
        
        # Update summary statistics
        if hasattr(self, "mag_total_editions_label"):
            self.mag_total_editions_label.setText(f"Total Editions: {total_editions}")
        if hasattr(self, "mag_total_sets_label"):
            self.mag_total_sets_label.setText(f"Question Sets: {total_sets}")

    def _populate_question_sets(self, question_sets: list[str] | None) -> None:
        if not hasattr(self, "question_list"):
            return
        self.question_list.clear()
        if not question_sets:
            return
        for idx, name in enumerate(question_sets, 1):
            item = QListWidgetItem(f"📝 {name}")
            item.setData(Qt.UserRole, name)
            self.question_list.addItem(item)

    def _populate_chapter_list(self, chapters: dict[str, list[dict]]) -> None:
        if not hasattr(self, "chapter_table"):
            return
        self.chapter_questions = chapters or {}
        self.chapter_table.setRowCount(0)
        if hasattr(self, "question_tree"):
            self.question_tree.clear()
        self.question_text_view.clear()
        if not self.chapter_questions:
            return
        sorted_chapters = sorted(
            self.chapter_questions.items(),
            key=lambda kv: (-len(kv[1]), kv[0].lower()),
        )
        self.chapter_table.setRowCount(len(sorted_chapters))
        for row, (chapter, questions) in enumerate(sorted_chapters):
            name_item = QTableWidgetItem(chapter)
            name_item.setData(Qt.UserRole, chapter)
            count_item = QTableWidgetItem(str(len(questions)))
            count_item.setTextAlignment(Qt.AlignCenter)
            self.chapter_table.setItem(row, 0, name_item)
            self.chapter_table.setItem(row, 1, count_item)
        self.chapter_table.resizeColumnToContents(0)
        self.chapter_table.resizeColumnToContents(1)
        if self.chapter_table.rowCount() > 0:
            self.chapter_table.selectRow(0)

    def _populate_question_table(self, questions: list[dict]) -> None:
        if not hasattr(self, "question_tree"):
            return
        self.all_questions = questions or []
        self._apply_question_search()

    def _apply_question_search(self, preserve_scroll: bool = False) -> None:
        """Apply normalized search terms and update the question tree."""
        # Save scroll position if requested
        scroll_value = None
        if preserve_scroll and hasattr(self, "question_tree"):
            scrollbar = self.question_tree.verticalScrollBar()
            if scrollbar:
                scroll_value = scrollbar.value()
        
        filtered = self.all_questions
        
        # Apply Question Set search (normalized)
        if self.question_set_search_term:
            normalized_search = self._normalize_label(self.question_set_search_term)
            filtered = [
                q for q in filtered
                if normalized_search in self._normalize_label(q.get("question_set_name", ""))
            ]
        
        # Apply Magazine search (normalized)
        if self.magazine_search_term:
            normalized_search = self._normalize_label(self.magazine_search_term)
            filtered = [
                q for q in filtered
                if normalized_search in self._normalize_label(q.get("magazine", ""))
            ]
        
        # Update tree with grouping
        self.current_questions = filtered
        self.question_tree.clear()
        self.question_text_view.clear()
        
        if not filtered:
            return
        
        # Group questions by similar question set names
        groups = {}
        for question in filtered:
            question_set_name = question.get("question_set_name", "Unknown")
            group_key = self._extract_group_key(question_set_name)
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append(question)
        
        # Apply tag filtering (multiple tags)
        if self.selected_tag_filters:
            filtered_groups = {}
            for group_key, group_questions in groups.items():
                tags = self.group_tags.get(group_key, [])
                # Check if any of the selected filter tags match any of the group's tags
                if any(filter_tag in tags for filter_tag in self.selected_tag_filters):
                    filtered_groups[group_key] = group_questions
            groups = filtered_groups
        
        # Sort groups by number of questions (descending) then by name
        sorted_groups = sorted(groups.items(), key=lambda x: (-len(x[1]), x[0]))
        
        # Build tree structure
        for group_key, group_questions in sorted_groups:
            # Create group node
            group_item = QTreeWidgetItem(self.question_tree)
            
            # Store group key for tagging
            group_item.setData(0, Qt.UserRole + 1, group_key)
            
            # Format group label with question count
            display_name = group_key.title()
            label_text = f"{display_name} ({len(group_questions)} questions)"
            
            # Always create a widget layout for consistent alignment
            tag_widget = QWidget()
            tag_layout = QHBoxLayout(tag_widget)
            tag_layout.setContentsMargins(0, 4, 0, 4)  # Add vertical padding for spacing
            tag_layout.setSpacing(8)  # Space between badges
            
            # Add the label with fixed minimum width for alignment
            label = QLabel(label_text)
            label.setStyleSheet("color: #1e40af; font-weight: bold;")
            label.setMinimumWidth(300)  # Fixed width to align all tags
            tag_layout.addWidget(label)
            
            # Add tag badges if tags exist for this group
            tags = self.group_tags.get(group_key, [])
            if tags:
                for tag in tags:
                    color = self._get_or_assign_tag_color(tag)
                    badge = TagBadge(tag, color)
                    badge.clicked.connect(self._on_tag_badge_clicked)
                    tag_layout.addWidget(badge)
            
            tag_layout.addStretch()
            
            # Set the widget in the tree
            self.question_tree.setItemWidget(group_item, 0, tag_widget)
            
            group_item.setExpanded(False)  # Collapsed by default
            
            # Add questions as children
            for question in group_questions:
                child = QTreeWidgetItem(group_item)
                child.setText(0, question.get("qno", ""))
                child.setText(1, question.get("page", ""))
                child.setText(2, question.get("question_set_name", ""))
                child.setText(3, question.get("magazine", ""))
                
                # Highlight magazine column with light blue background
                child.setBackground(3, QColor("#dbeafe"))
                
                # Store full question data for selection
                child.setData(0, Qt.UserRole, question)
        
        self.question_tree.resizeColumnToContents(0)
        self.question_tree.resizeColumnToContents(1)
        self.question_tree.resizeColumnToContents(2)
        
        # Restore scroll position if it was saved
        if scroll_value is not None:
            scrollbar = self.question_tree.verticalScrollBar()
            if scrollbar:
                scrollbar.setValue(scroll_value)

    def on_question_set_search_changed(self, text: str) -> None:
        """Handle Question Set search box text change."""
        self.question_set_search_term = text.strip()
        self._apply_question_search()

    def on_magazine_search_changed(self, text: str) -> None:
        """Handle Magazine search box text change."""
        self.magazine_search_term = text.strip()
        self._apply_question_search()

    def clear_question_search(self) -> None:
        """Clear all search terms."""
        self.question_set_search_term = ""
        self.magazine_search_term = ""
        self.selected_tag_filters = []
        if hasattr(self, "question_set_search"):
            self.question_set_search.clear()
        if hasattr(self, "tag_filter_label"):
            self.tag_filter_label.setText("None")
        if hasattr(self, "magazine_search"):
            self.magazine_search.clear()
        self._apply_question_search()

    def _show_tag_filter_dialog(self) -> None:
        """Show dialog to select multiple tags for filtering."""
        # Get all existing tags across all groups
        all_tags = sorted(set(tag for tags in self.group_tags.values() for tag in tags))
        
        if not all_tags:
            QMessageBox.information(self, "No Tags", "No tags have been created yet. Assign tags to groups first.")
            return
        
        # Show multi-select dialog
        dialog = MultiSelectTagDialog(
            all_tags, 
            self.selected_tag_filters, 
            "Filter by Tags",
            self.tag_colors,
            self.available_tag_colors,
            self
        )
        
        if dialog.exec() == QDialog.Accepted:
            self.selected_tag_filters = dialog.get_selected_tags()
            # Update label
            if self.selected_tag_filters:
                self.tag_filter_label.setText(", ".join(self.selected_tag_filters))
            else:
                self.tag_filter_label.setText("None")
            self._apply_question_search()

    def on_magazine_search_changed(self, text: str) -> None:
        """Handle Magazine search box text change."""
        self.magazine_search_term = text.strip()
        self._apply_question_search()

    def clear_question_search(self) -> None:
        """Clear all search terms."""
        self.question_set_search_term = ""
        self.magazine_search_term = ""
        self.selected_tag_filters = []
        if hasattr(self, "question_set_search"):
            self.question_set_search.clear()
        if hasattr(self, "tag_filter_label"):
            self.tag_filter_label.setText("None")
        if hasattr(self, "magazine_search"):
            self.magazine_search.clear()
        self._apply_question_search()

    def on_mag_tree_search_changed(self, text: str) -> None:
        """Filter magazine tree based on search text."""
        search_term = text.strip().lower()
        
        if not hasattr(self, "mag_tree"):
            return
        
        # Show all if search is empty
        if not search_term:
            for i in range(self.mag_tree.topLevelItemCount()):
                parent = self.mag_tree.topLevelItem(i)
                parent.setHidden(False)
                for j in range(parent.childCount()):
                    parent.child(j).setHidden(False)
            self.mag_tree.expandAll()
            return
        
        # Filter tree items
        for i in range(self.mag_tree.topLevelItemCount()):
            parent = self.mag_tree.topLevelItem(i)
            parent_text = parent.text(0).lower()
            parent_matches = search_term in parent_text
            
            visible_children = 0
            for j in range(parent.childCount()):
                child = parent.child(j)
                child_text = f"{child.text(0)} {child.text(1)}".lower()
                child_matches = search_term in child_text
                
                child.setHidden(not (parent_matches or child_matches))
                if parent_matches or child_matches:
                    visible_children += 1
            
            parent.setHidden(visible_children == 0)
            if visible_children > 0:
                parent.setExpanded(True)
    
    def on_magazine_select(self) -> None:
        selected = self.mag_tree.selectedItems()
        if not selected:
            self.question_label.setText("Select an edition to view question sets")
            self._populate_question_sets([])
            return
        item = selected[0]
        data = item.data(0, Qt.UserRole)
        if not isinstance(data, dict) or data.get("type") != "edition":
            self.question_label.setText("Select an edition to view question sets")
            self._populate_question_sets([])
            return

        question_sets = data.get("question_sets", [])
        label = f"{data.get('display_name', 'Magazine')} - {data.get('edition_label', 'Edition')}"
        if question_sets:
            self.question_label.setText(f"✓ {label} • {len(question_sets)} question set(s)")
            self._populate_question_sets(question_sets)
        else:
            self.question_label.setText(f"⚠ {label} • No question sets found")
            self._populate_question_sets([])

    def on_group_selected(self) -> None:
        if not hasattr(self, "group_list"):
            return
        item = self.group_list.currentItem()
        self.group_chapter_list.clear()
        if not item:
            return
        group = item.data(Qt.UserRole) or item.text()
        for chapter in sorted(self.chapter_groups.get(group, []), key=lambda value: value.lower()):
            chapter_item = QListWidgetItem(chapter)
            chapter_item.setData(Qt.UserRole, chapter)
            self.group_chapter_list.addItem(chapter_item)

    def move_selected_chapter(self) -> None:
        group_item = getattr(self, "group_list", None)
        chapter_list = getattr(self, "group_chapter_list", None)
        target_combo = getattr(self, "move_target_combo", None)
        if (
            not group_item
            or not chapter_list
            or not target_combo
            or not group_item.currentItem()
            or not chapter_list.currentItem()
        ):
            return
        current_group = group_item.currentItem().data(Qt.UserRole)
        chapter_name = chapter_list.currentItem().data(Qt.UserRole)
        target_group = target_combo.currentText()
        if not chapter_name or not target_group or current_group == target_group:
            return
        self.move_chapter_to_group(chapter_name, target_group)

    def move_chapter_to_group(self, chapter_name: str, target_group: str, stay_on_group: str | None = None) -> None:
        if not chapter_name or not target_group:
            return
        for group, values in self.chapter_groups.items():
            if chapter_name in values:
                values.remove(chapter_name)
                break
        self.chapter_groups.setdefault(target_group, [])
        if chapter_name not in self.chapter_groups[target_group]:
            self.chapter_groups[target_group].append(chapter_name)
        self._save_chapter_grouping()
        self._refresh_grouping_ui()
        self._select_group_in_ui(stay_on_group or target_group)
        self.on_group_selected()

    def reassign_question(self, question: dict, target_group: str) -> None:
        if not question or not target_group:
            return
        if self.current_workbook_path is None or self.high_level_column_index is None:
            QMessageBox.warning(self, "Unavailable", "Workbook must be loaded before regrouping questions.")
            return
        old_group = question.get("group")
        if old_group == target_group:
            return
        row_number = question.get("row_number")
        if not isinstance(row_number, int):
            QMessageBox.warning(self, "Unavailable", "Question row information is missing.")
            return
        qno = question.get("qno", "")
        prompt = f"Move question '{qno}' to '{target_group}'?"
        if QMessageBox.question(self, "Confirm Reassignment", prompt) != QMessageBox.Yes:
            return
        try:
            workbook = load_workbook(self.current_workbook_path)
            sheet = workbook[workbook.sheetnames[0]]
            sheet.cell(row=row_number, column=self.high_level_column_index, value=target_group)
            workbook.save(self.current_workbook_path)
        except Exception as exc:
            QMessageBox.critical(self, "Update Failed", f"Unable to update workbook: {exc}")
            return
        self.log(f"Question '{qno}' moved to '{target_group}'. Reloading data...")
        self.update_row_count()

    def on_chapter_selected(self) -> None:
        if not hasattr(self, "chapter_table"):
            return
        row = self.chapter_table.currentRow()
        if row < 0:
            self._populate_question_table([])
            return
        item = self.chapter_table.item(row, 0)
        if not item:
            self._populate_question_table([])
            return
        chapter_key = item.data(Qt.UserRole) or item.text()
        questions = self.chapter_questions.get(chapter_key, [])
        self._populate_question_table(questions)

    def on_question_selected(self) -> None:
        if not hasattr(self, "question_tree"):
            return
        selected_items = self.question_tree.selectedItems()
        if not selected_items:
            self.question_text_view.clear()
            return
        
        # Get the first selected item
        item = selected_items[0]
        
        # Skip if it's a group header (has children)
        if item.childCount() > 0:
            self.question_text_view.clear()
            return
        
        # Get question data from UserRole
        question = item.data(0, Qt.UserRole)
        if question:
            html = (
                f"<div style='background-color: #0f172a; color: #e2e8f0; font-family: Arial, sans-serif; padding: 10px;'>"
                f"<span style='color: #60a5fa; font-weight: bold;'>Qno:</span> <span style='color: #cbd5e1;'>{question.get('qno','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Page No:</span> <span style='color: #cbd5e1;'>{question.get('page','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Question Set:</span> <span style='color: #cbd5e1;'>{question.get('question_set_name','')}</span> &nbsp;&nbsp;"
                f"<span style='color: #60a5fa; font-weight: bold;'>Magazine Edition:</span> <span style='color: #cbd5e1;'>{question.get('magazine','')}</span><br/>"
                f"<hr style='border: none; border-top: 1px solid #334155; margin: 12px 0;'/>"
                f"<div style='color: #e2e8f0; line-height: 1.7; font-size: 14px;'>{question.get('text','').replace(chr(10), '<br/>')}</div>"
                f"</div>"
            )
            self.question_text_view.setHtml(html)

    def _on_tag_badge_clicked(self, tag: str) -> None:
        """Handle tag badge click - add to filter list."""
        if tag not in self.selected_tag_filters:
            self.selected_tag_filters.append(tag)
            if hasattr(self, "tag_filter_label"):
                self.tag_filter_label.setText(", ".join(self.selected_tag_filters))
            self._apply_question_search()

    def _show_group_context_menu(self, position) -> None:
        """Show context menu for group items."""
        item = self.question_tree.itemAt(position)
        if not item:
            return
        
        # Only show menu for group items (items with children)
        if item.childCount() == 0:
            return
        
        group_key = item.data(0, Qt.UserRole + 1)
        if not group_key:
            return
        
        menu = QMenu(self.question_tree)
        
        # Add tag action
        add_tag_action = menu.addAction("Add Tag")
        
        # Remove tag actions
        existing_tags = self.group_tags.get(group_key, [])
        if existing_tags:
            remove_menu = menu.addMenu("Remove Tag")
            for tag in existing_tags:
                remove_menu.addAction(tag)
        
        action = menu.exec(self.question_tree.viewport().mapToGlobal(position))
        
        if action:
            if action.text() == "Add Tag":
                self._assign_tag_to_group(group_key)
            else:
                # Remove tag
                self._remove_tag_from_group(group_key, action.text())

    def _assign_tag_to_group(self, group_key: str) -> None:
        """Show dialog to assign multiple tags to a group."""
        # Get all existing tags across all groups
        all_tags = sorted(set(tag for tags in self.group_tags.values() for tag in tags))
        
        # Get currently assigned tags for this group
        current_tags = self.group_tags.get(group_key, [])
        
        # Show multi-select dialog
        dialog = MultiSelectTagDialog(
            all_tags, 
            current_tags, 
            f"Add Tags to '{group_key.title()}'",
            self.tag_colors,
            self.available_tag_colors,
            self
        )
        
        if dialog.exec() == QDialog.Accepted:
            selected_tags = dialog.get_selected_tags()
            if selected_tags:
                self.group_tags[group_key] = selected_tags
                self._save_group_tags()
                self._apply_question_search(preserve_scroll=True)  # Refresh tree to show new tags

    def _remove_tag_from_group(self, group_key: str, tag: str) -> None:
        """Remove a tag from a group."""
        if group_key in self.group_tags and tag in self.group_tags[group_key]:
            self.group_tags[group_key].remove(tag)
            if not self.group_tags[group_key]:
                # Remove empty tag list
                del self.group_tags[group_key]
            self._save_group_tags()
            self._apply_question_search(preserve_scroll=True)  # Refresh tree to update display
        else:
            self.question_text_view.clear()

    def _load_saved_question_lists(self) -> None:
        """Load all saved question lists from QuestionList folder."""
        if not hasattr(self, "saved_lists_widget"):
            return
        
        self.saved_lists_widget.clear()
        self.question_lists.clear()
        
        if not QUESTION_LIST_DIR.exists():
            return
        
        for json_file in sorted(QUESTION_LIST_DIR.glob("*.json")):
            try:
                data = json.loads(json_file.read_text(encoding="utf-8"))
                list_name = data.get("name", json_file.stem)
                questions = data.get("questions", [])
                self.question_lists[list_name] = questions
                
                item = QListWidgetItem(f"{list_name} ({len(questions)})")
                item.setData(Qt.UserRole, list_name)
                self.saved_lists_widget.addItem(item)
            except Exception as exc:
                self.log(f"Error loading list {json_file.name}: {exc}")
    
    def _save_question_list(self, list_name: str) -> None:
        """Save a question list to file."""
        if list_name not in self.question_lists:
            return
        
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in list_name)
        file_path = QUESTION_LIST_DIR / f"{safe_name}.json"
        
        data = {
            "name": list_name,
            "questions": self.question_lists[list_name],
        }
        
        try:
            file_path.write_text(json.dumps(data, indent=2), encoding="utf-8")
            self.log(f"Saved question list: {list_name}")
        except Exception as exc:
            QMessageBox.critical(self, "Save Error", f"Failed to save list: {exc}")
    
    def create_new_question_list(self) -> None:
        """Create a new question list."""
        from PySide6.QtWidgets import QInputDialog
        
        list_name, ok = QInputDialog.getText(self, "New Question List", "Enter list name:")
        if not ok or not list_name.strip():
            return
        
        list_name = list_name.strip()
        if list_name in self.question_lists:
            QMessageBox.warning(self, "Duplicate Name", f"A list named '{list_name}' already exists.")
            return
        
        self.question_lists[list_name] = []
        self._save_question_list(list_name)
        self._load_saved_question_lists()
        self.log(f"Created new question list: {list_name}")
    
    def rename_question_list(self) -> None:
        """Rename selected question list."""
        from PySide6.QtWidgets import QInputDialog
        
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            QMessageBox.information(self, "No Selection", "Please select a list to rename.")
            return
        
        old_name = current_item.data(Qt.UserRole)
        new_name, ok = QInputDialog.getText(self, "Rename List", "Enter new name:", text=old_name)
        if not ok or not new_name.strip():
            return
        
        new_name = new_name.strip()
        if new_name == old_name:
            return
        
        if new_name in self.question_lists:
            QMessageBox.warning(self, "Duplicate Name", f"A list named '{new_name}' already exists.")
            return
        
        # Delete old file
        safe_old_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in old_name)
        old_file = QUESTION_LIST_DIR / f"{safe_old_name}.json"
        old_file.unlink(missing_ok=True)
        
        # Update data structure
        self.question_lists[new_name] = self.question_lists.pop(old_name)
        self._save_question_list(new_name)
        self._load_saved_question_lists()
        self.log(f"Renamed list from '{old_name}' to '{new_name}'")
    
    def delete_question_list(self) -> None:
        """Delete selected question list."""
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            QMessageBox.information(self, "No Selection", "Please select a list to delete.")
            return
        
        list_name = current_item.data(Qt.UserRole)
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete the list '{list_name}'?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        
        # Delete file
        safe_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in list_name)
        file_path = QUESTION_LIST_DIR / f"{safe_name}.json"
        file_path.unlink(missing_ok=True)
        
        # Update data structure
        del self.question_lists[list_name]
        self._load_saved_question_lists()
        self.list_question_table.setRowCount(0)
        self.list_name_label.setText("Select a list to view questions")
        self.log(f"Deleted question list: {list_name}")
    
    def add_selected_to_list(self) -> None:
        """Add selected questions from question tree to a list."""
        from PySide6.QtWidgets import QInputDialog
        
        if not self.question_lists:
            reply = QMessageBox.question(
                self,
                "No Lists",
                "No question lists exist. Create a new list?",
                QMessageBox.Yes | QMessageBox.No,
            )
            if reply == QMessageBox.Yes:
                self.create_new_question_list()
            return
        
        selected_items = self.question_tree.selectedItems()
        if not selected_items:
            QMessageBox.information(self, "No Selection", "Please select questions to add.")
            return
        
        # Filter out group headers, only get leaf items (questions)
        selected_questions = []
        for item in selected_items:
            if item.childCount() == 0:  # Leaf node
                question_data = item.data(0, Qt.UserRole)
                if question_data:
                    selected_questions.append(question_data)
        
        if not selected_questions:
            QMessageBox.information(self, "No Questions", "Please select actual questions, not group headers.")
            return
        
        # Select list
        list_names = sorted(self.question_lists.keys())
        list_name, ok = QInputDialog.getItem(
            self, "Select List", "Choose a list to add questions to:", list_names, 0, False
        )
        if not ok or not list_name:
            return
        
        # Add selected questions
        added_count = 0
        for question in selected_questions:
            # Check for duplicates based on row_number
            if not any(q.get("row_number") == question.get("row_number") for q in self.question_lists[list_name]):
                self.question_lists[list_name].append(question.copy())
                added_count += 1
        
        self._save_question_list(list_name)
        self._load_saved_question_lists()
        self.log(f"Added {added_count} question(s) to list '{list_name}'")
        QMessageBox.information(self, "Success", f"Added {added_count} question(s) to '{list_name}'.")
    
    def remove_selected_from_list(self) -> None:
        """Remove selected questions from current list."""
        if not self.current_list_name:
            return
        
        selected_rows = self.list_question_table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.information(self, "No Selection", "Please select questions to remove.")
            return
        
        # Sort in reverse to avoid index issues
        rows_to_remove = sorted([index.row() for index in selected_rows], reverse=True)
        
        for row in rows_to_remove:
            if 0 <= row < len(self.question_lists[self.current_list_name]):
                del self.question_lists[self.current_list_name][row]
        
        self._save_question_list(self.current_list_name)
        self._load_saved_question_lists()
        self._populate_list_question_table(self.current_list_name)
        self.log(f"Removed {len(rows_to_remove)} question(s) from '{self.current_list_name}'")
    
    def on_saved_list_selected(self) -> None:
        """Handle selection of a saved question list."""
        current_item = self.saved_lists_widget.currentItem()
        if not current_item:
            self.list_question_table.setRowCount(0)
            self.list_name_label.setText("Select a list to view questions")
            self.current_list_name = None
            return
        
        list_name = current_item.data(Qt.UserRole)
        self.current_list_name = list_name
        self._populate_list_question_table(list_name)
    
    def _populate_list_question_table(self, list_name: str) -> None:
        """Populate the list question table with questions from the selected list."""
        if list_name not in self.question_lists:
            return
        
        questions = self.question_lists[list_name]
        self.current_list_questions = questions
        self.list_name_label.setText(f"{list_name} ({len(questions)} questions)")
        
        self.list_question_table.setRowCount(0)
        self.list_question_text_view.clear()
        if not questions:
            return
        
        self.list_question_table.setRowCount(len(questions))
        for row, question in enumerate(questions):
            qno_item = QTableWidgetItem(str(question.get("qno", "")))
            page_item = QTableWidgetItem(str(question.get("page", "")))
            question_set_item = QTableWidgetItem(str(question.get("question_set", "")))
            magazine_item = QTableWidgetItem(str(question.get("magazine", "")))
            
            self.list_question_table.setItem(row, 0, qno_item)
            self.list_question_table.setItem(row, 1, page_item)
            self.list_question_table.setItem(row, 2, question_set_item)
            self.list_question_table.setItem(row, 3, magazine_item)
        
        self.list_question_table.resizeColumnsToContents()
    
    def on_list_question_selected(self) -> None:
        """Handle selection of a question in the list question table."""
        if not hasattr(self, "list_question_table"):
            return
        selection_model = self.list_question_table.selectionModel()
        if selection_model is None:
            self.list_question_text_view.clear()
            return
        selected_rows = selection_model.selectedRows()
        if not selected_rows:
            self.list_question_text_view.clear()
            return
        row = selected_rows[0].row()
        if 0 <= row < len(self.current_list_questions):
            question = self.current_list_questions[row]
            html_parts = [
                '<div style="background-color: #0f172a; color: #e2e8f0; padding: 12px; font-family: Arial, sans-serif;">',
                f'<p><span style="color: #60a5fa; font-weight: bold;">Question No:</span> <span style="color: #cbd5e1;">{question.get("qno", "N/A")}</span></p>',
                f'<p><span style="color: #60a5fa; font-weight: bold;">Page:</span> <span style="color: #cbd5e1;">{question.get("page", "N/A")}</span></p>',
                f'<p><span style="color: #60a5fa; font-weight: bold;">Question Set:</span> <span style="color: #cbd5e1;">{question.get("question_set", "N/A")}</span></p>',
                f'<p><span style="color: #60a5fa; font-weight: bold;">Magazine:</span> <span style="color: #cbd5e1;">{question.get("magazine", "N/A")}</span></p>',
                f'<p><span style="color: #60a5fa; font-weight: bold;">Chapter:</span> <span style="color: #cbd5e1;">{question.get("group", "N/A")}</span></p>',
                '<hr style="border: 1px solid #475569;">',
                f'<div style="color: #e2e8f0; line-height: 1.6;">{question.get("question_text", "No question text available")}</div>',
                '</div>',
            ]
            self.list_question_text_view.setHtml("".join(html_parts))
        else:
            self.list_question_text_view.clear()

    def log(self, message: str) -> None:
        timestamp = time.strftime("%H:%M:%S")
        self.log_view.append(f"[{timestamp}] {message}")
        self.log_view.moveCursor(QTextCursor.End)

    def select_input_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.input_edit.setText(folder)
            self._save_last_selection()
            self.refresh_file_list()

    def select_output_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel Workbook",
            "",
            "Excel files (*.xlsx)",
        )
        if file_path:
            self.output_edit.setText(file_path)
            self._save_last_selection()
            self.update_row_count()

    def refresh_file_list(self) -> None:
        input_path = Path(self.input_edit.text().strip())
        if not input_path.exists():
            self.log("Input folder does not exist.")
            return

        tsv_files = sorted(input_path.glob("*.tsv"))
        for tsv_file in tsv_files:
            self._ensure_row(tsv_file.name)
        self.log(f"Found {len(tsv_files)} TSV file(s) in input folder.")

    def _ensure_row(self, filename: str) -> int:
        row = self.file_rows.get(filename)
        if row is not None:
            return row
        row = self.file_table.rowCount()
        self.file_table.insertRow(row)
        self.file_table.setItem(row, 0, QTableWidgetItem(filename))
        self.file_table.setItem(row, 1, QTableWidgetItem("Pending"))
        self.file_table.setItem(row, 2, QTableWidgetItem("Awaiting processing"))
        self.file_rows[filename] = row
        return row

    def update_file_status(self, filename: str, status: str, message: str) -> None:
        row = self._ensure_row(filename)
        for col, value in enumerate((filename, status, message)):
            item = self.file_table.item(row, col)
            if item is None:
                item = QTableWidgetItem(value)
                self.file_table.setItem(row, col, item)
            else:
                item.setText(value)
        if status.lower() == "error":
            self.file_errors[filename] = message
        else:
            self.file_errors.pop(filename, None)

    def on_file_double_clicked(self, row: int, column: int) -> None:  # noqa: ARG002
        filename_item = self.file_table.item(row, 0)
        status_item = self.file_table.item(row, 1)
        message_item = self.file_table.item(row, 2)
        if not filename_item or not status_item:
            return
        if status_item.text().lower() != "error":
            return
        filename = filename_item.text()
        detail = self.file_errors.get(filename, message_item.text() if message_item else "")
        QMessageBox.critical(self, "Validation Error", f"{filename}\n\n{detail}")

    def start_watching(self) -> None:
        if self.watch_thread and self.watch_thread.is_alive():
            QMessageBox.information(self, "Watcher", "Watcher is already running.")
            return

        input_path = Path(self.input_edit.text().strip())
        workbook_path = Path(self.output_edit.text().strip())

        if not input_path.is_dir():
            QMessageBox.critical(self, "Invalid Input", "Please select a valid input folder.")
            return
        if not workbook_path.is_file():
            QMessageBox.critical(self, "Invalid Workbook", "Please select an existing Excel workbook.")
            return

        self.stop_event.clear()
        self.watch_thread = threading.Thread(
            target=self._watch_loop, args=(input_path, workbook_path), daemon=True
        )
        self.watch_thread.start()
        self.start_button.setEnabled(False)
        self.log("Started watching for TSV files.")

    def stop_watching(self) -> None:
        self.stop_event.set()
        if self.watch_thread and self.watch_thread.is_alive():
            self.watch_thread.join(timeout=0.1)
        self.start_button.setEnabled(True)
        self.log("Stopped watching.")

    def _watch_loop(self, input_dir: Path, workbook_path: Path) -> None:
        poll_interval = 3.0
        while not self.stop_event.is_set():
            tsv_files = sorted(input_dir.glob("*.tsv"))
            if not tsv_files:
                time.sleep(poll_interval)
                continue

            for tsv_file in tsv_files:
                if self.stop_event.is_set():
                    break
                self.event_queue.put(("status", tsv_file.name, "Processing", "Validating..."))
                try:
                    result_message = process_tsv(tsv_file, workbook_path)
                except Exception as exc:
                    self.event_queue.put(("status", tsv_file.name, "Error", str(exc)))
                    self.event_queue.put(("log", f"Error processing {tsv_file.name}: {exc}"))
                else:
                    self.event_queue.put(("status", tsv_file.name, "Completed", result_message))
                    self.event_queue.put(("log", f"Processed {tsv_file.name}: {result_message}"))
                    self.event_queue.put(("rowcount",))

            time.sleep(poll_interval)

    def _process_queue(self) -> None:
        while True:
            try:
                event = self.event_queue.get_nowait()
            except queue.Empty:
                break

            event_type = event[0]
            if event_type == "status":
                _, filename, status, message = event
                self.update_file_status(filename, status, message)
            elif event_type == "log":
                _, message = event
                self.log(message)
            elif event_type == "metrics":
                _, req_id, row_count, details, warnings, chapter_data, question_col, raw_chapter_inputs, detected_magazine = event
                if req_id != self.metrics_request_id:
                    continue
                self.high_level_column_index = question_col
                
                # Magazine grouping already loaded in worker thread, just log it
                if detected_magazine:
                    self.log(f"Loaded chapter grouping for: {detected_magazine or 'default (Physics)'}")
                
                self.row_count_label.setText(f"Total rows: {row_count}")
                total_editions = sum(len(entry["editions"]) for entry in details)
                self._set_magazine_summary(
                    f"Magazines loaded: {len(details)}",
                    f"Tracked editions: {total_editions}",
                )
                self._populate_magazine_tree(details)
                self._populate_question_sets([])
                missing_qset_warning = next(
                    (msg for msg in warnings if "question set" in msg.lower()), None
                )
                if missing_qset_warning:
                    label_message = missing_qset_warning
                elif details:
                    label_message = "Select an edition to view question sets."
                else:
                    label_message = warnings[0] if warnings else "No magazine editions found."
                self.question_label.setText(label_message)
                self._auto_assign_chapters(raw_chapter_inputs)
                self._populate_chapter_list(chapter_data)
                self._refresh_grouping_ui()
                for warning in warnings:
                    self.log(warning)
            elif event_type == "metrics_error":
                _, req_id, error_message = event
                if req_id != self.metrics_request_id:
                    continue
                self.high_level_column_index = None
                self.row_count_label.setText("Total rows: Error")
                self._set_magazine_summary("Magazines: Error", "Missing ranges: Error")
                self._populate_magazine_tree([])
                self._populate_question_sets([])
                self._populate_chapter_list({})
                self._refresh_grouping_ui()
                self.question_label.setText("Unable to load editions.")
                self.log(f"Unable to read workbook rows: {error_message}")
            elif event_type == "rowcount":
                self.update_row_count()
        # Timer will trigger this method again; no manual reschedule needed.

    def closeEvent(self, event) -> None:
        self.stop_watching()
        super().closeEvent(event)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TSVWatcherWindow()
    window.show()
    sys.exit(app.exec())
